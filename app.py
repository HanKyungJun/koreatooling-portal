import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from flask import Flask, render_template, jsonify, request, send_from_directory
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from datetime import datetime, timezone
import mimetypes
import openpyxl

app = Flask(__name__)

# .env 파일 로드 (python-dotenv 설치된 경우. 없으면 시스템 환경변수 사용)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ── 설정 ──────────────────────────────────────────────────────────────────────
# 2026-05-11 보안 정리:
# - client_secret*.json, token*.json, sheets_id.json은 .gitignore로 git 추적 제외됨.
# - FORM_SHEET_ID는 환경변수로 분리 (.env.example 참조).
# - 파일명(상수)은 비밀 아니므로 코드에 유지.
# 관련: SECURITY_NOTES.md, wiki/_handoff/decisions.md 2026-05-11 항목
CREDS_FILE       = os.environ.get(
    'GOOGLE_CREDS_FILE',
    'client_secret_469311829534-3jnh14mv6tqbu4g2lurvhmdrblc9vpst.apps.googleusercontent.com.json'
)  # 2026-06-12: umv75na9 OAuth 클라이언트 삭제 → 신규 클라이언트로 교체. .env GOOGLE_CREDS_FILE 우선 적용.
TOKEN_RW         = 'token.json'
TOKEN_RO         = 'token_readonly.json'
TOKEN_SHEETS     = 'token_sheets.json'
SCOPES_RW        = ['https://www.googleapis.com/auth/drive.file']
SCOPES_RO        = ['https://www.googleapis.com/auth/drive.readonly']
SCOPES_SHEETS    = ['https://www.googleapis.com/auth/spreadsheets',
                    'https://www.googleapis.com/auth/drive.file']
LOCAL_ROOT       = 'wiki'
DRIVE_SYNC_ROOT  = 'cnc-wiki'
DRIVE_LIST_ROOT  = 'cnc-wiki 분석자료'
FORM_SHEET_ID    = os.getenv('FORM_SHEET_ID')
FORM_TABS        = ['재연마의뢰', '불량신고', '진행문의', '소모품요청']

if not FORM_SHEET_ID:
    print("⚠️ 경고: FORM_SHEET_ID 환경변수가 없습니다. /api/form-submissions 엔드포인트가 작동하지 않습니다.")
    print("   .env 파일에 FORM_SHEET_ID=... 형식으로 설정하세요. (.env.example 참조)")


# ── 인증 ──────────────────────────────────────────────────────────────────────
def get_service(token_file, scopes):
    creds = None
    if os.path.exists(token_file):
        creds = Credentials.from_authorized_user_file(token_file, scopes)
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
        with open(token_file, 'w') as f:
            f.write(creds.to_json())
    return build('drive', 'v3', credentials=creds)


# ── 유틸 ──────────────────────────────────────────────────────────────────────
def fmt_size(b):
    if b is None: return '-'
    b = int(b)
    if b >= 1024*1024: return f'{b/1024/1024:.1f} MB'
    if b >= 1024:      return f'{b/1024:.1f} KB'
    return f'{b} B'

def fmt_date(iso):
    if not iso: return '-'
    dt = datetime.fromisoformat(iso.replace('Z', '+00:00'))
    return dt.astimezone().strftime('%Y-%m-%d %H:%M')

def local_mtime(path):
    return datetime.fromtimestamp(os.path.getmtime(path), tz=timezone.utc)

def drive_mtime(iso):
    return datetime.fromisoformat(iso.replace('Z', '+00:00'))

def get_or_create_folder(service, name, parent_id=None):
    parent_q = f"and '{parent_id}' in parents" if parent_id else "and 'root' in parents"
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false {parent_q}"
    res = service.files().list(q=q, fields='files(id)').execute()
    folders = res.get('files', [])
    if folders:
        return folders[0]['id']
    meta = {'name': name, 'mimeType': 'application/vnd.google-apps.folder'}
    if parent_id:
        meta['parents'] = [parent_id]
    return service.files().create(body=meta, fields='id').execute()['id']

def get_drive_files(service, folder_id):
    q = f"'{folder_id}' in parents and trashed=false and mimeType != 'application/vnd.google-apps.folder'"
    res = service.files().list(q=q, fields='files(id, name, modifiedTime)').execute()
    return {f['name']: f for f in res.get('files', [])}


# ── 대시보드 헬퍼 ──────────────────────────────────────────────────────────────
def _drive_root_id(service):
    q = f"name='{DRIVE_SYNC_ROOT}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    res = service.files().list(q=q, fields='files(id)').execute()
    folders = res.get('files', [])
    return folders[0]['id'] if folders else None

def navigate_folder(service, path_parts):
    """cnc-wiki 루트에서 path_parts 순서대로 폴더를 탐색해 최종 folder id 반환"""
    current_id = _drive_root_id(service)
    if not current_id:
        return None
    for part in path_parts:
        q = f"'{current_id}' in parents and name='{part}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        res = service.files().list(q=q, fields='files(id)').execute()
        folders = res.get('files', [])
        if not folders:
            return None
        current_id = folders[0]['id']
    return current_id

def find_file(service, folder_id, name):
    q = f"'{folder_id}' in parents and name='{name}' and trashed=false"
    res = service.files().list(q=q, fields='files(id, name, modifiedTime)').execute()
    files = res.get('files', [])
    return files[0] if files else None

def find_latest_file(service, folder_id, contains=''):
    q = f"'{folder_id}' in parents and trashed=false and mimeType != 'application/vnd.google-apps.folder'"
    if contains:
        q += f" and name contains '{contains}'"
    res = service.files().list(q=q, fields='files(id, name, modifiedTime)', orderBy='modifiedTime desc').execute()
    files = res.get('files', [])
    return files[0] if files else None

def download_bytes(service, file_id):
    req = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return buf

def parse_shipping(buf):
    wb = openpyxl.load_workbook(buf, data_only=True)
    ws = wb['요약']
    months = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월']
    customers = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[1]:
            break
        monthly = [row[i + 2] or 0 for i in range(12)]
        customers.append({'name': row[1], 'monthly': monthly, 'total': sum(monthly)})
    customers.sort(key=lambda x: x['total'], reverse=True)
    return {'months': months, 'customers': customers}

def parse_daily(buf):
    wb = openpyxl.load_workbook(buf, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=1, max_row=12, values_only=True))

    def v(r, c):
        try: return rows[r][c]
        except: return None

    return {
        'title':    v(0, 0),
        'meta':     v(1, 0),
        'fast': {
            'qty_actual':  v(3, 2), 'qty_target': v(3, 1), 'qty_avg': v(3, 5),
            'time_actual': v(4, 2),
        },
        'gx7': {
            'qty_actual':  v(5, 2), 'qty_target': v(5, 1), 'qty_avg': v(5, 5),
            'time_actual': v(6, 2),
        },
        'total': {
            'qty_actual':  v(7, 2), 'qty_target': v(7, 1),
            'time_actual': v(8, 2),
        },
        'cumulative':      v(10, 2),
        'cumulative_info': v(10, 0),
    }


# ── 라우트 ────────────────────────────────────────────────────────────────────
DIST_DIR = os.path.join(os.path.dirname(__file__), 'dist')


@app.route('/')
def index():
    if os.path.exists(os.path.join(DIST_DIR, 'index.html')):
        return send_from_directory(DIST_DIR, 'index.html')
    return render_template('index.html')


@app.route('/dashboard')
def dashboard():
    return send_from_directory(DIST_DIR, 'dashboard.html')


@app.route('/<path:filename>', methods=['GET', 'POST'])
def serve_dist(filename):
    # POST = 폼 제출 (로컬 테스트용 — 실제 저장은 GAS에서만)
    if request.method == 'POST':
        return 'ok', 200
    fpath = os.path.join(DIST_DIR, filename)
    if os.path.exists(fpath):
        return send_from_directory(DIST_DIR, filename)
    return '파일을 찾을 수 없습니다.', 404


@app.route('/api/dashboard-data')
def api_dashboard_data():
    try:
        service = get_service(TOKEN_RW, SCOPES_RW)

        # 출하현황
        comp_folder = navigate_folder(service, ['comparisons'])
        if not comp_folder:
            return jsonify(ok=False, error='Drive에서 comparisons 폴더를 찾을 수 없습니다.')
        shipping_file = find_file(service, comp_folder, '출하현황_납품처별_월별분석_2025.xlsx')
        if not shipping_file:
            return jsonify(ok=False, error='출하현황_납품처별_월별분석_2025.xlsx 를 찾을 수 없습니다.')
        shipping = parse_shipping(download_bytes(service, shipping_file['id']))

        # 최신 일일보고
        daily, daily_filename = None, None
        reports_folder = navigate_folder(service, ['reports', 'daily'])
        if reports_folder:
            latest = find_latest_file(service, reports_folder, '일일보고')
            if latest:
                daily = parse_daily(download_bytes(service, latest['id']))
                daily_filename = latest['name']

        return jsonify(ok=True, shipping=shipping, daily=daily, daily_file=daily_filename,
                       fetched_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    except Exception as e:
        import traceback
        return jsonify(ok=False, error=str(e), trace=traceback.format_exc())


@app.route('/api/form-submissions')
def api_form_submissions():
    try:
        creds = Credentials.from_authorized_user_file(TOKEN_SHEETS, SCOPES_SHEETS)
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
            with open(TOKEN_SHEETS, 'w') as f:
                f.write(creds.to_json())
        svc = build('sheets', 'v4', credentials=creds)

        result = {}
        for tab in FORM_TABS:
            res = svc.spreadsheets().values().get(
                spreadsheetId=FORM_SHEET_ID,
                range=f'{tab}!A1:Z500'
            ).execute()
            values = res.get('values', [])
            if len(values) > 1:
                result[tab] = {'headers': values[0], 'rows': list(reversed(values[1:]))}
            else:
                result[tab] = {'headers': values[0] if values else [], 'rows': []}

        return jsonify(ok=True, data=result,
                       fetched_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    except Exception as e:
        import traceback
        return jsonify(ok=False, error=str(e), trace=traceback.format_exc())


@app.route('/api/sync', methods=['POST'])
def api_sync():
    try:
        service = get_service(TOKEN_RW, SCOPES_RW)
        root_id = get_or_create_folder(service, DRIVE_SYNC_ROOT)
        lines = [f'Drive 폴더: {DRIVE_SYNC_ROOT} (id={root_id})', '']
        stats = {'uploaded': 0, 'skipped': 0}

        def sync_dir(local_dir, drive_id, rel=''):
            drive_files = get_drive_files(service, drive_id)
            for entry in sorted(os.scandir(local_dir), key=lambda e: e.name):
                rp = os.path.join(rel, entry.name) if rel else entry.name
                if entry.is_dir():
                    sub_id = get_or_create_folder(service, entry.name, drive_id)
                    sync_dir(entry.path, sub_id, rp)
                elif entry.is_file():
                    if entry.name.startswith('~$'):
                        stats['skipped'] += 1
                        continue
                    existing = drive_files.get(entry.name)
                    need = True
                    if existing:
                        if local_mtime(entry.path) <= drive_mtime(existing['modifiedTime']):
                            need = False
                    if need:
                        mime, _ = mimetypes.guess_type(entry.path)
                        mime = mime or 'application/octet-stream'
                        media = MediaFileUpload(entry.path, mimetype=mime, resumable=True)
                        if existing:
                            service.files().update(fileId=existing['id'], media_body=media).execute()
                            lines.append(f'[업데이트] {rp}')
                        else:
                            meta = {'name': entry.name, 'parents': [drive_id]}
                            service.files().create(body=meta, media_body=media, fields='id').execute()
                            lines.append(f'[신규]     {rp}')
                        stats['uploaded'] += 1
                    else:
                        stats['skipped'] += 1

        sync_dir(LOCAL_ROOT, root_id)
        lines.append('')
        lines.append(f'✅ 완료 — 업로드 {stats["uploaded"]}개 / 건너뜀 {stats["skipped"]}개')
        lines.append(f'https://drive.google.com/drive/folders/{root_id}')
        return jsonify(ok=True, output='\n'.join(lines))
    except Exception as e:
        return jsonify(ok=False, output=f'오류: {e}')


@app.route('/api/list')
def api_list():
    try:
        service = get_service(TOKEN_RW, SCOPES_RW)

        # cnc-wiki 폴더 조회
        q = f"name='{DRIVE_SYNC_ROOT}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        res = service.files().list(q=q, fields='files(id)').execute()
        folders = res.get('files', [])
        if not folders:
            return jsonify(ok=False, output='cnc-wiki 폴더를 찾을 수 없습니다.')

        fid = folders[0]['id']
        lines = [f'📁 {DRIVE_SYNC_ROOT}', f'https://drive.google.com/drive/folders/{fid}', '']

        def list_dir(folder_id, indent=0):
            q2 = f"'{folder_id}' in parents and trashed=false"
            res2 = service.files().list(
                q=q2,
                fields='files(id, name, size, mimeType, modifiedTime)',
                orderBy='folder,name'
            ).execute()
            items = res2.get('files', [])
            pad = '  ' * indent
            for item in items:
                is_dir = item['mimeType'] == 'application/vnd.google-apps.folder'
                icon = '📁' if is_dir else '📄'
                size = '' if is_dir else f"  {fmt_size(item.get('size')):>8}"
                date = fmt_date(item.get('modifiedTime'))
                lines.append(f"{pad}{icon} {item['name']}{size}  {date}")
                if is_dir:
                    list_dir(item['id'], indent + 1)

        list_dir(fid)
        return jsonify(ok=True, output='\n'.join(lines))
    except Exception as e:
        return jsonify(ok=False, output=f'오류: {e}')


@app.route('/api/browse', methods=['POST'])
def api_browse():
    try:
        folder_name = request.json.get('folder', '').strip()
        if not folder_name:
            return jsonify(ok=False, output='폴더명을 입력하세요.')

        service = get_service(TOKEN_RO, SCOPES_RO)
        q = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        res = service.files().list(q=q, fields='files(id, name)').execute()
        folders = res.get('files', [])

        if not folders:
            return jsonify(ok=False, output=f'폴더를 찾을 수 없습니다: {folder_name}')

        lines = []
        for folder in folders:
            fid = folder['id']
            lines.append(f'📁 {folder["name"]}')
            lines.append(f'https://drive.google.com/drive/folders/{fid}')
            lines.append('')

            def list_dir(folder_id, indent=0):
                q2 = f"'{folder_id}' in parents and trashed=false"
                res2 = service.files().list(
                    q=q2,
                    fields='files(id, name, size, mimeType, modifiedTime)',
                    orderBy='folder,name'
                ).execute()
                items = res2.get('files', [])
                pad = '  ' * indent
                for item in items:
                    is_dir = item['mimeType'] == 'application/vnd.google-apps.folder'
                    icon = '📁' if is_dir else '📄'
                    size = '' if is_dir else f"  {fmt_size(item.get('size')):>8}"
                    date = fmt_date(item.get('modifiedTime'))
                    lines.append(f"{pad}{icon} {item['name']}{size}  {date}")
                    if is_dir:
                        list_dir(item['id'], indent + 1)
                return len(items)

            count = list_dir(fid)
            lines.append('')
            lines.append(f'총 {count}개 항목')

        return jsonify(ok=True, output='\n'.join(lines))
    except Exception as e:
        return jsonify(ok=False, output=f'오류: {e}')


if __name__ == '__main__':
    print('cnc-wiki Drive 관리 서버 시작')
    print('브라우저에서 열기: http://localhost:5000')
    app.run(debug=False, host='0.0.0.0', port=5000)
