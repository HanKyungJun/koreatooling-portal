"""
weibull_tool_life.py — 공구 수명 Weibull 분석 스크립트
==========================================================
cnc-wiki / TOOLKOREA 재연마 공구 수명 분석 전용.

## 용도
- 재연마 횟수·가공 수량 데이터로 Weibull 분포를 추정합니다.
- B10/B50/B63.2 수명과 형상 파라미터(β)로 고장 모드를 진단합니다.
- 결과를 PNG 확률지(Probability Plot)와 콘솔로 출력합니다.

## 실행
    python weibull_tool_life.py --input data/weibull_data.csv
    python weibull_tool_life.py --input data/weibull_data.csv --output results/

## 입력 CSV 형식 (wiki/reports/03_계산기·도표/weibull_template.csv 참조)
    tool_id, tool_type, diameter_mm, failure_life, censored
    (censored=1 이면 아직 사용 중 = 우측 검열, 0이면 파손/교체 완료)

## Weibull 2-파라미터 모델
    F(t) = 1 - exp(-(t / η)^β)
    β (형상): <1=초기고장, 1=우발고장, >1=마모고장 (공구는 보통 1~5)
    η (척도): 63.2% 고장 수명 (특성 수명)

## 파라미터 추정
    MLE (Maximum Likelihood Estimation) — scipy.optimize 사용.
    검열 데이터(censored) 포함 처리 가능.

## 출력
    - 콘솔: β, η, B10/B50/B90, 95% CI
    - PNG: Weibull 확률지
    - (선택) Excel: 결과 요약 시트

## 데이터 최소 요건
    ⚠️ Weibull 추정 신뢰도 확보를 위해 완전 고장 데이터(censored=0) 최소 6개 이상 권장.
    샘플 수 < 6: 추정은 가능하나 신뢰구간 매우 넓음 — "추정값" 으로만 사용.

저작: cnc-wiki Cowork / 한경준 (TOOLKOREA) / 2026-06-08
"""

import argparse
import math
import os
import sys
import warnings
import numpy as np
import pandas as pd
from scipy import optimize, stats
import matplotlib
matplotlib.use('Agg')   # 헤드리스 환경 호환
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker

# ─────────────────────────────────────────────
# 1. 데이터 로드
# ─────────────────────────────────────────────

REQUIRED_COLS = {'failure_life', 'censored'}

def load_data(path: str) -> pd.DataFrame:
    """CSV 로드 및 기본 검증."""
    df = pd.read_csv(path, comment='#')
    df.columns = df.columns.str.strip().str.lower()
    missing = REQUIRED_COLS - set(df.columns)
    if missing:
        sys.exit(f"[오류] CSV에 필수 컬럼 누락: {missing}\n"
                 f"  → wiki/reports/03_계산기·도표/weibull_template.csv 참조")
    df = df.dropna(subset=['failure_life'])
    df['failure_life'] = pd.to_numeric(df['failure_life'], errors='coerce')
    df['censored']     = pd.to_numeric(df['censored'],     errors='coerce').fillna(0).astype(int)
    df = df[df['failure_life'] > 0]
    return df


# ─────────────────────────────────────────────
# 2. Weibull MLE (검열 데이터 포함)
# ─────────────────────────────────────────────

def weibull_log_likelihood(params, t_fail, t_cens):
    """
    2-파라미터 Weibull 음의 로그우도 (최소화 대상).
      params = [log(β), log(η)]  — 로그 변환으로 양수 제약 자동 처리
    """
    beta = np.exp(params[0])
    eta  = np.exp(params[1])
    # 완전 고장 항
    ll_fail = np.sum(np.log(beta / eta) + (beta - 1) * np.log(t_fail / eta)
                     - (t_fail / eta) ** beta)
    # 검열 항 (아직 파손 안 됨 → 생존 확률)
    ll_cens = -np.sum((t_cens / eta) ** beta) if len(t_cens) > 0 else 0.0
    return -(ll_fail + ll_cens)


def fit_weibull(df: pd.DataFrame):
    """Weibull 파라미터 추정 + 95% 신뢰구간 (Fisher 정보행렬)."""
    t_fail = df.loc[df['censored'] == 0, 'failure_life'].values
    t_cens = df.loc[df['censored'] == 1, 'failure_life'].values

    n_fail = len(t_fail)
    if n_fail < 2:
        sys.exit(f"[오류] 완전 고장 데이터(censored=0)가 {n_fail}개입니다. 최소 2개 필요.")
    if n_fail < 6:
        warnings.warn(f"[경고] 완전 고장 데이터 {n_fail}개 — 신뢰구간 매우 넓음. 추정값으로만 사용.")

    # 초기값: MoM 추정
    log_t = np.log(t_fail)
    beta0  = np.pi / (np.sqrt(6) * np.std(log_t, ddof=1)) if np.std(log_t) > 0 else 1.5
    eta0   = np.exp(np.mean(log_t))
    x0 = [np.log(beta0), np.log(eta0)]

    result = optimize.minimize(
        weibull_log_likelihood, x0, args=(t_fail, t_cens),
        method='Nelder-Mead',
        options={'xatol': 1e-8, 'fatol': 1e-8, 'maxiter': 10000}
    )
    if not result.success:
        warnings.warn(f"[경고] 최적화 수렴 불완전: {result.message}")

    beta_hat = np.exp(result.x[0])
    eta_hat  = np.exp(result.x[1])

    # 수치 헤시안으로 분산 추정 → 95% CI
    try:
        hess = _numerical_hessian(weibull_log_likelihood, result.x, t_fail, t_cens)
        cov  = np.linalg.inv(hess)
        se   = np.sqrt(np.diag(cov))
        # 델타법: CI on log scale, 역변환
        z95  = stats.norm.ppf(0.975)
        beta_ci = (np.exp(result.x[0] - z95 * se[0]),
                   np.exp(result.x[0] + z95 * se[0]))
        eta_ci  = (np.exp(result.x[1] - z95 * se[1]),
                   np.exp(result.x[1] + z95 * se[1]))
    except Exception:
        beta_ci = eta_ci = (None, None)

    return beta_hat, eta_hat, beta_ci, eta_ci, n_fail, len(t_cens)


def _numerical_hessian(f, x, *args, eps=1e-5):
    """중앙차분으로 헤시안 근사."""
    n = len(x)
    H = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            xi, xj = x.copy(), x.copy()
            xij = x.copy()
            xi[i] += eps; xj[j] += eps; xij[i] += eps; xij[j] += eps
            H[i, j] = (f(xij, *args) - f(xi, *args) - f(xj, *args) + f(x, *args)) / eps**2
    return H


# ─────────────────────────────────────────────
# 3. Weibull 수명 계산
# ─────────────────────────────────────────────

def weibull_quantile(p, beta, eta):
    """P% 누적 고장 확률에 대응하는 수명 t."""
    return eta * (-np.log(1 - p)) ** (1 / beta)


def reliability(t, beta, eta):
    """시간 t에서의 신뢰도 R(t) = 생존 확률."""
    return np.exp(-(t / eta) ** beta)


def failure_mode_label(beta):
    if beta < 0.75:
        return "초기고장 (Infant Mortality) — 생산·조립 불량 의심"
    elif beta < 1.25:
        return "우발고장 (Random Failure) — 외부 충격·사용 조건 변동"
    elif beta < 3.0:
        return "초기 마모고장 (Early Wear-out) — 정상 공구 마모"
    elif beta < 5.0:
        return "마모고장 (Wear-out) — 명확한 수명 한계 존재"
    else:
        return "급속 마모고장 (Rapid Wear-out) — 가공 조건 재검토 권장"


# ─────────────────────────────────────────────
# 4. Weibull 확률지 (Probability Plot) 출력
# ─────────────────────────────────────────────

def median_rank(n, k):
    """Median rank (Bernard's approximation): (k - 0.3) / (n + 0.4)."""
    return (k - 0.3) / (n + 0.4)


def plot_weibull(df, beta, eta, beta_ci, eta_ci, output_dir, title=""):
    t_fail = np.sort(df.loc[df['censored'] == 0, 'failure_life'].values)
    n      = len(t_fail)
    ranks  = np.array([median_rank(n, k+1) for k in range(n)])

    # Weibull 변환: x=ln(t), y=ln(-ln(1-F))
    x_data = np.log(t_fail)
    y_data = np.log(-np.log(1 - ranks))

    t_line = np.linspace(t_fail.min() * 0.5, t_fail.max() * 2, 200)
    x_line = np.log(t_line)
    y_line = np.log(-np.log(1 - (1 - np.exp(-(t_line / eta) ** beta))))

    fig, ax = plt.subplots(figsize=(9, 6))
    ax.scatter(x_data, y_data, color='steelblue', zorder=5, label='실측 데이터 (median rank)')
    ax.plot(x_line, y_line, 'r-', lw=2, label=f'Weibull 적합 (β={beta:.2f}, η={eta:.1f})')

    # 검열 데이터 표시
    t_cens = df.loc[df['censored'] == 1, 'failure_life'].values
    if len(t_cens) > 0:
        ax.axvline(np.log(t_cens.max()), color='gray', ls='--', lw=1,
                   label=f'검열 최대 수명 ({t_cens.max():.0f})')

    # B10/B50 수직선
    for p, color in [(0.1, 'green'), (0.5, 'orange')]:
        t_p = weibull_quantile(p, beta, eta)
        ax.axvline(np.log(t_p), color=color, ls=':', lw=1.5,
                   label=f'B{int(p*100)}={t_p:.1f}')
        ax.axhline(np.log(-np.log(1 - p)), color=color, ls=':', lw=1.0)

    ax.set_xlabel('ln(수명)', fontsize=11)
    ax.set_ylabel('ln(-ln(1-F(t)))', fontsize=11)
    ax.set_title(f'Weibull 확률지 — {title}' if title else 'Weibull 확률지', fontsize=13)
    ax.legend(fontsize=9)
    ax.grid(True, alpha=0.3)

    # 보조축: 실제 수명 / 확률 값
    ax2 = ax.twiny()
    t_ticks = np.array([t_fail.min(), weibull_quantile(0.1, beta, eta),
                        weibull_quantile(0.5, beta, eta), eta, t_fail.max()])
    t_ticks = np.unique(np.round(t_ticks, 0))
    ax2.set_xlim(ax.get_xlim())
    ax2.set_xticks(np.log(t_ticks))
    ax2.set_xticklabels([f'{int(t)}' for t in t_ticks], fontsize=8)
    ax2.set_xlabel('수명 (원래 단위)', fontsize=9)

    fig.tight_layout()
    fname = os.path.join(output_dir, 'weibull_plot.png')
    fig.savefig(fname, dpi=150)
    plt.close(fig)
    print(f"  → 확률지 저장: {fname}")
    return fname


# ─────────────────────────────────────────────
# 5. 결과 출력
# ─────────────────────────────────────────────

def print_results(beta, eta, beta_ci, eta_ci, n_fail, n_cens, life_unit="개"):
    b10 = weibull_quantile(0.1,  beta, eta)
    b50 = weibull_quantile(0.5,  beta, eta)
    b90 = weibull_quantile(0.9,  beta, eta)
    mttf = eta * math.gamma(1 + 1/beta)

    ci_str = lambda ci: f"[{ci[0]:.2f}, {ci[1]:.2f}]" if ci[0] else "계산 불가"

    print()
    print("=" * 55)
    print("  Weibull 수명 분석 결과")
    print("=" * 55)
    print(f"  데이터: 완전고장 {n_fail}개 / 검열 {n_cens}개")
    print()
    print(f"  β (형상 파라미터):  {beta:.3f}  95%CI {ci_str(beta_ci)}")
    print(f"  η (척도 파라미터):  {eta:.1f} {life_unit}  95%CI {ci_str(eta_ci)}")
    print()
    print(f"  고장 모드: {failure_mode_label(beta)}")
    print()
    print(f"  B10 수명 (10% 고장): {b10:>8.1f} {life_unit}")
    print(f"  B50 수명 (50% 고장): {b50:>8.1f} {life_unit}")
    print(f"  B90 수명 (90% 고장): {b90:>8.1f} {life_unit}")
    print(f"  MTTF (평균 수명):    {mttf:>8.1f} {life_unit}")
    print()

    if n_fail < 6:
        print("  ⚠️  샘플 수 부족 (<6) — 위 수치는 추정값입니다.")
        print("     데이터 누적 후 재분석 권장.")
    print("=" * 55)


def save_excel(df, beta, eta, beta_ci, eta_ci, n_fail, n_cens, life_unit, output_dir):
    """결과를 Excel로 저장."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [안내] openpyxl 없음 — Excel 출력 생략.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weibull 결과"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    # 파라미터 표
    rows = [
        ["파라미터", "값", "95% CI 하한", "95% CI 상한"],
        ["β (형상)", round(beta, 4),
         round(beta_ci[0], 4) if beta_ci[0] else "—",
         round(beta_ci[1], 4) if beta_ci[1] else "—"],
        ["η (척도)", round(eta, 2),
         round(eta_ci[0], 2) if eta_ci[0] else "—",
         round(eta_ci[1], 2) if eta_ci[1] else "—"],
        [],
        ["수명 지표", f"수명 ({life_unit})", "", ""],
        ["B10 (10% 고장)", round(weibull_quantile(0.1, beta, eta), 2), "", ""],
        ["B50 (50% 고장)", round(weibull_quantile(0.5, beta, eta), 2), "", ""],
        ["B90 (90% 고장)", round(weibull_quantile(0.9, beta, eta), 2), "", ""],
        ["MTTF (평균 수명)", round(eta * math.gamma(1 + 1/beta), 2), "", ""],
        [],
        ["고장 모드", failure_mode_label(beta), "", ""],
        ["데이터 수 (완전)", n_fail, "", ""],
        ["데이터 수 (검열)", n_cens, "", ""],
    ]
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            if i == 1 or i == 5:
                cell.font = header_font
                cell.fill = header_fill

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # 원본 데이터 시트
    ws2 = wb.create_sheet("원본 데이터")
    for col_i, col_name in enumerate(df.columns, 1):
        ws2.cell(row=1, column=col_i, value=col_name).font = Font(bold=True)
    for row_i, row in enumerate(df.itertuples(index=False), 2):
        for col_i, val in enumerate(row, 1):
            ws2.cell(row=row_i, column=col_i, value=val)

    fpath = os.path.join(output_dir, "weibull_results.xlsx")
    wb.save(fpath)
    print(f"  → Excel 저장: {fpath}")


# ─────────────────────────────────────────────
# 6. CLI 진입점
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="공구 수명 Weibull 분석 — cnc-wiki TOOLKOREA",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument('--input',  '-i', required=True,
                        help='입력 CSV 파일 경로\n  예: data/weibull_data.csv')
    parser.add_argument('--output', '-o', default='.',
                        help='결과 저장 폴더 (기본: 현재 폴더)')
    parser.add_argument('--unit',   '-u', default='개',
                        help='수명 단위 표시 문자열 (기본: 개)')
    parser.add_argument('--title',  '-t', default='',
                        help='그래프 제목 (공구 종류 등)')
    parser.add_argument('--no-plot', action='store_true',
                        help='확률지 PNG 미출력')
    parser.add_argument('--excel',   action='store_true',
                        help='Excel 결과 파일 출력')
    args = parser.parse_args()

    os.makedirs(args.output, exist_ok=True)

    print(f"[1] 데이터 로드: {args.input}")
    df = load_data(args.input)
    print(f"    → {len(df)}행 로드 완료 (고장: {(df['censored']==0).sum()}, 검열: {(df['censored']==1).sum()})")

    print("[2] Weibull MLE 추정 중...")
    beta, eta, beta_ci, eta_ci, n_fail, n_cens = fit_weibull(df)

    print_results(beta, eta, beta_ci, eta_ci, n_fail, n_cens, args.unit)

    if not args.no_plot:
        print("[3] 확률지 출력...")
        plot_weibull(df, beta, eta, beta_ci, eta_ci, args.output, args.title)

    if args.excel:
        print("[4] Excel 저장...")
        save_excel(df, beta, eta, beta_ci, eta_ci, n_fail, n_cens, args.unit, args.output)

    print("\n완료.")


if __name__ == '__main__':
    main()
