# -*- coding: utf-8 -*-
"""
재연마 단가 재계산 (26.09 정가 인상 반영)
  신단가 = ROUNDUP(정가 x 할인율, -1)   ← 10원 단위 올림
  * 할인율은 파일에 이미 입력된 값을 그대로 사용 (한경준님 확인 2026-08-21: "정가만 인상, 할인율은 그대로")
  * 원본 파일은 수정하지 않고 별도 파일로 저장 (서식·수식·드롭다운 보존)
사용법:  python recalc.py <입력폴더> <출력폴더>
"""
import sys, math, shutil
from pathlib import Path
import openpyxl, pandas as pd

SHEET = '품목일괄등록'
DATA_START = 3          # 1-based: 1행 대분류헤더, 2행 세부헤더, 3행부터 데이터
C_CODE, C_NAME, C_MAKER = 2, 3, 6
C_LIST, C_SDISC, C_SPRICE, C_PDISC, C_PPRICE = 28, 29, 30, 31, 32   # 정가/매출할인율/매출단가/매입할인율/매입단가

def roundup10(x):
    """10원 단위 올림. 부동소수 오차 방지용 6자리 정규화 후 ceil."""
    return int(math.ceil(round(x, 6) / 10.0)) * 10

def load_pricelist(pl_path):
    """정가표 시트별 고유 금액 집합 → 정가 갱신 여부 판정용"""
    xl = pd.ExcelFile(pl_path)
    def vals(sh):
        d = pd.read_excel(pl_path, sheet_name=sh, header=None)
        out = set()
        for r in range(d.shape[0]):
            for c in range(d.shape[1]):
                v = d.iat[r, c]
                if isinstance(v, (int, float)) and pd.notna(v) and v >= 1000:
                    out.add(int(v))
        return out
    v09 = vals('정가표(26.09)') if '정가표(26.09)' in xl.sheet_names else set()
    v05 = vals('정가표(26.05)') if '정가표(26.05)' in xl.sheet_names else set()
    return v09, v05

def classify(listprice, v09, v05):
    if listprice is None:      return '정가없음'
    lp = int(listprice)
    if lp in v09 and lp in v05: return '공통값(판정보류)'
    if lp in v09:               return '신정가(26.09)'
    if lp in v05:               return '구정가(26.05)-갱신필요'
    return '정가표에없음'

def process(src, dst, v09, v05):
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb[SHEET]
    rows = []
    for r in range(DATA_START, ws.max_row + 1):
        code = ws.cell(r, C_CODE).value
        if code in (None, ''):
            continue
        lp    = ws.cell(r, C_LIST).value
        sd    = ws.cell(r, C_SDISC).value
        pd_   = ws.cell(r, C_PDISC).value
        old_s = ws.cell(r, C_SPRICE).value
        old_p = ws.cell(r, C_PPRICE).value
        stat  = classify(lp if isinstance(lp,(int,float)) else None, v09, v05)

        new_s = new_p = None
        if isinstance(lp, (int, float)) and isinstance(sd, (int, float)):
            new_s = 0 if sd == 0 else roundup10(lp * sd)
            ws.cell(r, C_SPRICE).value = new_s
        if isinstance(lp, (int, float)) and isinstance(pd_, (int, float)):
            new_p = 0 if pd_ == 0 else roundup10(lp * pd_)
            ws.cell(r, C_PPRICE).value = new_p

        rows.append(dict(
            파일=Path(src).stem, 품목코드=code, 품목명=ws.cell(r, C_NAME).value,
            메이커=ws.cell(r, C_MAKER).value, 정가상태=stat, 정가=lp,
            매출할인율=sd, 구_매출단가=old_s, 신_매출단가=new_s,
            매출_증감=(new_s - old_s) if isinstance(old_s,(int,float)) and new_s is not None else None,
            매입할인율=pd_, 구_매입단가=old_p, 신_매입단가=new_p,
            매입_증감=(new_p - old_p) if isinstance(old_p,(int,float)) and new_p is not None else None,
        ))
    wb.save(dst)
    return rows

def main():
    src_dir, out_dir = Path(sys.argv[1]), Path(sys.argv[2])
    out_dir.mkdir(parents=True, exist_ok=True)
    pl = src_dir / '26.09 재연마정가표_정리.xlsx'
    v09, v05 = load_pricelist(pl)
    print(f'정가표 고유금액 — 26.09: {len(v09)} / 26.05: {len(v05)} / 26.05전용: {len(v05-v09)}')
    allrows = []
    for f in sorted(src_dir.glob('260901 단가 수정*.xlsx')):
        out = out_dir / (f.stem + '_단가재계산.xlsx')
        rows = process(f, out, v09, v05)
        allrows += rows
        print(f'  {f.name:<34} {len(rows):>6}행 → {out.name}')
    df = pd.DataFrame(allrows)
    # 검수 플래그
    def flag(r):
        f=[]
        if r['정가상태']=='구정가(26.05)-갱신필요': f.append('정가 미갱신')
        if r['정가상태']=='정가표에없음':          f.append('정가표 미등재')
        if r['정가상태']=='정가없음':              f.append('정가 공란')
        if (r['매출_증감'] or 0) < 0 or (r['매입_증감'] or 0) < 0: f.append('단가 인하(역주행)')
        return ' / '.join(f)
    df['검수플래그'] = df.apply(flag, axis=1)
    rpt = out_dir / '단가변경_대조표.xlsx'
    with pd.ExcelWriter(rpt, engine='openpyxl') as w:
        df.to_excel(w, sheet_name='전체', index=False)
        df[df['검수플래그']!=''].to_excel(w, sheet_name='확인필요', index=False)
        df[df['검수플래그']==''].to_excel(w, sheet_name='이상없음', index=False)
        df.groupby(['파일','정가상태']).size().rename('건수').reset_index() \
          .to_excel(w, sheet_name='요약_정가상태', index=False)
        df.assign(플래그=df['검수플래그'].replace('','(이상없음)')) \
          .groupby(['파일','플래그']).size().rename('건수').reset_index() \
          .to_excel(w, sheet_name='요약_검수플래그', index=False)
    print(f'\n대조표 → {rpt.name}  (전체 {len(df)}행)')
    print('\n=== 파일 x 정가상태 ===')
    print(pd.crosstab(df['파일'], df['정가상태']).to_string())
    n_ok=(df['검수플래그']=='').sum()
    print(f"\n검수 결과: 이상없음 {n_ok}행 / 확인필요 {len(df)-n_ok}행")

if __name__ == '__main__':
    main()
