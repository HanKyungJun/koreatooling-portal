#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
재연마 표준 공수 DB 워크북 생성

실행:
    python scripts/parse_worklog.py      # 먼저 실행 (CSV 생성)
    python scripts/build_costdb.py       # 이 스크립트

입력: outputs/worklog_parsed/jobs_all.csv, meter_all.csv
출력: wiki/measurements/재연마_표준공수DB_2022-2026.xlsx  (10개 시트)

━━━ 정규화 규칙 (2026-08-28 한경준님 확인) ━━━
  평 = 스퀘어        → 통합 ('평(스퀘어)')
  면취 = 챔퍼        → 통합 ('면취')
  HSS               → 형상에서 분리, '소재' 열로
  NC드릴 != 드릴     → 분리 유지 (서로 다른 가공 파일, 형상 상이)
  각도 표기          → 중요도 낮음, 집계 제외
  단위시간 <= 1초     → 집계 제외 (전량 1.0초 단일값 = 미입력 자리표시).
                       중앙값에는 영향 없음, 평균·P80 보호 목적.

주의: 「가공시간」은 장비 사이클 타임이다 — 셋업·대기·검사 미포함.
      개당 원가는 「일 인건비 / 일 생산량」으로 내야 하며 보정계수는 불필요하다.
      (「개당 원가 배분」 시트 참조)

주의: LibreOffice 없이 실행하면 수식 셀의 캐시값이 비어 있다.
      Excel에서 열면 자동 계산되므로 실사용에는 문제 없다.
"""
from pathlib import Path
BASE = Path(__file__).resolve().parent.parent
DATA = BASE / 'outputs' / 'worklog_parsed'
OUTF = BASE / 'wiki' / 'measurements' / '재연마_표준공수DB_2022-2026.xlsx'

import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table,TableStyleInfo

F='맑은 고딕'
NAVY='1F3864'; HDRB='2F5597'; W='FFFFFF'; LGRAY='F2F2F2'; WARN='FFF2CC'; BAD='FADBD8'; OK='E2EFDA'
thin=Side(style='thin',color='BFBFBF'); B=Border(left=thin,right=thin,top=thin,bottom=thin)


import re


def _normalize(df):
    """원시 파싱 결과에 형상·코팅 정규화 열을 붙인다. (규칙은 파일 상단 주석 참조)"""
    def one(x):
        s = str(x).strip()
        out = {'라핑': False, '소재': '초경', '각도': ''}
        if '라핑' in s:
            out['라핑'] = True
            s = re.sub(r'[_ ]?라핑', '', s)
        m = re.search(r'[_,. ]?(HSS|Hss|hss)\b', s)
        if m:
            out['소재'] = 'HSS'
            s = s[:m.start()] + s[m.end():]
        m = re.search(r"[_ ]?([0-9]{2,3})\s*(?:도|°|˚|`|')", s)
        if m:
            out['각도'] = m.group(1)
            s = s[:m.start()] + s[m.end():]
        m = re.search(r'[_ ]([0-9]{2,3})$', s)
        if m and s.split('_')[0] in ('면취', 'NC', 'nc', '드릴'):
            out['각도'] = m.group(1)
            s = s[:m.start()]
        s = s.replace('엔드밀', '').strip(' _,.')
        low = s.lower()
        if '코너' in s or re.fullmatch(r'[rR]\s*[0-9.]+', s):
            m2 = re.search(r'([0-9]+(?:\.[0-9]+)?)', s)
            s = '코너R' + ('%g' % float(m2.group(1))) if m2 else '코너R'
        elif low in ('볼', '봃', '볼_롱'):
            s = '볼'
        elif low.startswith('평') or low.startswith('스퀘어'):
            s = '평(스퀘어)'
        elif '면취' in s or '챔퍼' in s or low.startswith('c'):
            s = '면취'
        elif low.startswith('nc'):
            s = 'NC드릴'
        elif '드릴' in s:
            s = '드릴'
        out['형상'] = s
        return out

    n = pd.DataFrame([one(v) for v in df['형상']])
    df = df.rename(columns={'형상': '형상_원본', '코팅': '코팅_원본'})
    df = pd.concat([df.reset_index(drop=True), n], axis=1)
    df['직경'] = pd.to_numeric(df['날경'], errors='coerce').round(2)
    df['날수'] = pd.to_numeric(df['날수F'], errors='coerce')
    CO = {'뮤': '뮤', '알파': '알파', '알': '알파', '비': '비코팅', '비코팅': '비코팅',
          '블루': '블루', '블': '블루', 'ts': 'TS', 'TS': 'TS', '틴': 'TiN', '티진': 'TiN',
          '알틴': 'AlTiN', 'TICN': 'TiCN', '다이아': '다이아', '메가크롬': '메가크롬', 'AB': 'AB'}
    df['코팅_정규'] = df['코팅_원본'].astype(str).str.strip().map(
        lambda x: CO.get(x, x if x and x != 'nan' else '미기재'))
    return df


df = _normalize(pd.read_csv(DATA / 'jobs_all.csv'))
OUTLIER=df['단위시간s']<=1
clean=df[~OUTLIER].copy()

def wstats(g):
    v=np.repeat(g['단위시간s'].values,g['수량'].values.astype(int))
    return pd.Series({'표본수':len(v),'작업건수':len(g),'중앙값':np.median(v),'평균':v.mean(),
                      'P20':np.percentile(v,20),'P80':np.percentile(v,80),'최소':v.min(),'최대':v.max(),'표준편차':v.std()})
db=clean.groupby(['형상','직경','날수','소재'],dropna=False).apply(wstats,include_groups=False).reset_index()
db=db[db['표본수']>=10].copy()
db['변동계수%']=(db['표준편차']/db['평균']*100).round(1)
db['중앙값_분']=(db['중앙값']/60).round(2)
for c in ['중앙값','평균','P20','P80','최소','최대','표준편차']: db[c]=db[c].round(0).astype(int)
db=db.sort_values(['형상','직경','날수']).reset_index(drop=True)

shp=clean.groupby('형상').apply(wstats,include_groups=False).reset_index().sort_values('표본수',ascending=False)
for c in ['중앙값','평균','P20','P80','최소','최대','표준편차']: shp[c]=shp[c].round(0).astype(int)

TOP=[('볼',6.0,2),('볼',10.0,2),('볼',8.0,2),('볼',12.0,2),('평(스퀘어)',10.0,4),('평(스퀘어)',12.0,4),
     ('평(스퀘어)',6.0,4),('면취',6.0,2),('면취',8.0,2),('면취',10.0,2),('코너R1',10.0,4),('코너R1',12.0,4),
     ('코너R0.5',6.0,2),('코너R0.5',8.0,4),('드릴',6.0,2)]
tr=[]
for s,d_,f_ in TOP:
    sub=clean[(clean['형상']==s)&(clean['직경']==d_)&(clean['날수']==f_)]
    if len(sub)<10: continue
    row={'형상':s,'직경':d_,'날수':int(f_)}
    for y in range(2022,2027):
        yy=sub[sub['연']==y]
        if len(yy)==0: row[str(y)]=None; continue
        v=np.repeat(yy['단위시간s'].values,yy['수량'].values.astype(int))
        row[str(y)]=int(np.median(v)) if len(v) else None
    row['표본수']=int(sub['수량'].sum())
    tr.append(row)
tr=pd.DataFrame(tr)

eq=clean.pivot_table(index='형상',columns='설비',values='수량',aggfunc='sum',fill_value=0).reset_index()
if 'FG' not in eq: eq['FG']=0
if 'GX7' not in eq: eq['GX7']=0
eq['합계']=eq['FG']+eq['GX7']
eq['FG비중%']=(eq['FG']/eq['합계']*100).round(1)
eq=eq.sort_values('합계',ascending=False)

yr=clean.groupby('연').agg(작업건수=('수량','size'),총수량=('수량','sum'),
    총사이클h=('시간합계s',lambda x:round(x.sum()/3600,1))).reset_index()
yr['조업일수']=clean.groupby('연').apply(lambda g:g[['월','일']].drop_duplicates().shape[0],include_groups=False).values
yr['일평균수량']=(yr['총수량']/yr['조업일수']).round(1)
_dh=clean.groupby(['연','월','일'])['시간합계s'].sum().reset_index(); _dh['h']=_dh['시간합계s']/3600
_n=_dh.groupby('연')['h'].agg(일수량중앙=('median'),초과12h=(lambda x:(x>12).sum())).reset_index()
_n.columns=['연','사이클중앙h','12h초과일']
_n['사이클중앙h']=_n['사이클중앙h'].round(2)
yr=yr.merge(_n,on='연',how='left')
yr['근무형태']=yr['연'].map({2022:'야간 운용',2023:'야간 운용(축소)',2024:'주간만',2025:'주간만',2026:'주간만 (초과일=로더전담)'})

wb=Workbook(); wb.remove(wb.active)

def style_hdr(ws,row,ncol,bg=HDRB):
    for c in range(1,ncol+1):
        cl=ws.cell(row=row,column=c)
        cl.font=Font(name=F,bold=True,color=W,size=13)
        cl.fill=PatternFill('solid',fgColor=bg)
        cl.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        cl.border=B
    ws.row_dimensions[row].height=30

def put_df(ws,d,start=1,hdr_row=None):
    hr=hdr_row or start
    for j,col in enumerate(d.columns,1):
        ws.cell(row=hr,column=j,value=str(col))
    style_hdr(ws,hr,len(d.columns))
    for i,(_,r) in enumerate(d.iterrows(),hr+1):
        for j,col in enumerate(d.columns,1):
            v=r[col]
            if pd.isna(v): v=None
            elif isinstance(v,(np.integer,)): v=int(v)
            elif isinstance(v,(np.floating,)): v=float(v)
            elif isinstance(v,(np.bool_,)): v=bool(v)
            cl=ws.cell(row=i,column=j,value=v)
            cl.font=Font(name=F,size=13)
            cl.border=B
            cl.alignment=Alignment(horizontal='center' if j>2 else 'left',vertical='center')
            if i%2==0: cl.fill=PatternFill('solid',fgColor=LGRAY)
    return hr+len(d)

def autow(ws,widths):
    for i,w_ in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w_

def title(ws,text,ncol,row=1,sz=22):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncol)
    c=ws.cell(row=row,column=1,value=text)
    c.font=Font(name=F,bold=True,size=sz,color=W)
    c.fill=PatternFill('solid',fgColor=NAVY)
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[row].height=36

# ── 1. 개요 ──
ws=wb.create_sheet('개요')
title(ws,'재연마 표준 공수 DB — 2022~2026 실적 기반',6)
rows=[
 ('','',''),
 ('■ 목적','','재연마 품목별 실제 사이클 타임을 집계해, 원가 산정·견적·설비 투자 검토의 정량 근거로 사용한다.'),
 ('','',''),
 ('■ 데이터 범위','',''),
 ('원본','','raw/출하현황/재연마 작업일지(2022~2026)/재연마_월간생산일지 (M월).xls — 60개 파일'),
 ('추출 방식','','각 일자 시트의 FAST GRIND·GX7 블록을 동적 탐지 후 작업 행만 추출 (합계행·계기값행 제외)'),
 ('작업 레코드','',13145),
 ('총 재연마 수량','',84010),
 ('총 사이클 시간','','6,298 시간'),
 ('집계 기준일','','2026-08-28'),
 ('','',''),
 ('■ 「사이클 타임」의 성격 — 반드시 확인','',''),
 ('정의','','ANCA 장비가 기록한 개당 사이클 타임 [신뢰도: 사내 확인 — 한경준님, 2026-08-28]'),
 ('포함되지 않는 것','','셋업 시간, 대기 시간, 검사·측정 시간, 휠 드레싱, 공구 교체'),
 ('⚠️ 원가 산정 시','','이 값을 그대로 개당 원가에 쓰면 과소평가된다. 셋업·간접시간 보정계수를 별도로 적용할 것.'),
 ('','',''),
 ('■ 통계 방법','',''),
 ('가중 방식','','수량 가중 — 1개 작업 행(수량 n개)을 n개 표본으로 전개해 집계'),
 ('대표값','','중앙값(median). 평균은 이상치에 끌려가므로 중앙값을 1차 기준으로 삼는다.'),
 ('P80','','상위 20% 지점. 견적·납기 산정 시 안전측 값으로 사용.'),
 ('수록 기준','','표본 10개 이상인 조합만 수록'),
 ('','',''),
 ('■ 신뢰도','','실측 검증 — 원본 작업일지 직접 판독. 단, 입력값 자체의 정확도는 현장 입력에 의존한다.'),
 ('','',''),
 ('■ 제외·주의','',''),
 ('이상치 제외','','단위시간이 정확히 1.0초인 레코드 511건(3.89%)을 집계에서 제외. 전량 단일값 1.0초 — 실제 가공시간이 아니라 미입력 자리표시로 판단.'),
 ('','','  └ 제외해도 중앙값은 변하지 않는다(볼 Ø6 2날: 제외 240초 = 포함 240초). 평균·P80을 보호하기 위해서만 제외한다.'),
 ('금액 미수록','','원본의 가공비·금액 열은 실거래 단가이므로 추출하지 않았다 (CLAUDE.md §4 대외비).'),
 ('계기값 미수록','','일자 시트 하단 계기값 행(741건)은 리셋이 확인됐으나 시간 단위로 해석되지 않아 미수록 — 「설비 계기값」 시트 참조.'),
 ('NC드릴/드릴 분리','','NC드릴과 드릴은 서로 다른 가공 파일을 쓰고 형상도 달라 별도 항목으로 유지 [사내 확인, 2026-08-28].'),
]
r=3
for a,b_,c_ in rows:
    ws.cell(row=r,column=1,value=a).font=Font(name=F,bold=a.startswith('■'),size=15 if a.startswith('■') else 13,color=NAVY if a.startswith('■') else '000000')
    ws.cell(row=r,column=3,value=c_).font=Font(name=F,size=13)
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6)
    ws.cell(row=r,column=3).alignment=Alignment(vertical='center',wrap_text=True)
    r+=1
autow(ws,[20,2,40,20,20,20])
ws.column_dimensions['C'].width=30
for i in range(4,7): ws.column_dimensions[get_column_letter(i)].width=22

# ── 2. 표준 공수 DB ──
ws=wb.create_sheet('표준 공수 DB')
title(ws,'표준 공수 DB — 형상 × 직경 × 날수 × 소재',13)
ws.cell(row=2,column=1,value='단위: 초/개 · 수량 가중 중앙값 기준 · 표본 10개 이상만 수록 · 중앙값·P20·P80·표준편차는 수량가중 계산값이라 셀 수식으로 표현 불가(원천데이터 시트에서 재현 가능)').font=Font(name=F,size=11,italic=True)
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=13)
d=db[['형상','직경','날수','소재','표본수','작업건수','중앙값','중앙값_분','P20','P80','최소','최대','변동계수%']].copy()
d['난이도지수']=0.0
put_df(ws,d,hdr_row=3)
for _i in range(4,4+len(d)):
    _c=ws.cell(row=_i,column=14,value=f"=ROUND(G{_i}/'개당 원가 배분'!$C$10,2)")
    _c.font=Font(name=F,size=13); _c.border=B; _c.alignment=Alignment(horizontal='center',vertical='center')
    if _i%2==0: _c.fill=PatternFill('solid',fgColor=LGRAY)
autow(ws,[14,8,7,7,9,10,10,11,9,9,9,9,11,11])
ws.freeze_panes='A4'
for i in range(4,4+len(d)):
    cv=ws.cell(row=i,column=13).value
    if cv is not None and cv>=35: ws.cell(row=i,column=13).fill=PatternFill('solid',fgColor=WARN)

# ── 3. 형상별 요약 ──
ws=wb.create_sheet('형상별 요약')
title(ws,'형상별 요약',10)
s=shp[['형상','표본수','작업건수','중앙값','평균','P20','P80','최소','최대','변동계수%']].copy() if '변동계수%' in shp else shp
s=shp[['형상','표본수','작업건수','중앙값','평균','P20','P80','최소','최대']].copy()
s['표본수']=s['표본수'].astype(int); s['작업건수']=s['작업건수'].astype(int)
end=put_df(ws,s,hdr_row=3)
ws.cell(row=end+1,column=1,value='합계').font=Font(name=F,bold=True,size=13)
ws.cell(row=end+1,column=2,value=f'=SUM(B4:B{end})').font=Font(name=F,bold=True,size=13)
ws.cell(row=end+1,column=3,value=f'=SUM(C4:C{end})').font=Font(name=F,bold=True,size=13)
for c in range(1,11): ws.cell(row=end+1,column=c).fill=PatternFill('solid',fgColor=WARN); ws.cell(row=end+1,column=c).border=B
autow(ws,[16,10,10,10,10,9,9,9,9,11])
ws.freeze_panes='A4'

# ── 4. 연도별 추이 ──
ws=wb.create_sheet('연도별 추이')
title(ws,'연도별 추이',10)
ws.cell(row=2,column=1,value='① 전체 생산 추이').font=Font(name=F,bold=True,size=15,color=NAVY)
e1=put_df(ws,yr,hdr_row=3)
r2=e1+3
ws.cell(row=r2-1,column=1,value='② 주요 품목 사이클 타임 중앙값 추이 (초/개)').font=Font(name=F,bold=True,size=15,color=NAVY)
cols=['형상','직경','날수','2022','2023','2024','2025','2026','표본수']
tr2=tr[[c for c in cols if c in tr.columns]]
e2=put_df(ws,tr2,hdr_row=r2)
ws.cell(row=e2+2,column=1,value='※ 빈 칸은 해당 연도 표본 부족. 조업일수 = 작업 기록이 1건 이상 있는 날의 수.').font=Font(name=F,size=11,italic=True)
ws.cell(row=e2+3,column=1,value='※ 야간 작업은 2023년까지만 운용됐다 [사내 확인 — 한경준님, 2026-08-28]. 2024년 이후 주간 근무만이다.').font=Font(name=F,size=11,italic=True)
ws.cell(row=e2+4,column=1,value='※ 🔴 2022년 상반기가 야간 작업기다 — 119 조업일 중 43일(36%)이 12시간 초과, 일 수량 중앙값 136개. 2022 하반기 87개 → 2023 상반기 78개로 급감.').font=Font(name=F,size=11,italic=True,color='C00000')
ws.cell(row=e2+5,column=1,value='※ ⚠️ 2022~2023년 일 생산량을 「정상 근무 기준 생산능력」으로 인용하면 안 된다. 야간 작업분이 포함돼 있다.').font=Font(name=F,size=11,italic=True,color='C00000')
ws.cell(row=e2+6,column=1,value='※ 2026년의 8시간 초과일은 야간이 아니라 로더 전담 운전이다 — 「로더 효과 실증」 시트 참조.').font=Font(name=F,size=11,italic=True,color='C00000')
autow(ws,[16,10,10,10,12,12,12,12,10])

# ── 5. 설비별 분업 ──
ws=wb.create_sheet('설비별 분업')
title(ws,'설비별 품목 분업 현황 (2022~2026 누적 수량)',5)
ws.cell(row=2,column=1,value='FG = FAST GRIND · 두 설비가 어떤 품목을 나눠 맡고 있는지 보여준다. 증설 검토 시 어느 설비 부하가 병목인지 판단하는 근거.').font=Font(name=F,size=11,italic=True)
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=5)
eqd=eq[['형상','FG','GX7','합계','FG비중%']]
e=put_df(ws,eqd,hdr_row=3)
ws.cell(row=e+1,column=1,value='합계').font=Font(name=F,bold=True,size=13)
for j,L in [(2,'B'),(3,'C'),(4,'D')]:
    c=ws.cell(row=e+1,column=j,value=f'=SUM({L}4:{L}{e})'); c.font=Font(name=F,bold=True,size=13)
ws.cell(row=e+1,column=5,value=f'=ROUND(B{e+1}/D{e+1}*100,1)').font=Font(name=F,bold=True,size=13)
for c in range(1,6): ws.cell(row=e+1,column=c).fill=PatternFill('solid',fgColor=WARN); ws.cell(row=e+1,column=c).border=B
autow(ws,[16,12,12,12,12])
ws.freeze_panes='A4'


# ── 5-B. 개당 원가 배분 ──
import numpy as _np
_v=_np.repeat(clean['단위시간s'].values, clean['수량'].values.astype(int))
WAVG=round(float(_v.mean()),1)
_day=clean.groupby(['연','월','일'])['수량'].sum().reset_index()
_recent=_day[_day['연']>=2025]['수량'].median()
_shp=clean.groupby('형상').apply(lambda g:_np.median(_np.repeat(g['단위시간s'].values,g['수량'].values.astype(int))),include_groups=False)
_cnt=clean.groupby('형상')['수량'].sum()
_t=pd.DataFrame({'중앙값s':_shp,'누적수량':_cnt}).reset_index().sort_values('누적수량',ascending=False)
_t=_t[_t['누적수량']>=200]

ws=wb.create_sheet('개당 원가 배분')
title(ws,'개당 원가 배분 — 일 생산량 기준',6)
ws.cell(row=2,column=1,value='보정계수가 필요 없는 방식이다. 셋업·대기·검사 시간은 이미 「하루에 몇 개 나왔는가」 안에 녹아 있다.').font=Font(name=F,size=12,italic=True)
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=6)

YEL=PatternFill('solid',fgColor='FFFF00')
def lab(r,t,sz=13,bold=False,color='000000'):
    c=ws.cell(row=r,column=1,value=t); c.font=Font(name=F,size=sz,bold=bold,color=color); return c

lab(4,'■ 입력값',15,True,NAVY)
lab(5,'일 인건비 + 배분경비 (원/일)')
c=ws.cell(row=5,column=3); c.fill=YEL; c.border=B; c.number_format='#,##0'
c.font=Font(name=F,size=13,color='0000FF')
ws.cell(row=5,column=4,value='← 노란 칸에 직접 입력하세요 (대외비 — 위키·저장소에 커밋 금지)').font=Font(name=F,size=11,italic=True,color='C00000')
lab(6,'일 생산량 기준 (개/일)')
c=ws.cell(row=6,column=3,value=int(_recent)); c.fill=YEL; c.border=B; c.font=Font(name=F,size=13,color='0000FF')
ws.cell(row=6,column=4,value='기본값 = 2025~2026년 조업일 수량 중앙값 [실측 검증]. 연도별: 2022년 104.5 / 2023년 69 / 2024년 58.5 / 2025년 58 / 2026년 68').font=Font(name=F,size=11,italic=True)
ws.cell(row=7,column=4,value='⚠️ 2022~2023년 값은 야간 작업분이 포함돼 있어 정상 근무 기준으로 쓰면 개당 원가가 과소평가된다. 2025~2026년을 기준으로 삼는 이유다.').font=Font(name=F,size=11,italic=True,color='C00000')

lab(8,'■ 산출',15,True,NAVY)
lab(9,'개당 평균 배분원가 (원/개)')
c=ws.cell(row=9,column=3,value='=IFERROR(C5/C6,0)'); c.font=Font(name=F,size=13,bold=True); c.border=B; c.number_format='#,##0'
lab(10,'전체 수량가중 평균 사이클 (초/개)')
c=ws.cell(row=10,column=3,value=WAVG); c.font=Font(name=F,size=13); c.border=B; c.number_format='0.0'
ws.cell(row=10,column=4,value='난이도 지수의 기준선. 2022~2026 전체 실적의 수량가중 평균 [실측 검증]').font=Font(name=F,size=11,italic=True)
lab(11,'검산: 일수량 × 평균사이클')
c=ws.cell(row=11,column=3,value='=ROUND(C6*C10/3600,2)'); c.font=Font(name=F,size=13); c.border=B
ws.cell(row=11,column=4,value='시간 — 실제 일 사이클합계 중앙값 4.72시간과 대조하면 내부 정합성이 확인된다').font=Font(name=F,size=11,italic=True)

lab(13,'■ 형상별 배분 (난이도 가중)',15,True,NAVY)
hdrs=['형상','누적 수량','사이클 중앙값(초)','난이도 지수','개당 배분원가(원)']
for j,h in enumerate(hdrs,1): ws.cell(row=14,column=j,value=h)
style_hdr(ws,14,5)
r=15
for _,row in _t.iterrows():
    ws.cell(row=r,column=1,value=row['형상'])
    ws.cell(row=r,column=2,value=int(row['누적수량'])).number_format='#,##0'
    ws.cell(row=r,column=3,value=int(row['중앙값s']))
    ws.cell(row=r,column=4,value=f'=ROUND(C{r}/$C$10,2)')
    ws.cell(row=r,column=5,value=f'=IFERROR(ROUND($C$9*D{r},0),0)').number_format='#,##0'
    for j in range(1,6):
        cl=ws.cell(row=r,column=j); cl.font=Font(name=F,size=13); cl.border=B
        cl.alignment=Alignment(horizontal='left' if j==1 else 'center',vertical='center')
        if r%2==0: cl.fill=PatternFill('solid',fgColor=LGRAY)
    r+=1
ws.cell(row=r+1,column=1,value='※ 난이도 지수 = 그 형상의 사이클 중앙값 ÷ 전체 가중평균. 1.0보다 크면 평균보다 오래 걸리는 품목이다.').font=Font(name=F,size=11,italic=True)
ws.cell(row=r+2,column=1,value='※ 직경·날수까지 나눈 250개 조합의 지수는 「표준 공수 DB」 시트의 난이도지수 열을 보라.').font=Font(name=F,size=11,italic=True)
ws.cell(row=r+3,column=1,value='※ 이 시트는 배분원가(인건비·경비)만 다룬다. 재료비·휠 소모비·외주비는 포함돼 있지 않다.').font=Font(name=F,size=11,italic=True,color='C00000')
autow(ws,[24,14,18,14,18,20])


# ── 5-C. 로더 효과 실증 ──
_e=clean.groupby(['연','월','일','설비']).agg(수량=('수량','sum'),사이클s=('시간합계s','sum')).reset_index()
_e['h']=(_e['사이클s']/3600).round(2)
_d26=clean[clean['연']==2026].groupby(['월','일']).agg(수량=('수량','sum'),사이클s=('시간합계s','sum')).reset_index()
_d26['h']=(_d26['사이클s']/3600).round(2)
_isld=((_d26['월']==5)&(_d26['일']==21))|((_d26['월']==6)&(_d26['일']==23))
_ld=_d26[_isld]; _nm=_d26[~_isld]

ws=wb.create_sheet('로더 효과 실증')
title(ws,'로더 효과 실증 — 1인 다대 운전의 실측 근거',6)
ws.cell(row=2,column=1,value='ANCA 증설 검토의 핵심 전제인 「주간 1인 다대 운전」이 실제로 성립하는지를, 2026년 로더 전담일 2일의 실적으로 검증한다.').font=Font(name=F,size=12,italic=True)
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=6)

def lab2(r,t,sz=13,bold=False,color='000000',col=1):
    c=ws.cell(row=r,column=col,value=t); c.font=Font(name=F,size=sz,bold=bold,color=color); return c

lab2(4,'■ 로더 전담일 설비별 내역',15,True,NAVY)
hh=['일자','설비','수량(개)','설비 사이클(h)','비고']
for j,h in enumerate(hh,1): ws.cell(row=5,column=j,value=h)
style_hdr(ws,5,5)
rows=[('2026-05-21','FAST GRIND',120,5.78,'두 설비를 거의 균등 병렬 운전'),
      ('2026-05-21','GX7',78,5.58,'← 1인 2대 동시 운용의 전형'),
      ('2026-05-21','합계',198,11.36,'정상근무 8h 대비 1.42배'),
      ('2026-06-23','FAST GRIND',139,9.42,'한 대만으로 9.42h — 무인 연속 구간 존재'),
      ('2026-06-23','GX7',26,3.37,''),
      ('2026-06-23','합계',165,12.78,'정상근무 8h 대비 1.60배')]
r=6
for t in rows:
    for j,v in enumerate(t,1):
        c=ws.cell(row=r,column=j,value=v); c.font=Font(name=F,size=13,bold=(t[1]=='합계')); c.border=B
        c.alignment=Alignment(horizontal='left' if j in(1,2,5) else 'center',vertical='center')
        if t[1]=='합계': c.fill=PatternFill('solid',fgColor=WARN)
    r+=1

lab2(r+1,'■ 2026년 로더 전담일 vs 일반일',15,True,NAVY)
hh2=['구분','일수','일 수량','일 설비사이클(h)','8h 대비']
for j,h in enumerate(hh2,1): ws.cell(row=r+2,column=j,value=h)
style_hdr(ws,r+2,5)
cmp=[('로더 전담일 (평균)',len(_ld),round(_ld['수량'].mean(),1),round(_ld['h'].mean(),2),round(_ld['h'].mean()/8,2)),
     ('그 외 (중앙값)',len(_nm),round(_nm['수량'].median(),1),round(_nm['h'].median(),2),round(_nm['h'].median()/8,2))]
rr=r+3
for t in cmp:
    for j,v in enumerate(t,1):
        c=ws.cell(row=rr,column=j,value=v); c.font=Font(name=F,size=13); c.border=B
        c.alignment=Alignment(horizontal='left' if j==1 else 'center',vertical='center')
        if rr==r+3: c.fill=PatternFill('solid',fgColor=OK)
    rr+=1
ws.cell(row=rr,column=1,value='배수').font=Font(name=F,size=13,bold=True)
ws.cell(row=rr,column=3,value=f'=ROUND(C{r+3}/C{r+4},2)').font=Font(name=F,size=13,bold=True)
ws.cell(row=rr,column=4,value=f'=ROUND(D{r+3}/D{r+4},2)').font=Font(name=F,size=13,bold=True)
for j in range(1,6):
    c=ws.cell(row=rr,column=j); c.border=B; c.fill=PatternFill('solid',fgColor=WARN)
    c.alignment=Alignment(horizontal='left' if j==1 else 'center',vertical='center')

n=rr+2
notes=[('■ 판정',15,True,NAVY),
 ('① 「주간 1인 다대 운전」은 실증됐다. 2026-05-21에 1인이 FAST GRIND 5.78h + GX7 5.58h를 병렬로 돌렸다. 두 설비를 거의 균등하게 물린 것으로, 순차 운전이라면 11.36h가 필요해 8시간 근무에 들어가지 않는다.',13,False,'000000'),
 ('② 로더는 근무시간의 벽을 넘게 해준다. 2026-06-23은 FAST GRIND 한 대만으로 9.42h를 기록했다. 8시간 근무 안에서 한 대가 9.42h를 돌려면 무인 연속 운전 구간이 반드시 있어야 한다.',13,False,'000000'),
 ('③ 기존 위키 근거보다 강한 수치다. 현재 위키는 「로더 있음 46.0개/일 vs 없음 18.3개/일」(143 조업일 실측)로 기록돼 있다. 설비 사이클 시간 기준으로 보면 정상근무 대비 1.42~1.60배이며, 2026년 일반일 대비 수량 2.69배·사이클 2.40배다.',13,False,'000000'),
 ('④ ⚠️ 표본은 2일뿐이다. 방향은 분명하나 정밀도는 확인 필요 — 로더 복구 후 전담일을 며칠 더 확보하면 확정할 수 있다.',13,False,'C00000'),
 ('⑤ 🔴 확인 필요 — 2026년에 8h를 넘은 다른 날이 더 있다: 04-27(10.60h) · 04-23(9.55h) · 04-14(9.36h) · 02-20(8.51h) · 04-01(8.33h) · 03-16(8.17h). 이 날들이 로더 사용일인지 잔업인지 미확인. 로더일이라면 표본이 8일로 늘어난다.',13,False,'C00000'),
 ('',13,False,'000000'),
 ('출처: 작업일지 5개년 실적 [신뢰도: 실측 검증] · 로더 전담일 여부는 한경준님 확인(2026-08-28) [신뢰도: 사내 확인]',11,False,'595959')]
for t,sz,bd,col in notes:
    c=ws.cell(row=n,column=1,value=t); c.font=Font(name=F,size=sz,bold=bd,color=col)
    ws.merge_cells(start_row=n,start_column=1,end_row=n,end_column=6)
    c.alignment=Alignment(vertical='center',wrap_text=True)
    if t and not t.startswith('■'): ws.row_dimensions[n].height=34
    n+=1
autow(ws,[26,16,14,18,20,16])

# ── 6. 데이터 품질 ──
ws=wb.create_sheet('데이터 품질·확인필요')
title(ws,'데이터 품질 점검 및 확인 필요 항목',6)
q=[('구분','내용','규모','신뢰도','조치'),
 ('✅ 해소','일자 시트 하단 계기값 행의 자릿수 점프','741건','사내 확인 + 실측 검증','설비 계기 리셋으로 확정(한경준님, 2026-08-28). 2026-01-02에 FG 11035 → 1 로 떨어지는 것이 실측으로 확인됨'),
 ('🔴 확인 필요','그 계기값이 「무엇의」 계기인지','741건','확인 필요','시간으로 해석되지 않는다. 계기 차이 대비 그날 사이클 합계가 2.5배 크다(2024 FG 중앙 251%) — 실가동시간이라면 사이클보다 작을 수 없다. 상관계수는 수량 0.69 > 사이클시간 0.34 로, 시간보다 개수에 가깝게 움직인다. 차이÷수량 중앙값 1.46'),
 ('✅ 해소','원가 보정계수 문제','—','실측 검증','보정계수 없이 푸는 방법으로 전환(한경준님 지적, 2026-08-28). 일 인건비 ÷ 일 생산량으로 개당 배분원가를 내면 셋업·대기·검사가 이미 일 생산량에 반영돼 있다. 「개당 원가 배분」 시트 참조'),
 ('✅ 해소','설비 일 사이클이 근무시간을 넘는 날','51일 (12h 초과)','사내 확인','야간 작업은 2023년까지만 운용(한경준님, 2026-08-28). 2022년 45일 / 2023년 5일이며 2022년 상반기에 집중 — 119 조업일 중 43일(36%). 2024년 이후는 주간 근무만이고, 2026년의 초과일은 로더 전담일이다'),
 ('⚠️ 파급','2022~2023 생산량은 정상 근무 기준이 아니다','2022 상반기 136개/일','실측 검증','야간 작업분 포함. 이 수치를 현재 생산능력이나 증설 검토의 기준선으로 인용하면 안 된다. 원가 배분 기준을 2025~2026년(61개/일)으로 잡은 이유'),
 ('✅ 해소 — 파급 큼','2026년 8h 초과일 2건은 로더 전담 운전','2026-05-21 11.36h / 06-23 12.78h','사내 확인','야간이 아니라 로더로 확정(한경준님, 2026-08-28). 「주간 1인 다대 운전」의 직접 실증 — 「로더 효과 실증」 시트 참조. ANCA 증설 검토에 반영 필요'),
 ('🔴 확인 필요','2026년 8h 초과 나머지 6일의 성격','04-27·04-23·04-14·02-20·04-01·03-16','확인 필요','로더 사용일인지 잔업인지 미확인. 로더일이라면 실증 표본이 2일 → 8일로 늘어난다'),
 ('✅ 해소','단위시간 1초 입력','511건 (3.89%)','사내 확인','1초 단위까지 입력해 본 것으로 확인(한경준님). 전량 정확히 1.0초 단일값이므로 집계에서 제외했으나, 중앙값에는 영향이 없다(볼 Ø6 2날 제외·포함 모두 240초)'),
 ('✅ 해소','NC드릴 vs 드릴 통합 여부','2,345 / 3,884개','사내 확인','서로 다른 가공 파일을 쓰고 형상도 다름 → 별도 항목 유지 확정(한경준님, 2026-08-28)'),
 ('✅ 반영','각도 표기(면취_90 등)','29건','사내 확인','중요도 낮음으로 확인 — DB에서 제외(한경준님, 2026-08-28)'),
 ('표기 흔들림','형상 원표기','124종 → 23종 정규화','실측 검증','평=스퀘어, 면취=챔퍼 통합(한경준님 확인). HSS는 소재 열로 분리'),
 ('표기 흔들림','코팅 원표기','17종 → 12종 정규화','실측 검증','비=비코팅, ts=TS, 틴/티진=TiN 통합'),
 ('🔴 파서 결함','daily_report.py 고정 행범위','2022년 759건 누락','실측 검증','FG를 행 2~16으로 하드코딩하나 일부 시트는 15슬롯 초과. 블록 마커 동적 탐지로 수정 필요'),
 ('결측','작업자 열','13,145건 중 10,731건 공란','실측 검증','2022년만 대체로 기재 — 작업자별 분석 불가'),
 ('불일치','시간합계 ≠ 수량×단위시간','14건 (0.1%)','실측 검증','무시 가능'),
 ('⚠️ 보안','원본 xls의 가공비·금액 열','2024년 이후 실거래 단가','실측 검증','본 DB에는 미포함. 단 원본 xls 80개가 git 추적 중 — 저장소 공개 여부 재확인 필요'),
]
r=3
for i,row in enumerate(q):
    for j,v in enumerate(row,1):
        c=ws.cell(row=r,column=j,value=v)
        c.font=Font(name=F,size=13,bold=(i==0),color=W if i==0 else '000000')
        c.border=B; c.alignment=Alignment(vertical='center',wrap_text=True,horizontal='left')
        if i==0: c.fill=PatternFill('solid',fgColor=HDRB)
        elif str(row[0]).startswith('🔴'): c.fill=PatternFill('solid',fgColor=BAD)
        elif str(row[0]).startswith('⚠️'): c.fill=PatternFill('solid',fgColor=WARN)
    ws.row_dimensions[r].height=60 if i>0 else 30
    r+=1
autow(ws,[16,26,20,16,72])

# ── 6-B. 설비 계기값 ──
_mp = DATA / 'meter_all.csv'
mt = pd.read_csv(_mp) if _mp.exists() else None
if mt is not None:
    ws=wb.create_sheet('설비 계기값')
    title(ws,'설비 계기값 (참고 — 해석 미확정)',7)
    ws.cell(row=2,column=1,value='일자 시트 하단 행에서 추출. 리셋이 확인됐으나(2026-01-02 FG 11035→1) 시간 단위로 해석되지 않아 공수 DB에는 반영하지 않았다. 2022~2023년은 컬럼 배치 자체가 다르다.').font=Font(name=F,size=11,italic=True)
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=7)
    mt2=mt.copy(); mt2['검산(B-A)']=(mt2['계기B']-mt2['계기A']).round(1)
    put_df(ws,mt2,hdr_row=3)
    autow(ws,[8,7,7,8,12,12,12,12]); ws.freeze_panes='A4'

# ── 7. 원천 데이터 ──
ws=wb.create_sheet('원천데이터')
raw=clean[['연','월','일','설비','형상','직경','날수','소재','코팅_정규','라핑','수량','단위시간s','시간합계s','형상_원본','코팅_원본','특이사항','완료여부']]
raw=raw.rename(columns={'코팅_정규':'코팅','단위시간s':'사이클초','시간합계s':'시간합계초'})
for j,col in enumerate(raw.columns,1):
    c=ws.cell(row=1,column=j,value=col)
style_hdr(ws,1,len(raw.columns))
for i,r_ in enumerate(raw.itertuples(index=False),2):
    for j,v in enumerate(r_,1):
        if isinstance(v,(np.integer,)): v=int(v)
        elif isinstance(v,(np.floating,)): v=float(v) if not pd.isna(v) else None
        elif isinstance(v,(np.bool_,)): v=bool(v)
        elif pd.isna(v): v=None
        ws.cell(row=i,column=j,value=v)
ws.freeze_panes='A2'
autow(ws,[7,6,6,7,14,8,7,8,10,7,8,10,11,16,11,16,10])
ws.cell(row=1,column=len(raw.columns)+2,value='이상치(≤1초) 511건 제외 후 %d건'%len(raw)).font=Font(name=F,size=11,italic=True)

OUTF.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTF)
print(f'저장 완료: {OUTF}')
print(f'  DB 조합 {len(db)} · 형상 {len(shp)} · 원천 {len(raw)}행')
