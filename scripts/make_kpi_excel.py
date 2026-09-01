import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd, warnings
warnings.filterwarnings('ignore')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# ── 스타일 상수 ────────────────────────────────────────────
FONT      = "맑은 고딕"
C_HDR_BG  = "1F497D"; C_HDR_FG  = "FFFFFF"
C_SUB_BG  = "4472C4"; C_SUB_FG  = "FFFFFF"
C_FG_BG   = "D6E4F7"   # FAST GRIND 연파랑
C_GX_BG   = "E2EFDA"   # GX7 연초록
C_TOT_BG  = "FFF2CC"   # 합계 연노랑
C_WARN_BG = "FFE0E0"   # 경고 (낮은 달) 연빨강
C_OK_BG   = "E2EFDA"   # 양호 연초록
C_GRAY    = "D9D9D9"
C_ASM_BG  = "FFF2CC"   # 가정(assumption) 셀 노란
PCT_FMT   = '0.0%'
NUM_FMT   = '#,##0'

thin   = Side(style='thin', color='BBBBBB')
med    = Side(style='medium', color='888888')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MBORDER= Border(left=med, right=med, top=med, bottom=med)

def s(cell, bold=False, fg=None, bg=None, halign='center', valign='center',
      wrap=False, sz=10, border=True):
    cell.font      = Font(name=FONT, bold=bold, color=fg or "000000", size=sz)
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if border: cell.border = BORDER

def hdr(cell, txt=None, sub=False, center=True):
    if txt is not None: cell.value = txt
    cell.font      = Font(name=FONT, bold=True, color=C_SUB_FG if sub else C_HDR_FG, size=10)
    cell.fill      = PatternFill("solid", fgColor=C_SUB_BG if sub else C_HDR_BG)
    cell.alignment = Alignment(horizontal='center' if center else 'left', vertical='center', wrap_text=True)
    cell.border    = BORDER

def tot(cell):
    cell.font      = Font(name=FONT, bold=True, size=10)
    cell.fill      = PatternFill("solid", fgColor=C_GRAY)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = BORDER

def title_row(ws, row, text, ncols, sz=13):
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name=FONT, bold=True, size=sz, color=C_HDR_FG)
    c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 22

MONTHS  = list(range(1, 13))
MO_KR   = [f'{m}월' for m in MONTHS]
BIZNOTE = {6: '* GX7 고장(3~22일)', 8: '* 생산량 누락', 10: '* 추석(3~9일)'}

# ── 원본 데이터 로드 ─────────────────────────────────────
src = 'raw/출하현황/25년 월별 생산 실적.xlsx'
raw = pd.read_excel(src, sheet_name='25', header=None)
rows = raw.iloc[3:15].copy().reset_index(drop=True)   # 1~12월

data = []
for i, r in rows.iterrows():
    data.append({
        '월':     int(r[1]),
        '근무일':  int(r[2]),
        '근무시간': int(r[2]) * 28800,  # 근무일수 × 28,800초(8h×60m×60s)
        'FG수량':  int(r[4]),
        'FG시간':  int(r[5]),    # 초
        'GX수량':  int(r[6]),
        'GX시간':  int(r[7]),    # 초
        '합수량':  int(r[8]),
        '합시간':  int(r[9]),    # 초
        'FG표기':  r[10],
        'GX표기':  r[11],
        '출하수량': int(r[12]),
        '불량':    int(r[13]),
    })

# 표준생산량 (row18)
avg_row   = raw.iloc[18]
FG_STD    = float(avg_row[4])    # 927.25
GX7_STD   = float(avg_row[6])    # 533.9167

wb = Workbook()

# ══════════════════════════════════════════════════════════════════
# 1. 종합 대시보드
# ══════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "종합 대시보드"
ws.freeze_panes = "B4"
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 24

title_row(ws, 1, "2025년 KPI 종합 대시보드  —  설비가동율 · 작업효율 · 양품율", 11)

# 가정 셀 (표준생산량)
ws['A2'] = "【가정】 FG 표준생산량"
s(ws['A2'], bold=True, halign='left', sz=9)
ws.merge_cells('A2:C2')
s(ws.cell(2,2)); s(ws.cell(2,3))
ws['D2'] = FG_STD
ws['D2'].number_format = '#,##0.00'
ws['D2'].fill = PatternFill("solid", fgColor=C_ASM_BG)
ws['D2'].font = Font(name=FONT, bold=True, color="0000FF", size=10)
ws['D2'].border = BORDER; ws['D2'].alignment = Alignment(horizontal='center', vertical='center')

ws['E2'] = "GX7 표준생산량"
s(ws['E2'], bold=True, halign='left', sz=9)
ws.merge_cells('E2:G2')
s(ws.cell(2,6)); s(ws.cell(2,7))
ws['H2'] = GX7_STD
ws['H2'].number_format = '#,##0.00'
ws['H2'].fill = PatternFill("solid", fgColor=C_ASM_BG)
ws['H2'].font = Font(name=FONT, bold=True, color="0000FF", size=10)
ws['H2'].border = BORDER; ws['H2'].alignment = Alignment(horizontal='center', vertical='center')

ws['I2'] = "※ 표준생산량 = 25년+26년1~3월 15개월 월평균 (출처: 25년 월별 생산 실적.xlsx)"
ws['I2'].font = Font(name=FONT, size=8, color="666666", italic=True)
ws.merge_cells('I2:K2')

# 헤더
hdrs = ['월', '근무일(일)', '설비가동율\nFAST', '설비가동율\nGX7', '설비가동율\n합계',
        '작업효율\nFAST', '작업효율\nGX7', '양품율', '출하량\n(개)', '불량\n(개)', '비고']
for ci, h in enumerate(hdrs, 1):
    hdr(ws.cell(3, ci, value=h))

# 데이터
for i, d in enumerate(data):
    r = i + 4
    m = d['월']
    ws.row_dimensions[r].height = 18

    # 설비가동율
    fg_util  = f'=C{r}*2/근무시간'   # 실제론 수식으로
    # 실제 수식으로 작성
    fg_util_val  = d['FG시간'] / d['근무시간']
    gx_util_val  = d['GX시간'] / d['근무시간']
    sum_util_val = fg_util_val + gx_util_val
    fg_eff_val   = d['FG수량'] / FG_STD
    gx_eff_val   = d['GX수량'] / GX7_STD
    yp_val       = (d['출하수량'] - d['불량']) / d['출하수량']

    bg_util = C_WARN_BG if sum_util_val < 0.50 else (C_OK_BG if sum_util_val >= 0.65 else None)
    bg_eff  = C_WARN_BG if min(fg_eff_val, gx_eff_val) < 0.65 else (C_OK_BG if min(fg_eff_val, gx_eff_val) >= 0.80 else None)
    bg_yp   = C_WARN_BG if yp_val < 0.985 else C_OK_BG

    vals = [
        (m,            None,    None),
        (d['근무일'],   NUM_FMT, None),
        (fg_util_val,  PCT_FMT, bg_util),
        (gx_util_val,  PCT_FMT, bg_util),
        (sum_util_val, PCT_FMT, bg_util),
        (fg_eff_val,   PCT_FMT, bg_eff),
        (gx_eff_val,   PCT_FMT, bg_eff),
        (yp_val,       PCT_FMT, bg_yp),
        (d['출하수량'], NUM_FMT, None),
        (d['불량'],     NUM_FMT, None),
        (BIZNOTE.get(m, ''), None, None),
    ]
    for ci, (val, fmt, bg) in enumerate(vals, 1):
        c = ws.cell(r, ci, value=val)
        s(c, bg=bg, halign='left' if ci in (1,11) else 'center')
        if fmt: c.number_format = fmt
        if ci == 1: c.font = Font(name=FONT, bold=True, size=10)

# 합계/평균 행
tot_r = len(data) + 4
ws.merge_cells(f'A{tot_r}:B{tot_r}')
tot(ws.cell(tot_r, 1, value='연간 평균')); tot(ws.cell(tot_r, 2))

for ci, col_letter in enumerate(['C','D','E','F','G','H'], 3):
    c = ws.cell(tot_r, ci, value=f'=AVERAGE({col_letter}4:{col_letter}{tot_r-1})')
    tot(c); c.number_format = PCT_FMT

c = ws.cell(tot_r, 9, value=f'=SUM(I4:I{tot_r-1})')
tot(c); c.number_format = NUM_FMT
c = ws.cell(tot_r, 10, value=f'=SUM(J4:J{tot_r-1})')
tot(c); c.number_format = NUM_FMT
tot(ws.cell(tot_r, 11, value=''))

# 열 너비
widths = [6, 8, 11, 11, 11, 11, 11, 11, 10, 8, 22]
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

# ══════════════════════════════════════════════════════════════════
# 2. 설비가동율 상세
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("설비가동율")
ws2.freeze_panes = "B4"
title_row(ws2, 1, "설비가동율 상세  —  총 가공시간(초) ÷ 근무시간(초)", 10)

# 범례 설명
ws2['A2'] = "▶ 설비가동율 = 기계 실제 가공시간 ÷ 근무가능시간  |  합계 = FAST + GX7 가동율 합산"
ws2['A2'].font = Font(name=FONT, size=9, color="444444", italic=True)
ws2.merge_cells('A2:J2')
ws2.row_dimensions[2].height = 16

hdrs2_top = ['월', '근무일\n(일)', '근무시간\n(초)', 'FAST GRIND', '', '', 'GX7', '', '', '합계 가동율']
for ci, h in enumerate(hdrs2_top, 1):
    c = ws2.cell(3, ci, value=h)
    if ci in (4,5,6): hdr(c, sub=True)
    elif ci in (7,8,9): hdr(c, sub=True, txt=h)
    else: hdr(c)

# 서브헤더 병합 (3행은 그대로, 별도 헤더로)
for ci, h in enumerate(['월','근무일(일)','근무시간(초)','생산수량(개)','가공시간(초)','가동율(%)','생산수량(개)','가공시간(초)','가동율(%)','가동율 합계(%)'], 1):
    c = ws2.cell(3, ci, value=h)
    if ci in (4,5,6): hdr(c, sub=True)
    elif ci in (7,8,9): hdr(c, sub=True)
    else: hdr(c)

for i, d in enumerate(data):
    r = i + 4
    m = d['월']
    ws2.row_dimensions[r].height = 18
    fg_u = d['FG시간'] / d['근무시간']
    gx_u = d['GX시간'] / d['근무시간']
    su   = fg_u + gx_u
    bg_s = C_WARN_BG if su < 0.50 else (C_OK_BG if su >= 0.65 else None)

    row_vals = [
        (m,           None,    None),
        (d['근무일'],  NUM_FMT, None),
        (d['근무시간'],NUM_FMT, None),
        (d['FG수량'],  NUM_FMT, C_FG_BG),
        (d['FG시간'],  NUM_FMT, C_FG_BG),
        (fg_u,         PCT_FMT, C_FG_BG),
        (d['GX수량'],  NUM_FMT, C_GX_BG),
        (d['GX시간'],  NUM_FMT, C_GX_BG),
        (gx_u,         PCT_FMT, C_GX_BG),
        (su,           PCT_FMT, bg_s),
    ]
    for ci, (val, fmt, bg) in enumerate(row_vals, 1):
        c = ws2.cell(r, ci, value=val)
        s(c, bg=bg, bold=(ci==1))
        if fmt: c.number_format = fmt

tot_r2 = len(data) + 4
ws2.merge_cells(f'A{tot_r2}:B{tot_r2}')
tot(ws2.cell(tot_r2, 1, value='연간 평균')); tot(ws2.cell(tot_r2, 2))
for ci, col in enumerate(['C','D','E','F','G','H','I','J'], 3):
    c = ws2.cell(tot_r2, ci)
    if col in ['E','F','H','I']:
        c.value = f'=AVERAGE({col}4:{col}{tot_r2-1})'
        c.number_format = PCT_FMT if col in ['F','I','J'] else NUM_FMT
    elif col == 'J':
        c.value = f'=AVERAGE(J4:J{tot_r2-1})'
        c.number_format = PCT_FMT
    else:
        c.value = f'=SUM({col}4:{col}{tot_r2-1})'
        c.number_format = NUM_FMT
    tot(c)

c = ws2.cell(tot_r2, 10, value=f'=AVERAGE(J4:J{tot_r2-1})')
tot(c); c.number_format = PCT_FMT

widths2 = [6,8,13,11,13,11,11,13,11,12]
for ci, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

# ══════════════════════════════════════════════════════════════════
# 3. 작업효율 상세
# ══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("작업효율")
ws3.freeze_panes = "B5"
title_row(ws3, 1, "작업효율 상세  —  실생산량 ÷ 표준생산량", 9)

ws3['A2'] = "▶ 작업효율 = 월 실제 생산수량 ÷ 표준생산량(15개월 월평균)  |  100% 초과 = 표준 대비 초과 생산"
ws3['A2'].font = Font(name=FONT, size=9, color="444444", italic=True)
ws3.merge_cells('A2:I2')
ws3.row_dimensions[2].height = 16

# 가정 셀
ws3['A3'] = "FG 표준생산량"
s(ws3['A3'], bold=True, halign='center', sz=9)
ws3['B3'] = FG_STD
ws3['B3'].fill = PatternFill("solid", fgColor=C_ASM_BG)
ws3['B3'].font = Font(name=FONT, bold=True, color="0000FF", size=10)
ws3['B3'].number_format = '#,##0.00'
ws3['B3'].border = BORDER; ws3['B3'].alignment = Alignment(horizontal='center', vertical='center')
ws3['B3'].comment = None
ws3['C3'] = "개/월"
s(ws3['C3'], sz=9)
ws3['D3'] = "GX7 표준생산량"
s(ws3['D3'], bold=True, halign='center', sz=9)
ws3['E3'] = GX7_STD
ws3['E3'].fill = PatternFill("solid", fgColor=C_ASM_BG)
ws3['E3'].font = Font(name=FONT, bold=True, color="0000FF", size=10)
ws3['E3'].number_format = '#,##0.00'
ws3['E3'].border = BORDER; ws3['E3'].alignment = Alignment(horizontal='center', vertical='center')
ws3['F3'] = "개/월"
s(ws3['F3'], sz=9)
ws3['G3'] = "출처: 25년 월별 생산 실적.xlsx 15개월 평균"
ws3['G3'].font = Font(name=FONT, size=8, color="666666", italic=True)
ws3.merge_cells('G3:I3')
ws3.row_dimensions[3].height = 18

for ci, h in enumerate(['월','근무일(일)','FG 생산수량(개)','FG 표준수량(개)','FG 작업효율(%)','GX7 생산수량(개)','GX7 표준수량(개)','GX7 작업효율(%)','비고'], 1):
    c = ws3.cell(4, ci, value=h)
    if ci in (3,4,5): hdr(c, sub=True)
    elif ci in (6,7,8): hdr(c, sub=True)
    else: hdr(c)

for i, d in enumerate(data):
    r = i + 5
    m = d['월']
    ws3.row_dimensions[r].height = 18
    fe = d['FG수량'] / FG_STD
    ge = d['GX수량'] / GX7_STD
    bg_f = C_WARN_BG if fe < 0.70 else (C_OK_BG if fe >= 0.90 else None)
    bg_g = C_WARN_BG if ge < 0.70 else (C_OK_BG if ge >= 0.90 else None)

    row3_vals = [
        (m,          None,    None),
        (d['근무일'], NUM_FMT, None),
        (d['FG수량'], NUM_FMT, C_FG_BG),
        (FG_STD,      '#,##0.00', C_FG_BG),
        (fe,           PCT_FMT, bg_f),
        (d['GX수량'], NUM_FMT, C_GX_BG),
        (GX7_STD,     '#,##0.00', C_GX_BG),
        (ge,           PCT_FMT, bg_g),
        (BIZNOTE.get(m, ''), None, None),
    ]
    for ci, (val, fmt, bg) in enumerate(row3_vals, 1):
        c = ws3.cell(r, ci, value=val)
        s(c, bg=bg, bold=(ci==1), halign='left' if ci in (1,9) else 'center')
        if fmt: c.number_format = fmt

tot_r3 = len(data) + 5
ws3.merge_cells(f'A{tot_r3}:B{tot_r3}')
tot(ws3.cell(tot_r3, 1, value='연간 평균')); tot(ws3.cell(tot_r3, 2))
c = ws3.cell(tot_r3, 3, value=f'=AVERAGE(C5:C{tot_r3-1})')
tot(c); c.number_format = NUM_FMT
c = ws3.cell(tot_r3, 4, value=FG_STD)
tot(c); c.number_format = '#,##0.00'
c = ws3.cell(tot_r3, 5, value=f'=AVERAGE(E5:E{tot_r3-1})')
tot(c); c.number_format = PCT_FMT
c = ws3.cell(tot_r3, 6, value=f'=AVERAGE(F5:F{tot_r3-1})')
tot(c); c.number_format = NUM_FMT
c = ws3.cell(tot_r3, 7, value=GX7_STD)
tot(c); c.number_format = '#,##0.00'
c = ws3.cell(tot_r3, 8, value=f'=AVERAGE(H5:H{tot_r3-1})')
tot(c); c.number_format = PCT_FMT
tot(ws3.cell(tot_r3, 9, value=''))

widths3 = [6,8,14,13,13,14,13,13,22]
for ci, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

# ══════════════════════════════════════════════════════════════════
# 4. 양품율 상세
# ══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("양품율")
ws4.freeze_panes = "B4"
title_row(ws4, 1, "양품율 상세  —  (출하수량 - 불량수) ÷ 출하수량", 8)

ws4['A2'] = "▶ 양품율 = (출하수량 - 불량수) ÷ 출하수량  |  99% 이상 = 정상, 98.5% 미만 = 주의"
ws4['A2'].font = Font(name=FONT, size=9, color="444444", italic=True)
ws4.merge_cells('A2:H2')
ws4.row_dimensions[2].height = 16

for ci, h in enumerate(['월','근무일(일)','출하수량(개)','불량수(개)','양품수량(개)','양품율(%)','불량율(%)','비고'], 1):
    hdr(ws4.cell(3, ci, value=h))

for i, d in enumerate(data):
    r = i + 4
    m = d['월']
    ws4.row_dimensions[r].height = 18
    yp = (d['출하수량'] - d['불량']) / d['출하수량']
    bg = C_WARN_BG if yp < 0.985 else C_OK_BG

    row4_vals = [
        (m,                     None,    None),
        (d['근무일'],            NUM_FMT, None),
        (d['출하수량'],          NUM_FMT, None),
        (d['불량'],              NUM_FMT, C_WARN_BG if d['불량'] > 15 else None),
        (d['출하수량']-d['불량'], NUM_FMT, None),
        (yp,                    PCT_FMT, bg),
        (1 - yp,                PCT_FMT, C_WARN_BG if (1-yp) > 0.015 else None),
        (BIZNOTE.get(m, ''),    None,    None),
    ]
    for ci, (val, fmt, bgc) in enumerate(row4_vals, 1):
        c = ws4.cell(r, ci, value=val)
        s(c, bg=bgc, bold=(ci==1), halign='left' if ci in (1,8) else 'center')
        if fmt: c.number_format = fmt

tot_r4 = len(data) + 4
ws4.merge_cells(f'A{tot_r4}:B{tot_r4}')
tot(ws4.cell(tot_r4, 1, value='연간 합계/평균')); tot(ws4.cell(tot_r4, 2))
c = ws4.cell(tot_r4, 3, value=f'=SUM(C4:C{tot_r4-1})')
tot(c); c.number_format = NUM_FMT
c = ws4.cell(tot_r4, 4, value=f'=SUM(D4:D{tot_r4-1})')
tot(c); c.number_format = NUM_FMT
c = ws4.cell(tot_r4, 5, value=f'=SUM(E4:E{tot_r4-1})')
tot(c); c.number_format = NUM_FMT
c = ws4.cell(tot_r4, 6, value=f'=(C{tot_r4}-D{tot_r4})/C{tot_r4}')
tot(c); c.number_format = PCT_FMT
c = ws4.cell(tot_r4, 7, value=f'=D{tot_r4}/C{tot_r4}')
tot(c); c.number_format = PCT_FMT
tot(ws4.cell(tot_r4, 8, value=''))

widths4 = [6,8,12,10,12,11,10,22]
for ci, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(ci)].width = w

# ══════════════════════════════════════════════════════════════════
# 5. 원본 데이터
# ══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("원본 데이터")
title_row(ws5, 1, "원본 데이터 (25년 월별 생산 실적.xlsx — '25' 시트)", 13)
ws5['A2'] = "※ 이 시트는 참조 전용입니다. 원본 파일(raw/)은 수정하지 않습니다."
ws5['A2'].font = Font(name=FONT, size=9, color="CC0000", italic=True)
ws5.merge_cells('A2:M2')

raw_hdrs = ['월','근무일(일)','근무시간(초)','FG 생산수량','FG 가공시간(초)','GX7 생산수량',
            'GX7 가공시간(초)','합계 생산수량','합계 가공시간(초)','FG 장비표기','GX7 장비표기',
            '출하수량','불량수']
for ci, h in enumerate(raw_hdrs, 1):
    hdr(ws5.cell(3, ci, value=h))

for i, d in enumerate(data):
    r = i + 4
    row_vals = [d['월'],d['근무일'],d['근무시간'],d['FG수량'],d['FG시간'],
                d['GX수량'],d['GX시간'],d['합수량'],d['합시간'],
                d['FG표기'],d['GX표기'],d['출하수량'],d['불량']]
    for ci, val in enumerate(row_vals, 1):
        c = ws5.cell(r, ci, value=val)
        s(c, bold=(ci==1)); c.number_format = NUM_FMT if ci > 1 else 'General'

for ci in [1,2,10,11,12,13]:
    ws5.column_dimensions[get_column_letter(ci)].width = 8
for ci in [3,5,7,9]:
    ws5.column_dimensions[get_column_letter(ci)].width = 14
for ci in [4,6,8]:
    ws5.column_dimensions[get_column_letter(ci)].width = 12

out = 'wiki/comparisons/2025_KPI_설비가동율_작업효율_양품율.xlsx'
wb.save(out)
print(f"저장 완료: {out}")
