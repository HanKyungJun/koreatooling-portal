import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd, warnings, os
warnings.filterwarnings('ignore')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 스타일 ─────────────────────────────────────────────────
FONT   = "맑은 고딕"
C_H1   = "1F497D";  C_H1F = "FFFFFF"
C_H2   = "2E75B6";  C_H2F = "FFFFFF"
C_FG   = "D6E4F7"   # FAST GRIND
C_GX   = "E2EFDA"   # GX7
C_TOT  = "FFF2CC"   # 합계/평균
C_WARN = "FFE0E0"   # 경고
C_OK   = "E2EFDA"   # 양호
C_NA   = "F2F2F2"   # 데이터없음
C_GRAY = "D9D9D9"
C_BLUE = "0000FF"
C_ASM  = "FFFACD"

thin  = Side(style='thin',   color='BBBBBB')
med   = Side(style='medium', color='888888')
BD    = Border(left=thin, right=thin, top=thin, bottom=thin)
MBD   = Border(left=med,  right=med,  top=med,  bottom=med)

def cs(cell, bold=False, fg=None, bg=None, ha='center', va='center',
       wrap=False, sz=10, bd=True, italic=False):
    cell.font      = Font(name=FONT, bold=bold, color=fg or "000000", size=sz, italic=italic)
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if bd: cell.border = BD

def hdr(cell, v=None, sub=False):
    if v is not None: cell.value = v
    cell.font      = Font(name=FONT, bold=True, color=C_H2F if sub else C_H1F, size=10)
    cell.fill      = PatternFill("solid", fgColor=C_H2 if sub else C_H1)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = BD

def totc(cell, v=None):
    if v is not None: cell.value = v
    cell.font      = Font(name=FONT, bold=True, size=10)
    cell.fill      = PatternFill("solid", fgColor=C_GRAY)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = BD

def title_row(ws, row, text, ncols, sz=13, bg=C_H1):
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, bold=True, size=sz, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 22

def note_row(ws, row, text, ncols, fg="444444"):
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=9, color=fg, italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 15

PCT = '0.0%'; NUM = '#,##0'; DASH = '#,##0;-#,##0;"-"'

# ══════════════════════════════════════════════════════════
#  데이터 수집
# ══════════════════════════════════════════════════════════
# 표준생산량 (2025 참조파일 row18)
ref25 = pd.read_excel('raw/출하현황/25년 월별 생산 실적.xlsx', sheet_name='25', header=None)
FG_STD  = float(ref25.iloc[18][4])   # 927.25
GX_STD  = float(ref25.iloc[18][6])   # 533.9167

# 2024년 근무일수 (참조파일)
ref24 = pd.read_excel('raw/출하현황/25년 월별 생산 실적.xlsx', sheet_name='24', header=None)
biz24 = {int(ref24.iloc[i][1]): int(ref24.iloc[i][2]) for i in range(4, 16)}

# 2025년 데이터 (참조파일)
def load_2025():
    rows = ref25.iloc[3:15].copy().reset_index(drop=True)
    result = []
    BIZNOTE = {6:'* GX7 고장(3~22일)', 8:'* 생산량 누락', 10:'* 추석(3~9일)'}
    for i, r in rows.iterrows():
        m = int(r[1])
        result.append({
            '연도': 2025, '월': m,
            '근무일': int(r[2]), '근무시간': int(r[2]) * 28800,  # 근무일수 × 28,800초
            'FG수량': int(r[4]),  'FG시간': int(r[5]),   'FG시간유효': True,
            'GX수량': int(r[6]),  'GX시간': int(r[7]),   'GX시간유효': True,
            '합수량': int(r[8]),  '합시간': int(r[9]),
            '출하수량': int(r[12]), '불량': int(r[13]),
            '양품율유효': True,
            '비고': BIZNOTE.get(m, ''),
        })
    return result

# 2022/2023/2024 출하현황에서 월별 출하량 · 불량 추출
def load_shipping(year):
    fp = f'raw/출하현황/{str(year)[2:]}년_출하현황.xlsx'
    if not os.path.exists(fp):
        return {}
    df = pd.read_excel(fp, sheet_name='출하현황', header=0)
    df['출하일자'] = pd.to_datetime(df['출하일자'], errors='coerce')
    df['월'] = df['출하일자'].dt.month
    df['연마불가'] = pd.to_numeric(df['연마불가'], errors='coerce').fillna(0)
    grp = df.groupby('월').agg(출하수량=('출고','sum'), 불량=('연마불가','sum'))
    return {int(m): (int(row['출하수량']), int(row['불량'])) for m, row in grp.iterrows()}

shipping = {yr: load_shipping(yr) for yr in [2022, 2023, 2024]}

# 2026 출하현황: 기본 파일 + 월별 추가 파일 병합
def load_shipping_2026():
    import glob as _glob
    frames = []
    base = 'raw/출하현황/26년_출하현황.xlsx'
    if os.path.exists(base):
        frames.append(pd.read_excel(base, sheet_name='출하현황', header=0))
    for extra in sorted(_glob.glob('raw/출하현황/26년_출하현황_*.xlsx')):
        frames.append(pd.read_excel(extra, sheet_name='출하현황', header=0))
    if not frames:
        return {}
    df = pd.concat(frames, ignore_index=True)
    df['출하일자'] = pd.to_datetime(df['출하일자'], errors='coerce')
    df['월'] = df['출하일자'].dt.month
    df['연마불가'] = pd.to_numeric(df['연마불가'], errors='coerce').fillna(0)
    grp = df.groupby('월').agg(출하수량=('출고','sum'), 불량=('연마불가','sum'))
    return {int(m): (int(row['출하수량']), int(row['불량'])) for m, row in grp.iterrows()}

shipping[2026] = load_shipping_2026()

# 2022/2023/2024 작업일지에서 추출
def load_year(year, biz_days_override=None):
    base = f'raw/출하현황/재연마 작업일지({year})'
    ship = shipping.get(year, {})
    result = []
    for m in range(1, 13):
        fp = f'{base}/재연마_월간생산일지 ({m}월).xls'
        if not os.path.exists(fp): continue
        df = pd.read_excel(fp, sheet_name='월간합계표', header=None)
        tot = df[df[0] == '합계']
        if tot.empty: continue
        t = tot.iloc[0]

        fg_qty  = int(t[1])  if pd.notna(t[1])  else 0
        gx_qty  = int(t[7])  if pd.notna(t[7])  else 0
        tot_qty = int(t[13]) if pd.notna(t[13]) else 0

        # 2022년 1~10월: 가공시간 데이터 신뢰 불가 → N/A 처리
        if year == 2022 and m <= 10:
            fg_time = 0
            gx_time = 0
        else:
            fg_time = int(t[3]) if pd.notna(t[3]) else 0
            gx_time = int(t[9]) if pd.notna(t[9]) else 0

        # 근무일수
        if biz_days_override and m in biz_days_override:
            biz = biz_days_override[m]
        else:
            biz = int((df.iloc[2:33][13].replace(0, pd.NA).dropna()).count())
        biz_sec = biz * 28800  # 근무일수 × 28,800초(8h×60m×60s)

        # 가동시간 유효성: 2022년 1~10월은 원본 기록 오류로 N/A
        if year == 2022 and m <= 10:
            fg_ok = False
            gx_ok = False
        else:
            fg_ok = fg_time > 0 and fg_time < biz_sec
            gx_ok = gx_time > 0 and gx_time < biz_sec

        # 출하현황 데이터 병합
        if m in ship:
            ship_qty, ship_bad = ship[m]
            yp_valid = True
        else:
            ship_qty, ship_bad = None, None
            yp_valid = False

        result.append({
            '연도': year, '월': m,
            '근무일': biz, '근무시간': biz_sec,
            'FG수량': fg_qty,  'FG시간': fg_time if fg_ok else None,  'FG시간유효': fg_ok,
            'GX수량': gx_qty,  'GX시간': gx_time if gx_ok else None,  'GX시간유효': gx_ok,
            '합수량': tot_qty, '합시간': (fg_time + gx_time) if (fg_ok and gx_ok) else None,
            '출하수량': ship_qty, '불량': ship_bad,
            '양품율유효': yp_valid,
            '비고': '',
        })
    return result

all_data = {
    2022: load_year(2022),
    2023: load_year(2023),
    2024: load_year(2024, biz_days_override=biz24),
    2025: load_2025(),
    2026: load_year(2026),
}

# ══════════════════════════════════════════════════════════
#  워크북 생성
# ══════════════════════════════════════════════════════════
wb = Workbook()

# ──────────────────────────────────────────────────────────
# 시트1: 연도별 종합 비교
# ──────────────────────────────────────────────────────────
ws0 = wb.active
ws0.title = "연도별 종합 비교"
ws0.freeze_panes = "C4"

title_row(ws0, 1, "2022~2025년 재연마 KPI 연도별 종합 비교  —  설비가동율 · 작업효율 · 양품율", 17)
note_row(ws0,  2, "▶ 설비가동율 = 가공시간 ÷ 근무시간  |  작업효율 = 실생산량 ÷ 표준생산량(FG:927.25, GX7:533.92)  |  양품율 = (출하량-불량) ÷ 출하량", 17)
note_row(ws0,  3, "⚠️ 2022년 1~10월 가공시간 원본 기록 오류 → 설비가동율 N/A  |  2022년 출하현황 5~12월만 존재 → 1~4월 양품율 N/A", 17, fg="CC0000")

# 헤더
hdrs = ['연도','월','근무일(일)',
        'FG가동율','GX7가동율','합계가동율',
        'FG작업효율','GX7작업효율',
        '양품율',
        'FG생산','GX7생산','합계생산',
        'FG시간(초)','GX7시간(초)',
        '출하량','불량','비고']
for ci, h in enumerate(hdrs, 1):
    hdr(ws0.cell(4, ci, value=h),
        sub=(ci in [4,5,6,7,8]))
ws0.row_dimensions[4].height = 28

yr_colors = {2022:'FFF0F5', 2023:'F0FFF0', 2024:'F0F8FF', 2025:'FFFFF0'}

for yr in [2022, 2023, 2024, 2025]:
    for d in all_data[yr]:
        r = sum(len(all_data[y]) for y in [2022,2023,2024,2025] if y < yr) + d['월'] + 4
        ws0.row_dimensions[r].height = 17
        yr_bg = yr_colors[yr]
        m = d['월']

        # 計算
        fg_u  = (d['FG시간'] / d['근무시간']) if d['FG시간유효'] and d['근무시간'] else None
        gx_u  = (d['GX시간'] / d['근무시간']) if d['GX시간유효'] and d['근무시간'] else None
        su    = (fg_u + gx_u) if (fg_u is not None and gx_u is not None) else None
        fg_e  = d['FG수량'] / FG_STD
        gx_e  = d['GX수량'] / GX_STD
        yp    = ((d['출하수량'] - d['불량']) / d['출하수량']) if d['양품율유효'] else None

        bg_u  = C_WARN if (su is not None and su < 0.50) else (C_OK if (su is not None and su >= 0.65) else None)
        bg_e  = C_WARN if min(fg_e, gx_e) < 0.65 else (C_OK if min(fg_e, gx_e) >= 0.85 else None)
        bg_yp = (C_WARN if yp < 0.985 else C_OK) if yp is not None else C_NA

        row_vals = [
            (yr,          None, yr_bg),
            (m,           None, yr_bg),
            (d['근무일'],  NUM,  yr_bg),
            (fg_u,         PCT,  bg_u  or (C_NA if fg_u is None else yr_bg)),
            (gx_u,         PCT,  bg_u  or (C_NA if gx_u is None else yr_bg)),
            (su,           PCT,  bg_u  or (C_NA if su  is None else yr_bg)),
            (fg_e,         PCT,  bg_e),
            (gx_e,         PCT,  bg_e),
            (yp,           PCT,  bg_yp),
            (d['FG수량'],  NUM,  C_FG),
            (d['GX수량'],  NUM,  C_GX),
            (d['합수량'],  NUM,  C_TOT),
            (d['FG시간'] if d['FG시간유효'] else None, NUM, C_FG),
            (d['GX시간'] if d['GX시간유효'] else None, NUM, C_GX),
            (d['출하수량'], NUM, yr_bg),
            (d['불량'],     NUM, yr_bg),
            (d['비고'],     None, yr_bg),
        ]
        for ci, (val, fmt, bg) in enumerate(row_vals, 1):
            c = ws0.cell(r, ci, value='N/A' if val is None and fmt == PCT else val)
            cs(c, bg=bg, bold=(ci <= 2), ha='left' if ci == 17 else 'center')
            if fmt and val is not None: c.number_format = fmt
            if val is None and fmt == PCT:
                c.font = Font(name=FONT, color="AAAAAA", size=9, italic=True)

# 구분선 (연도 사이)
for yr_idx, yr in enumerate([2022, 2023, 2024]):
    sep_r = sum(len(all_data[y]) for y in [2022,2023,2024,2025] if y <= yr) + 4
    for ci in range(1, 18):
        c = ws0.cell(sep_r, ci)
        c.border = Border(left=thin, right=thin, top=thin, bottom=med)

# 열 너비
w0 = [6,4,8, 9,9,10, 10,10, 9, 9,9,9, 11,11, 8,6, 16]
for ci, w in enumerate(w0, 1):
    ws0.column_dimensions[get_column_letter(ci)].width = w

# ──────────────────────────────────────────────────────────
# 시트2~5: 연도별 상세 (하나의 함수로)
# ──────────────────────────────────────────────────────────
def make_year_sheet(wb, year, data, fg_std, gx_std):
    BIZNOTE_24 = {6:'* GX7 고장(3~22일)', 8:'* 생산량 누락', 9:'* 추석(3~9일)', 10:'* 추석'}
    ws = wb.create_sheet(title=f"{year}년")
    ws.freeze_panes = "B5"

    title_row(ws, 1, f"{year}년 재연마 KPI 상세  —  설비가동율 · 작업효율 · 양품율", 13)

    # 가정 셀
    ws.row_dimensions[2].height = 18
    for ci, (lab, val, extra) in enumerate([
        ("FG 표준생산량", fg_std, "개/월"),
        ("GX7 표준생산량", gx_std, "개/월"),
    ], 1):
        c_lab = ws.cell(2, ci*4-3, value=lab)
        cs(c_lab, bold=True, sz=9, ha='right')
        c_val = ws.cell(2, ci*4-2, value=val)
        c_val.fill = PatternFill("solid", fgColor=C_ASM)
        c_val.font = Font(name=FONT, bold=True, color=C_BLUE, size=10)
        c_val.number_format = '#,##0.00'; c_val.border = BD
        c_val.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(2, ci*4-1, value=extra).font = Font(name=FONT, size=9)

    note = "출처: 25년 월별 생산 실적.xlsx (2025~2026년 15개월 평균)"
    ws.cell(2, 9, value=note).font = Font(name=FONT, size=8, color="888888", italic=True)
    ws.merge_cells('I2:M2')

    # 데이터 경고
    ws.row_dimensions[3].height = 14
    if year == 2022:
        note_row(ws, 3, "⚠️ 2022년 1~10월 가공시간 원본 기록 오류 → 설비가동율 N/A  |  출하현황 5~12월만 존재 → 1~4월 양품율 N/A", 13, fg="CC0000")
    elif year in (2023, 2024):
        note_row(ws, 3, "⚠️ 근무일수: 작업일지 가동일수 기준 (2024년은 참조파일 사용)", 13, fg="CC6600")

    # 헤더
    ws.row_dimensions[4].height = 28
    hdrs = ['월','근무일',
            'FG 생산수량','FG 가공시간(초)','FG 근무시간(초)','FG 설비가동율','FG 작업효율',
            'GX7 생산수량','GX7 가공시간(초)','GX7 가동율','GX7 작업효율',
            '합계 생산수량','양품율']
    for ci, h in enumerate(hdrs, 1):
        hdr(ws.cell(4, ci, value=h),
            sub=(3 <= ci <= 7 or 8 <= ci <= 11))

    valid_util_rows_fg, valid_util_rows_gx, valid_eff_rows = [], [], []

    for i, d in enumerate(data):
        r = i + 5
        m = d['월']
        ws.row_dimensions[r].height = 18

        fg_u  = (d['FG시간'] / d['근무시간']) if d['FG시간유효'] and d['근무시간'] else None
        gx_u  = (d['GX시간'] / d['근무시간']) if d['GX시간유효'] and d['근무시간'] else None
        fg_e  = d['FG수량'] / fg_std
        gx_e  = d['GX수량'] / gx_std
        yp    = ((d['출하수량'] - d['불량']) / d['출하수량']) if d['양품율유효'] else None

        if fg_u is not None: valid_util_rows_fg.append(r)
        if gx_u is not None: valid_util_rows_gx.append(r)
        valid_eff_rows.append(r)

        bg_fu = C_WARN if (fg_u is not None and fg_u < 0.25) else (C_OK if (fg_u is not None and fg_u >= 0.35) else C_FG)
        bg_gu = C_WARN if (gx_u is not None and gx_u < 0.20) else (C_OK if (gx_u is not None and gx_u >= 0.30) else C_GX)
        bg_fe = C_WARN if fg_e < 0.65 else (C_OK if fg_e >= 0.90 else C_FG)
        bg_ge = C_WARN if gx_e < 0.65 else (C_OK if gx_e >= 0.90 else C_GX)
        bg_yp = (C_WARN if yp < 0.985 else C_OK) if yp is not None else C_NA

        row_data = [
            (m,                None,   None,    True),
            (d['근무일'],       NUM,   None,    True),
            (d['FG수량'],       NUM,   C_FG,    True),
            (d['FG시간'] if d['FG시간유효'] else None, NUM, C_FG, True),
            (d['근무시간'],      NUM,   C_FG,    True),
            (fg_u,              PCT,   bg_fu,   True),
            (fg_e,              PCT,   bg_fe,   True),
            (d['GX수량'],       NUM,   C_GX,    True),
            (d['GX시간'] if d['GX시간유효'] else None, NUM, C_GX, True),
            (gx_u,              PCT,   bg_gu,   True),
            (gx_e,              PCT,   bg_ge,   True),
            (d['합수량'],       NUM,   C_TOT,   True),
            (yp,                PCT,   bg_yp,   True),
        ]
        for ci, (val, fmt, bg, show) in enumerate(row_data, 1):
            c = ws.cell(r, ci)
            if val is None and fmt == PCT:
                c.value = 'N/A'
                cs(c, bg=C_NA or bg, ha='center')
                c.font = Font(name=FONT, color="AAAAAA", size=9, italic=True)
            else:
                c.value = val
                cs(c, bg=bg, bold=(ci==1), ha='center')
                if fmt and val is not None: c.number_format = fmt

    # 합계/평균 행
    tot_r = len(data) + 5
    ws.row_dimensions[tot_r].height = 18
    totc(ws.cell(tot_r, 1, value='연간 평균'))
    totc(ws.cell(tot_r, 2, value=f'=AVERAGE(B5:B{tot_r-1})'))
    ws.cell(tot_r, 2).number_format = '#,##0.0'

    # FG 생산수량 합계
    c = ws.cell(tot_r, 3, value=f'=SUM(C5:C{tot_r-1})')
    totc(c); c.number_format = NUM
    # FG 가공시간: 유효한 행만 평균
    if valid_util_rows_fg:
        refs = ','.join([f'D{r}' for r in valid_util_rows_fg])
        c = ws.cell(tot_r, 4, value=f'=AVERAGE({refs})'); totc(c); c.number_format = NUM
        c = ws.cell(tot_r, 5, value=f'=AVERAGE(E5:E{tot_r-1})'); totc(c); c.number_format = NUM
        refs_u = ','.join([f'F{r}' for r in valid_util_rows_fg])
        c = ws.cell(tot_r, 6, value=f'=AVERAGE({refs_u})'); totc(c); c.number_format = PCT
    else:
        for ci in [4,5,6]:
            c = ws.cell(tot_r, ci, value='N/A'); totc(c)
    c = ws.cell(tot_r, 7, value=f'=AVERAGE(G5:G{tot_r-1})'); totc(c); c.number_format = PCT

    c = ws.cell(tot_r, 8, value=f'=SUM(H5:H{tot_r-1})'); totc(c); c.number_format = NUM
    if valid_util_rows_gx:
        refs_gi = ','.join([f'I{r}' for r in valid_util_rows_gx])
        c = ws.cell(tot_r, 9, value=f'=AVERAGE({refs_gi})'); totc(c); c.number_format = NUM
        refs_gu = ','.join([f'J{r}' for r in valid_util_rows_gx])
        c = ws.cell(tot_r, 10, value=f'=AVERAGE({refs_gu})'); totc(c); c.number_format = PCT
    else:
        for ci in [9,10]:
            c = ws.cell(tot_r, ci, value='N/A'); totc(c)
    c = ws.cell(tot_r, 11, value=f'=AVERAGE(K5:K{tot_r-1})'); totc(c); c.number_format = PCT
    c = ws.cell(tot_r, 12, value=f'=SUM(L5:L{tot_r-1})'); totc(c); c.number_format = NUM
    if year == 2025:
        c = ws.cell(tot_r, 13, value=f'=AVERAGE(M5:M{tot_r-1})'); totc(c); c.number_format = PCT
    else:
        c = ws.cell(tot_r, 13, value='N/A'); totc(c)
        c.font = Font(name=FONT, color="AAAAAA", size=9, italic=True)

    # 열 너비
    ws_w = [5,8, 12,13,13,11,11, 12,13,11,11, 12,10]
    for ci, w in enumerate(ws_w, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

for yr in [2022, 2023, 2024, 2025]:
    make_year_sheet(wb, yr, all_data[yr], FG_STD, GX_STD)

# ──────────────────────────────────────────────────────────
# 시트7: 연간 트렌드 요약
# ──────────────────────────────────────────────────────────
wst = wb.create_sheet("연간 트렌드")
title_row(wst, 1, "2022~2025년 연간 KPI 트렌드 요약", 10)
note_row(wst,  2, "⚠️ 2022년 설비가동율은 11~12월(2개월)만 유효  |  2022년 양품율은 5~12월(8개월) 평균  |  2023~2025년은 전체 12개월", 10, fg="CC0000")

trend_hdrs = ['연도','FG 연간생산','GX7 연간생산','합계 생산',
              'FG 작업효율','GX7 작업효율',
              'FG 설비가동율(유효월)','GX7 설비가동율(유효월)',
              '양품율','비고']
for ci, h in enumerate(trend_hdrs, 1):
    hdr(wst.cell(3, ci, value=h))
wst.row_dimensions[3].height = 28

trend_notes = {
    2022: '가동율: 1~10월 기록오류 N/A / 11~12월만 유효 / 양품율: 5~12월',
    2023: '가동율+양품율 모두 유효(전체 12개월)',
    2024: '가동율+양품율 모두 유효(전체 12개월)',
    2025: '가동율+양품율 모두 유효(전체 12개월)',
}

yr_bgs = {2022:'FFF0F5', 2023:'F0FFF0', 2024:'F0F8FF', 2025:'FFFFF0'}

for ri, yr in enumerate([2022, 2023, 2024, 2025]):
    r = ri + 4
    wst.row_dimensions[r].height = 20
    data = all_data[yr]
    bg = yr_bgs[yr]

    fg_tot = sum(d['FG수량'] for d in data)
    gx_tot = sum(d['GX수량'] for d in data)
    ht_tot = sum(d['합수량'] for d in data)
    fg_eff = sum(d['FG수량'] for d in data) / (FG_STD * 12)
    gx_eff = sum(d['GX수량'] for d in data) / (GX_STD * 12)

    valid_fg_u = [(d['FG시간']/d['근무시간']) for d in data if d['FG시간유효'] and d['근무시간']]
    valid_gx_u = [(d['GX시간']/d['근무시간']) for d in data if d['GX시간유효'] and d['근무시간']]
    fg_util = sum(valid_fg_u)/len(valid_fg_u) if valid_fg_u else None
    gx_util = sum(valid_gx_u)/len(valid_gx_u) if valid_gx_u else None

    yp_vals = [((d['출하수량']-d['불량'])/d['출하수량']) for d in data
               if d['양품율유효'] and d['출하수량'] and d['출하수량'] > 0]
    yp = sum(yp_vals)/len(yp_vals) if yp_vals else None

    row_vals = [
        (yr,     None,  bg),
        (fg_tot, NUM,   C_FG),
        (gx_tot, NUM,   C_GX),
        (ht_tot, NUM,   C_TOT),
        (fg_eff, PCT,   C_WARN if fg_eff < 0.70 else (C_OK if fg_eff >= 0.90 else None)),
        (gx_eff, PCT,   C_WARN if gx_eff < 0.70 else (C_OK if gx_eff >= 0.90 else None)),
        (fg_util, PCT,  C_WARN if (fg_util is not None and fg_util < 0.25) else (C_OK if (fg_util is not None and fg_util >= 0.35) else C_NA)),
        (gx_util, PCT,  C_WARN if (gx_util is not None and gx_util < 0.20) else (C_OK if (gx_util is not None and gx_util >= 0.30) else C_NA)),
        (yp,     PCT,   (C_WARN if yp < 0.985 else C_OK) if yp else C_NA),
        (trend_notes[yr], None, bg),
    ]
    for ci, (val, fmt, bg_c) in enumerate(row_vals, 1):
        c = wst.cell(r, ci)
        if val is None and fmt == PCT:
            c.value = 'N/A'
            cs(c, bg=C_NA, ha='center')
            c.font = Font(name=FONT, color="AAAAAA", size=9, italic=True)
        else:
            c.value = val
            cs(c, bg=bg_c, bold=(ci==1), ha='left' if ci in (1,10) else 'center')
            if fmt and val is not None: c.number_format = fmt

# 열 너비
wt_w = [7,12,12,12, 11,11, 15,15, 10, 30]
for ci, w in enumerate(wt_w, 1):
    wst.column_dimensions[get_column_letter(ci)].width = w

# ──────────────────────────────────────────────────────────
# 저장
# ──────────────────────────────────────────────────────────
out = 'wiki/comparisons/KPI_설비가동율_작업효율_양품율_2022-2026.xlsx'
wb.save(out)
print(f"저장 완료: {out}")
print(f"  FG 표준생산량: {FG_STD} / GX7 표준생산량: {GX_STD}")
for yr, data in all_data.items():
    fg_s = sum(d['FG수량'] for d in data)
    gx_s = sum(d['GX수량'] for d in data)
    valid_u = sum(1 for d in data if d['FG시간유효'])
    print(f"  {yr}년: FG={fg_s:,} GX7={gx_s:,} 합={fg_s+gx_s:,}  가동율유효={valid_u}/12개월")
