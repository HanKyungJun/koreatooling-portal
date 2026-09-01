#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
확인 필요 대장 생성 — 위키 본문에 흩어진 미해결 표기를 한 표로 모은다.

실행:
    python scripts/extract_openq.py

입력: wiki/ 하위 모든 .md (아래 SKIP 제외)
출력: wiki/_handoff/확인필요_대장_YYYYMMDD.xlsx
      outputs/worklog_parsed/open_questions.csv  (중간 산출)

━━━ 규칙 ━━━
· 검색어: 확인 필요 / 확인필요 / 미확정 / 미확인 / 미판독 / 미해결
· 제외 대상: _handoff/ · log.md · worklog-archive/ · _backup · _private · templates/
  → 이력 기록물이라 "과거 시점에는 미확정이었다"는 서술이 섞여 있어 현재 미해결과 구분되지 않는다.
· 「관련일자」: 문장 안에 날짜가 있으면 그 날짜, 없으면 문서 frontmatter 의 updated.
  ⚠️ 「질문이 제기된 날」이 아니라 「그 사건이 일어난 날」인 경우가 있다.
· 우선순위: A-긴급(🔴) > B-장기미해결(365일+) > B-경과90일+ > C-일반 > C-해소추정 > D-중복
· 「해소추정」은 같은 줄에 ✅ 또는 「해소」가 있는 경우 — 문장 형태로만 판단하므로 오판 가능.
"""
import re, csv, datetime
from collections import defaultdict, Counter
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent
WIKI = BASE / 'wiki'
CSV_OUT = BASE / 'outputs' / 'worklog_parsed' / 'open_questions.csv'
TODAY = datetime.date.today()
XLSX_OUT = BASE / 'wiki' / '_handoff' / f'확인필요_대장_{TODAY:%Y%m%d}.xlsx'

MARK = re.compile(r'확인\s*필요|미확정|미확인|미판독|미해결')

# ── 2026-08-28 추가: 이미 닫힌 표기를 걸러내는 규칙 ───────────────────────────
#   C-해소추정 33건을 전수 확인한 결과, 대부분은 아직 열린 항목이 아니라
#   ①취소선으로 지워진 옛 표기 ②「"확인 필요" … 해소」처럼 종결을 서술한 문장
#   ③문서 하단의 작성·갱신 이력 줄 이었다. 세 유형만 정밀하게 제외한다.
#   ⚠️ 같은 줄에 ✅ 가 있다는 이유만으로 제외하지 않는다 — 다른 항목이 해소된
#      것일 뿐 이 표기는 열려 있는 경우가 실제로 있다(예: 잔여 확인 2건).
STRIKE = re.compile(r'~~.*?~~')
QUOTED_CLOSED = re.compile(
    r'[「"\'\u201c\u2018(]'                      # 여는 따옴표·괄호
    r'[^「」"\'\u201c\u201d\u2018\u2019()]{0,60}?'
    r'(?:확인\s*필요|미확정|미확인|미판독|미해결)'
    r'[^「」"\'\u201c\u201d\u2018\u2019()]{0,60}?'
    r'[」"\'\u201d\u2019)]'                       # 닫는 따옴표·괄호
    r'[^\n]{0,30}?(?:해소|해결|확정)')
FOOTER = re.compile(r'^\s*작성:\s*\S+\s*/\s*\d{4}-\d{2}-\d{2}')


def already_closed(line):
    """이 줄의 미해결 표기가 전부 이미 닫혔으면 True."""
    if FOOTER.match(line):
        return True                                   # 문서 하단 작성·갱신 이력
    if QUOTED_CLOSED.search(line):
        return True                                   # 「"확인 필요" … 해소」
    if not MARK.search(STRIKE.sub('', line)):
        return True                                   # 표기가 전부 취소선 안
    return False

SKIP = ('_handoff/', 'log.md', 'worklog-archive', '_backup', '_private', 'templates/')
CAT = {'repairs': '설비·정비', 'machines': '설비', 'tools': '공구·휠', 'materials': '소재',
       'cadcam': 'CAD/CAM', 'gcode': 'G코드', 'comparisons': '비교분석', 'projects': '프로젝트',
       'troubleshoot': '트러블슈팅', 'standards': '표준', 'compliance': '법규',
       'measurements': '측정', 'erp': 'ERP', 'scripts': '스크립트', 'reports': '보고'}

F = '맑은 고딕'
NAVY, HDRB, WHT, LGRAY = '1F3864', '2F5597', 'FFFFFF', 'F2F2F2'
WARN, BAD, OK = 'FFF2CC', 'FADBD8', 'E2EFDA'
THIN = Side(style='thin', color='BFBFBF')
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
YEL = PatternFill('solid', fgColor='FFFF00')


def clean(t):
    t = re.sub(r'\[\[([^\]|]+)\|?([^\]]*)\]\]', lambda m: m.group(2) or m.group(1), t)
    t = re.sub(r'\*\*|__|`', '', t)
    t = re.sub(r'^[\s>|#*\-–—0-9.)]+', '', t)
    return re.sub(r'\s+', ' ', t).strip(' |')


def collect():
    rows, seen = [], set()
    for f in sorted(WIKI.rglob('*.md')):
        rel = str(f.relative_to(WIKI)).replace('\\', '/')
        if any(k in rel for k in SKIP):
            continue
        try:
            lines = f.read_text(encoding='utf-8').splitlines()
        except Exception:
            continue
        upd = ''
        for l in lines[:15]:
            m = re.match(r'\s*updated:\s*(\d{4}-\d{2}-\d{2})', l)
            if m:
                upd = m.group(1)
                break
        cat = CAT.get(rel.split('/')[0], '기타')
        for i, l in enumerate(lines, 1):
            if not MARK.search(l):
                continue
            if already_closed(l):
                continue                              # 2026-08-28: 이미 닫힌 표기
            txt = clean(l)
            if len(txt) < 12:
                continue
            if l.strip().startswith('> 📌') or txt.startswith('📌'):
                continue                                  # 문서 보강 안내문
            if re.match(r'^[,\d.\s|]+\|', txt):
                continue                                  # 표 조각
            key = (rel, txt[:70])
            if key in seen:
                continue
            seen.add(key)
            # 2026-08-28 폐기 — 「같은 줄에 ✅ 가 있으면 해소추정」 규칙은 오판이 많았다.
            #   초판 33건을 전수 확인한 결과 11건만 실제 종결이었고(이제 already_closed 가
            #   정밀하게 제외한다), 나머지 22건은 ✅ 가 **다른 항목**의 종결 표시였을 뿐
            #   해당 표기 자체는 열려 있었다(예: "✅ 발주 완료 … 정확한 착수일은 확인 필요").
            #   잘못된 「값싼 정리」 유도를 없애기 위해 분류 자체를 폐지한다.
            resolved = False
            urgent = '🔴' in l
            mk = ('확인 필요' if re.search(r'확인\s*필요', l) else
                  '미확정' if '미확정' in l else
                  '미판독' if '미판독' in l else
                  '미해결' if '미해결' in l else '미확인')
            dm = re.search(r'20\d{2}-\d{2}-\d{2}', l)
            d = dm.group(0) if dm else upd
            try:
                elapsed = (TODAY - datetime.date(*map(int, d.split('-')))).days
            except Exception:
                elapsed = ''
            rows.append([cat, rel, i, mk, '🔴' if urgent else '',
                         '✅추정' if resolved else '', d, elapsed, txt[:300], ''])

    grp = defaultdict(list)
    for i, r in enumerate(rows):
        grp[re.sub(r'[^가-힣A-Za-z0-9]', '', r[8])[:34]].append(i)
    for v in grp.values():
        for j in v[1:]:
            rows[j][9] = '중복'

    for r in rows:
        el = r[7] if isinstance(r[7], int) else 0
        r.append('D-중복' if r[9] == '중복' else
                 'C-해소추정' if r[5] else
                 'A-긴급' if r[4] else
                 'B-장기미해결' if el > 365 else
                 'B-경과90일+' if el > 90 else 'C-일반')
    order = {'A-긴급': 0, 'B-장기미해결': 1, 'B-경과90일+': 2,
             'C-일반': 3, 'C-해소추정': 4, 'D-중복': 5}
    rows.sort(key=lambda r: (order[r[10]], -(r[7] if isinstance(r[7], int) else 0)))
    return rows


def to_csv(rows):
    CSV_OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(CSV_OUT, 'w', newline='', encoding='utf-8-sig') as fh:
        w = csv.writer(fh)
        w.writerow(['우선순위', '분류', '문서', '라인', '표기', '긴급',
                    '상태추정', '관련일자', '경과일', '중복', '내용'])
        w.writerows([[r[10], r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[9], r[8]]
                     for r in rows])


def title(ws, t, n, sz=22):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value=t)
    c.font = Font(name=F, bold=True, size=sz, color=WHT)
    c.fill = PatternFill('solid', fgColor=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36


def shdr(ws, row, n):
    for c in range(1, n + 1):
        cl = ws.cell(row=row, column=c)
        cl.font = Font(name=F, bold=True, color=WHT, size=13)
        cl.fill = PatternFill('solid', fgColor=HDRB)
        cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cl.border = BORD
    ws.row_dimensions[row].height = 30


def put(ws, d, hr):
    for j, c in enumerate(d.columns, 1):
        ws.cell(row=hr, column=j, value=str(c))
    shdr(ws, hr, len(d.columns))
    for i, (_, r) in enumerate(d.iterrows(), hr + 1):
        for j, c in enumerate(d.columns, 1):
            v = r[c]
            if pd.isna(v):
                v = None
            elif isinstance(v, np.integer):
                v = int(v)
            elif isinstance(v, np.floating):
                v = float(v)
            cl = ws.cell(row=i, column=j, value=v)
            cl.font = Font(name=F, size=13)
            cl.border = BORD
            cl.alignment = Alignment(horizontal='left' if j in (1, 2) else 'center',
                                     vertical='center')
            if i % 2 == 0:
                cl.fill = PatternFill('solid', fgColor=LGRAY)
    return hr + len(d)


def autow(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build(rows):
    df = pd.read_csv(CSV_OUT)
    df['경과일'] = pd.to_numeric(df['경과일'], errors='coerce')
    cnt = Counter(df['우선순위'])
    wb = Workbook(); wb.remove(wb.active)

    ws = wb.create_sheet('개요')
    title(ws, '확인 필요 대장 — cnc-wiki 미해결 항목 단일 인덱스', 6)
    info = [
        ('■ 목적', '위키 문서 본문에 흩어져 있던 「확인 필요」·「미확정」 표기를 한 표로 모아, '
                   '무엇이 아직 안 닫혔는지 한눈에 보고 우선순위를 정하기 위함.'),
        ('', ''),
        ('■ 수집 범위', ''),
        ('대상', f'wiki/ 하위 콘텐츠 문서 전체 ({TODAY} 기준)'),
        ('검색어', '「확인 필요」·「확인필요」·「미확정」·「미확인」·「미판독」·「미해결」'),
        ('제외', '_handoff/ · log.md · worklog-archive/ · _backup · _private · templates/ '
                 '— 이력 기록물이라 과거 시점의 표기가 섞여 있음'),
        ('제외 (표기 단위)', '2026-08-28 추가 — ①취소선(~~…~~)으로 지워진 표기 '
                          '②「"확인 필요" … 해소」처럼 종결을 서술한 문장 '
                          '③문서 하단 「작성: … / 갱신(…)」 이력 줄'),
        ('추출 건수', f'{len(df)}건'),
        ('', ''),
        ('■ 우선순위 정의', ''),
        ('A-긴급', f'본문에 🔴 표기가 붙은 항목 ({cnt.get("A-긴급",0)}건)'),
        ('B-장기미해결', f'관련일자로부터 365일 초과 ({cnt.get("B-장기미해결",0)}건)'),
        ('B-경과90일+', f'관련일자로부터 90일 초과 ({cnt.get("B-경과90일+",0)}건)'),
        ('C-일반', f'그 외 ({cnt.get("C-일반",0)}건)'),
        ('C-해소추정', '⛔ 2026-08-28 폐지 — 「같은 줄에 ✅」 규칙은 오판이 많았다(33건 중 '
                     '실제 종결은 11건뿐). 실제 종결분은 이제 추출 단계에서 제외되고, '
                     '나머지는 열린 항목이므로 C-일반으로 들어간다.'),
        ('D-중복', f'다른 문서에 같은 내용이 이미 있음 ({cnt.get("D-중복",0)}건)'),
        ('', ''),
        ('■ 「관련일자」의 의미',
         '⚠️ 문장 안에 날짜가 있으면 그 날짜를, 없으면 문서의 updated 날짜를 넣었다. '
         '따라서 「질문이 제기된 날」이 아니라 「그 사건이 일어난 날」인 경우가 있다 — '
         '예: 2021년 수리 이력의 경과일 2,000일은 그 수리가 2021년에 있었다는 뜻이지 '
         '질문이 2,000일 묵었다는 뜻이 아니다.'),
        ('', ''),
        ('■ 사용법', ''),
        ('1', 'A-긴급 → B-장기미해결 순으로 훑는다.'),
        ('2', '한 문서에 여러 건이 몰려 있으면 그 문서를 한 번에 여는 것이 가장 값싸다 — 「문서별 집계」 시트 참조.'),
        ('3', '조치·담당·완료일 열(노란 칸)에 기입하고, 닫힌 항목은 원본 문서에서 '
              '「확인 필요」 표기를 지운 뒤 이 대장을 다시 뽑는다.'),
        ('4', '원본 문서·라인 번호가 있으므로 바로 찾아갈 수 있다.'),
        ('', ''),
        ('■ 재생성', 'python scripts/extract_openq.py — 원본 문서가 바뀌면 다시 뽑는다. '
                   '파일명에 날짜가 들어가므로 이전 판은 덮어쓰지 않는다.'),
        ('', ''),
        ('■ 신뢰도', '실측 검증 — 위키 원문 직접 추출. 단 「해소추정」은 문장 형태로 판단한 것이라 '
                   '오판 가능 — 사람 확인 필요.'),
    ]
    r = 3
    for a, b in info:
        c = ws.cell(row=r, column=1, value=a)
        c.font = Font(name=F, bold=a.startswith('■'),
                      size=15 if a.startswith('■') else 13,
                      color=NAVY if a.startswith('■') else '000000')
        c2 = ws.cell(row=r, column=3, value=b)
        c2.font = Font(name=F, size=13)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        c2.alignment = Alignment(vertical='center', wrap_text=True)
        if b and len(b) > 90:
            ws.row_dimensions[r].height = 52
        r += 1
    autow(ws, [18, 2, 44, 24, 24, 24])

    ws = wb.create_sheet('대장')
    title(ws, f'확인 필요 대장 ({len(df)}건)', 14)
    ws.cell(row=2, column=1,
            value='노란 열(조치·담당·완료일·비고)은 직접 기입하는 칸입니다.').font = \
        Font(name=F, size=11, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)
    d = df[['우선순위', '분류', '문서', '라인', '표기', '관련일자', '경과일', '중복', '내용']].copy()
    d.insert(0, '번호', range(1, len(d) + 1))
    for c in ['조치', '담당', '완료일', '비고']:
        d[c] = ''
    end = put(ws, d, 3)
    col = {'A-긴급': BAD, 'B-장기미해결': WARN, 'B-경과90일+': WARN,
           'C-해소추정': OK, 'D-중복': LGRAY}
    for i in range(4, end + 1):
        pr = ws.cell(row=i, column=2).value
        if pr in col:
            for j in (1, 2):
                ws.cell(row=i, column=j).fill = PatternFill('solid', fgColor=col[pr])
        for j in range(11, 15):
            cl = ws.cell(row=i, column=j)
            cl.fill = YEL
            cl.border = BORD
        ws.cell(row=i, column=10).alignment = Alignment(horizontal='left', vertical='top',
                                                        wrap_text=True)
    autow(ws, [6, 14, 12, 42, 7, 10, 12, 8, 8, 80, 26, 10, 12, 20])
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:N{end}'

    ws = wb.create_sheet('문서별 집계')
    title(ws, '문서별 집계 — 어디에 몰려 있나', 6)
    g = (df.groupby(['분류', '문서'])
           .agg(건수=('내용', 'size'), 긴급=('긴급', lambda x: x.notna().sum()),
                해소추정=('상태추정', lambda x: x.notna().sum()), 최장경과일=('경과일', 'max'))
           .reset_index().sort_values('건수', ascending=False))
    e = put(ws, g, 3)
    ws.cell(row=e + 1, column=1, value='합계').font = Font(name=F, bold=True, size=13)
    ws.cell(row=e + 1, column=3, value=f'=SUM(C4:C{e})').font = Font(name=F, bold=True, size=13)
    for c in range(1, 7):
        cl = ws.cell(row=e + 1, column=c)
        cl.fill = PatternFill('solid', fgColor=WARN)
        cl.border = BORD
    autow(ws, [14, 46, 10, 10, 12, 14])
    ws.freeze_panes = 'A4'

    ws = wb.create_sheet('분류별 요약')
    title(ws, '분류별 · 우선순위별 요약', 8)
    p = pd.crosstab(df['분류'], df['우선순위'], margins=True, margins_name='합계').reset_index()
    put(ws, p, 3)
    autow(ws, [16] + [13] * (len(p.columns) - 1))

    XLSX_OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_OUT)
    print(f'저장 완료: {XLSX_OUT}')
    for k, v in Counter(df['우선순위']).most_common():
        print(f'  {k}: {v}건')


if __name__ == '__main__':
    rows = collect()
    print(f'추출 {len(rows)}건')
    to_csv(rows)
    build(rows)
