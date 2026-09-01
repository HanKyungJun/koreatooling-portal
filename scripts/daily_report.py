#!/usr/bin/env python3
"""
재연마 작업일지 일일보고 자동화
실행: python scripts/daily_report.py [YYYY-MM-DD]
출력: wiki/reports/daily/YYYY-MM-DD_일일보고.xlsx
"""
import sys, io, argparse
from datetime import datetime
from pathlib import Path
import pandas as pd
import warnings
warnings.filterwarnings('ignore')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 경로 ──────────────────────────────────────────────────
BASE = Path(__file__).resolve().parent.parent
RAW  = BASE / 'raw' / '출하현황'
OUT  = BASE / 'wiki' / 'reports' / 'daily'

# ── 스타일 상수 ────────────────────────────────────────────
FONT   = "맑은 고딕"
CBHDR  = "1F497D"; CFHDR = "FFFFFF"
CFGBG  = "D6E4F7"   # FAST GRIND 연파랑
CGXBG  = "E2EFDA"   # GX7 연초록
CTOTBG = "FFF2CC"   # 합계 연노랑
CGRAY  = "D9D9D9"
CUPBG  = "D5F5E3"   # 전일대비 증가 연초록
CDNBG  = "FADBD8"   # 전일대비 감소 연빨강
CFG_H  = "4472C4"   # FAST GRIND 헤더 파랑
CGX_H  = "70AD47"   # GX7 헤더 초록

thin  = Side(style='thin', color='BBBBBB')
BORD  = Border(left=thin, right=thin, top=thin, bottom=thin)
NUM   = '#,##0'
DNUM  = '+#,##0;-#,##0;"-"'
DPCT  = '+0.0%;-0.0%;"-"'


def s(c, bold=False, fg="000000", bg=None, h='center', sz=10, wrap=False):
    c.font      = Font(name=FONT, bold=bold, color=fg, size=sz)
    if bg:
        c.fill  = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=h, vertical='center', wrap_text=wrap)
    c.border    = BORD


def hdr(c, txt=None, bg=CBHDR, fg=CFHDR):
    if txt is not None:
        c.value = txt
    c.font      = Font(name=FONT, bold=True, color=fg, size=10)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border    = BORD


def title(ws, row, text, ncols):
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name=FONT, bold=True, size=13, color=CFHDR)
    c.fill      = PatternFill("solid", fgColor=CBHDR)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 24


def tot_s(c, txt=None):
    if txt is not None:
        c.value = txt
    c.font      = Font(name=FONT, bold=True, size=10)
    c.fill      = PatternFill("solid", fgColor=CGRAY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = BORD


def _int(v):
    try:
        return int(float(v)) if pd.notna(v) else 0
    except Exception:
        return 0


def sec_to_hms(sec: int) -> str:
    h, rem = divmod(abs(sec), 3600)
    m, s   = divmod(rem, 60)
    sign   = "-" if sec < 0 else ""
    return f"{sign}{h}:{m:02d}:{s:02d}"


# ── 파일 탐색 ──────────────────────────────────────────────
def find_xls(year: int, month: int) -> Path:
    p = RAW / f'재연마 작업일지({year})' / f'재연마_월간생산일지 ({month}월).xls'
    if not p.exists():
        raise FileNotFoundError(f"작업일지 파일 없음: {p}")
    return p


# ── 월간합계표 → 일별 요약 DataFrame ──────────────────────
def read_summary(path: Path) -> pd.DataFrame:
    """월간합계표 시트에서 일별 FG/GX7/합계 데이터 추출."""
    df = pd.read_excel(path, sheet_name=0, header=None)
    rows = []
    for i in range(2, 33):          # pandas 행 2~32 = 1~31일
        r = df.iloc[i]
        try:
            day = int(float(r[0]))
        except Exception:
            break
        rows.append({
            '일':       day,
            'FG수량':   _int(r[1]),  'FG금액':   _int(r[2]),  'FG시간':  _int(r[3]),
            'GX수량':   _int(r[7]),  'GX금액':   _int(r[8]),  'GX시간':  _int(r[9]),
            '합수량':   _int(r[13]), '합금액':   _int(r[14]), '합시간':  _int(r[15]),
        })
    return pd.DataFrame(rows)


# ── 일별 작업 상세 추출 ────────────────────────────────────
def read_detail(path: Path, day: int) -> dict:
    """해당 일 시트에서 FAST GRIND / GX7 작업 목록 추출."""
    xl    = pd.ExcelFile(path)
    sname = str(day)
    if sname not in xl.sheet_names:
        return {'fg': [], 'gx': []}
    df = pd.read_excel(xl, sheet_name=sname, header=None)

    def parse(start, end):
        jobs = []
        for i in range(start, end):
            r   = df.iloc[i]
            qty = _int(r[10])
            if qty <= 0:
                continue
            # 2026-08-28: 블록 범위를 동적으로 잡으면서 블록 끝의 「합계」행과
            #   시트 하단 「계기값」행이 범위에 들어온다. 둘 다 수량은 있으나
            #   형상이 비어 있으므로 이것으로 걸러낸다(parse_worklog.py 규칙 2와 동일).
            if not (str(r[1]).strip() if pd.notna(r[1]) else ''):
                continue
            jobs.append({
                '순서':       _int(r[0]),
                '형상':       str(r[1]).strip() if pd.notna(r[1]) else '',
                '날수(F)':    _int(r[2]),
                '날경(Ø)':    r[3] if pd.notna(r[3]) else '',
                '상크경(Ø)':  r[4] if pd.notna(r[4]) else '',
                '코팅':       str(r[5]).strip() if pd.notna(r[5]) else '',
                '특이사항':   str(r[6]).strip() if pd.notna(r[6]) else '',
                '완료여부':   str(r[8]).strip() if pd.notna(r[8]) else '',
                '수량':       qty,
                '시간합계':   _int(r[13]),   # 총 가공시간(초) = 수량 × 단위시간
                '금액':       _int(r[12]),
            })
        return jobs

    # ── 블록 범위 동적 탐지 ────────────────────────────────
    # 🔴 2026-08-28 결함 수정 — 기존에는 fg=행 2~16, gx=행 21~35로 **고정**이었다.
    #   2022년 일부 일자 시트는 FAST GRIND 슬롯이 15개를 넘어 GX7 블록이 아래로
    #   밀리는데, 고정 범위로 읽으면 초과분과 밀린 GX7분이 통째로 누락된다.
    #   ── 고정 범위 vs 동적 탐지 전수 대조 실측 (2026-08-28) ──
    #     2022년: 3,621건/26,106개 → 3,879건/27,209개  (+258건 / +1,103개, 43일분)
    #     2026년: 1,738건/10,863개 → 1,742건/10,874개  (+4건 / +11개, 06-09 하루)
    #   ⚠️ 2026-08-28 세션 초반에 언급된 「2022년 759건(+20.8%) 누락」은
    #      parse_worklog.py의 다른 집계 기준에서 나온 값이다. 이 함수 기준의
    #      실측 차이는 위 표가 정확하다. A열의 블록 마커(FAST GRIND / GX7)를 찾아
    #   마커+2행부터 다음 마커 직전까지를 데이터 구간으로 잡는다.
    marks = []
    for i in range(len(df)):
        v = df.iat[i, 0]
        tag = str(v).upper().replace(' ', '') if pd.notna(v) else ''
        if tag in ('FASTGRIND', 'GX7'):
            marks.append((i, 'fg' if tag == 'FASTGRIND' else 'gx'))

    span = {}
    for k, (row, key) in enumerate(marks):
        end = marks[k + 1][0] if k + 1 < len(marks) else len(df)
        span.setdefault(key, (row + 2, end))

    if not span:                       # 마커를 못 찾으면 기존 고정 범위로 폴백
        return {'fg': parse(2, 17), 'gx': parse(21, 36)}

    return {
        'fg': parse(*span['fg']) if 'fg' in span else [],
        'gx': parse(*span['gx']) if 'gx' in span else [],
    }


# ── 전일(마지막 작업일) 데이터 ─────────────────────────────
def find_prev(summary: pd.DataFrame, today_day: int):
    prev = summary[(summary['일'] < today_day) & (summary['합수량'] > 0)]
    return prev.iloc[-1].to_dict() if not prev.empty else None


# ── 월평균 (오늘 이전 작업일 기준) ────────────────────────
def calc_avg(summary: pd.DataFrame, today_day: int) -> dict:
    work = summary[(summary['일'] < today_day) & (summary['합수량'] > 0)]
    keys = ['FG수량', 'FG시간', 'GX수량', 'GX시간', '합수량', '합시간', '합금액']
    if work.empty:
        return {k: 0 for k in keys}
    return {k: work[k].mean() for k in keys}


# ══════════════════════════════════════════════════════════
#  Excel 보고서 생성
# ══════════════════════════════════════════════════════════
def make_report(target: datetime) -> Path:
    y, m, d = target.year, target.month, target.day

    xls_path = find_xls(y, m)
    summary  = read_summary(xls_path)

    today_row = summary[summary['일'] == d]
    if today_row.empty:
        return None  # 당일 데이터 미입력 상태 (오전 실행 등) — 조용히 스킵

    td     = today_row.iloc[0].to_dict()
    prev   = find_prev(summary, d)
    avg    = calc_avg(summary, d)
    detail = read_detail(xls_path, d)

    OUT.mkdir(parents=True, exist_ok=True)
    out_path = OUT / f'{y}-{m:02d}-{d:02d}_일일보고.xlsx'

    wb  = Workbook()
    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    # ── 시트 1: 일일 요약 ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "일일 요약"

    date_str = target.strftime("%Y년 %m월 %d일")
    prev_str = f"{m}월 {int(prev['일'])}일" if prev else "없음"

    # 작업일 통계
    work_days = summary[(summary['일'] <= d) & (summary['합수량'] > 0)]
    n_work    = len(work_days)
    n_prev    = len(summary[(summary['일'] < d) & (summary['합수량'] > 0)])

    title(ws1, 1, f"재연마 일일생산 보고  ─  {date_str}  (17:30 기준)", 9)

    ws1.merge_cells('A2:I2')
    ws1['A2'] = (f"생성: {now}   |   전일: {prev_str}   |"
                 f"   월평균 기준: 이전 작업일 {n_prev}일 평균   |   {m}월 누적 작업일: {n_work}일")
    ws1['A2'].font      = Font(name=FONT, size=9, color="555555", italic=True)
    ws1['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[2].height = 15

    # ── 헤더 행
    hdrs3 = ['항목', '오늘', '전일', '전일대비\n(▲▼수량)', '전일대비\n(%)',
             '월평균', '월평균대비\n(▲▼수량)', '월평균대비\n(%)', '비고']
    for ci, h in enumerate(hdrs3, 1):
        hdr(ws1.cell(3, ci, value=h))
    ws1.row_dimensions[3].height = 28

    # ── 데이터 행 정의
    ROWS = [
        ('FAST GRIND  수량 (개)',       'FG수량',  CFGBG, False),
        ('FAST GRIND  가공시간 (초)',   'FG시간',  CFGBG, True),
        ('GX7  수량 (개)',              'GX수량',  CGXBG, False),
        ('GX7  가공시간 (초)',          'GX시간',  CGXBG, True),
        ('합계  수량 (개)',             '합수량',  CTOTBG, False),
        ('합계  가공시간 (초)',         '합시간',  CTOTBG, True),
        ('합계  금액',                 '합금액',  CTOTBG, False),
    ]

    for ri, (label, key, rowbg, is_time) in enumerate(ROWS):
        r = ri + 4
        ws1.row_dimensions[r].height = 22

        tv  = td.get(key, 0)
        pv  = prev.get(key, 0) if prev else None
        av  = avg.get(key, 0)

        d_prev = (tv - pv) if pv is not None else None
        d_avg  = tv - av
        p_prev = (d_prev / pv) if (pv and pv != 0) else None
        p_avg  = (d_avg  / av) if av != 0 else 0

        bg_prev = (CUPBG if d_prev > 0 else CDNBG) if d_prev is not None and d_prev != 0 else None
        bg_avg  = (CUPBG if d_avg  > 0 else CDNBG) if d_avg  != 0 else None

        note = sec_to_hms(tv) if is_time else ''

        c = ws1.cell(r, 1, value=label);     s(c, bold=True, bg=rowbg, h='left')
        c = ws1.cell(r, 2, value=tv);        s(c, bold=True); c.number_format = NUM
        c = ws1.cell(r, 3, value=pv if pv is not None else '-')
        s(c); c.number_format = NUM if pv is not None else 'General'
        c = ws1.cell(r, 4, value=d_prev if d_prev is not None else '-')
        s(c, bg=bg_prev); c.number_format = DNUM if d_prev is not None else 'General'
        c = ws1.cell(r, 5, value=p_prev if p_prev is not None else '-')
        s(c, bg=bg_prev); c.number_format = DPCT if p_prev is not None else 'General'
        c = ws1.cell(r, 6, value=round(av, 1)); s(c); c.number_format = '#,##0.0'
        c = ws1.cell(r, 7, value=d_avg);        s(c, bg=bg_avg); c.number_format = DNUM
        c = ws1.cell(r, 8, value=p_avg);        s(c, bg=bg_avg); c.number_format = DPCT
        c = ws1.cell(r, 9, value=note);         s(c, fg="444444", sz=9)

    # ── 월 누적 합계 행
    tot_r = len(ROWS) + 4
    ws1.row_dimensions[tot_r].height = 22
    mtd = summary[(summary['일'] <= d) & (summary['합수량'] > 0)]

    ws1.merge_cells(f'A{tot_r}:B{tot_r}')
    tot_s(ws1.cell(tot_r, 1), f'{m}월 누적  ({n_work}일 / {d}일 중 작업일)')
    tot_s(ws1.cell(tot_r, 2))

    for ci, key in enumerate(['합수량', '합시간', '합금액'], 3):
        c = ws1.cell(tot_r, ci, value=int(mtd[key].sum()) if not mtd.empty else 0)
        tot_s(c); c.number_format = NUM
    for ci in range(6, 10):
        tot_s(ws1.cell(tot_r, ci))

    # ── 열 너비
    ws1.column_dimensions['A'].width = 28
    for ci in range(2, 10):
        ws1.column_dimensions[get_column_letter(ci)].width = 14

    # ── 시트 2: 작업 상세 ──────────────────────────────────
    ws2 = wb.create_sheet("작업 상세")
    title(ws2, 1, f"작업 상세  ─  {date_str}", 11)

    DHDR = ['순서', '형상', '날수(F)', '날경(Ø)', '상크경(Ø)',
            '코팅', '특이사항', '완료', '수량(개)', '시간합계(초)', '시간(H:M:S)']
    row_cur = 2

    for machine, jobs, mhdr_bg in [('FAST GRIND', detail['fg'], CFG_H),
                                    ('GX7',        detail['gx'], CGX_H)]:
        rowbg = CFGBG if machine == 'FAST GRIND' else CGXBG

        # 기계명 타이틀
        ws2.merge_cells(f'A{row_cur}:{get_column_letter(len(DHDR))}{row_cur}')
        c = ws2.cell(row_cur, 1, value=f'■ {machine}')
        c.font      = Font(name=FONT, bold=True, size=11, color=CFHDR)
        c.fill      = PatternFill("solid", fgColor=mhdr_bg)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border    = BORD
        ws2.row_dimensions[row_cur].height = 22
        row_cur += 1

        # 헤더
        for ci, h in enumerate(DHDR, 1):
            hdr(ws2.cell(row_cur, ci, value=h), bg=mhdr_bg)
        ws2.row_dimensions[row_cur].height = 20
        row_cur += 1

        if not jobs:
            ws2.merge_cells(f'A{row_cur}:{get_column_letter(len(DHDR))}{row_cur}')
            c = ws2.cell(row_cur, 1, value='(해당일 작업 없음)')
            s(c, fg="888888")
            row_cur += 1
        else:
            for job in jobs:
                ws2.row_dimensions[row_cur].height = 18
                vals = [
                    job['순서'], job['형상'], job['날수(F)'], job['날경(Ø)'],
                    job['상크경(Ø)'], job['코팅'], job['특이사항'], job['완료여부'],
                    job['수량'], job['시간합계'], sec_to_hms(job['시간합계']),
                ]
                for ci, v in enumerate(vals, 1):
                    c = ws2.cell(row_cur, ci, value=v)
                    is_left = ci in (2, 6, 7)
                    s(c, bg=rowbg, h='left' if is_left else 'center')
                    if ci in (9, 10):
                        c.number_format = NUM
                row_cur += 1

            # 소계
            tot_qty  = sum(j['수량']      for j in jobs)
            tot_time = sum(j['시간합계']  for j in jobs)
            ws2.merge_cells(f'A{row_cur}:H{row_cur}')
            tot_s(ws2.cell(row_cur, 1), f'{machine} 소계')
            for ci in range(2, 9):
                tot_s(ws2.cell(row_cur, ci))
            c = ws2.cell(row_cur, 9,  value=tot_qty);          tot_s(c); c.number_format = NUM
            c = ws2.cell(row_cur, 10, value=tot_time);         tot_s(c); c.number_format = NUM
            c = ws2.cell(row_cur, 11, value=sec_to_hms(tot_time)); tot_s(c)
            row_cur += 1

        row_cur += 1    # 기계 간 빈 행

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 14
    for ci in range(3, 6):
        ws2.column_dimensions[get_column_letter(ci)].width = 9
    ws2.column_dimensions['F'].width = 10
    ws2.column_dimensions['G'].width = 24
    ws2.column_dimensions['H'].width = 6
    ws2.column_dimensions['I'].width = 9
    ws2.column_dimensions['J'].width = 12
    ws2.column_dimensions['K'].width = 12

    wb.save(out_path)
    print(f"✓ 보고서 저장 완료: {out_path}")
    return out_path


# ── 메인 ──────────────────────────────────────────────────
if __name__ == '__main__':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    parser = argparse.ArgumentParser(description='재연마 일일보고 생성')
    parser.add_argument('date', nargs='?',
                        help='날짜 (YYYY-MM-DD). 생략 시 오늘')
    args = parser.parse_args()

    target = (datetime.strptime(args.date, '%Y-%m-%d')
              if args.date else datetime.now())

    try:
        result_path = make_report(target)
    except FileNotFoundError as e:
        print(f"[오류] {e}")
        sys.exit(1)
    except Exception as e:
        import traceback
        traceback.print_exc()
        sys.exit(1)

    if result_path is None:
        print(f"[스킵] {target.strftime('%Y-%m-%d')} 데이터가 아직 없습니다 — 재실행 불필요")
        sys.exit(0)

    # 일일보고 생성 후 현황판 자동 업로드
    import subprocess, os, time
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    def _ts():
        return datetime.now().strftime('%H:%M:%S')

    # 2026-07-07: 16:00 정기 실행만 30분+ 소요되는 병목 조사용 — generate.py 전체를
    # 감싸는 시작/종료 타임스탬프를 기록해, generate.py 내부 로그(단계별 타이밍 추가됨)와
    # 대조해 어느 구간에서 시간이 소요되는지 다음 16:00 실행에서 특정한다.
    # 관련: wiki/_handoff/tasks.md "16:00 정기 실행만 30분+ 소요되는 원인 조사" (P1)
    _t0 = time.time()
    print(f"[{_ts()}] [현황판] generate.py 실행 중...")
    try:
        result = subprocess.run(
            [sys.executable, os.path.join(base, 'generate.py')],
            cwd=base,
            timeout=1800  # 2026-07-06: 600s(10분)에서 상향. repo 비대화(.git 830MB, Pages 아티팩트 478MB)로
                          # generate.py 전체(git add+push 포함) 실행이 10분을 넘겨 push까지 스킵되는 현상 확인.
                          # Task Scheduler ExecutionTimeLimit(PT0S, 무제한)과 별개의 파이썬 레벨 안전장치로 30분 유지.
        )
        _dt = time.time() - _t0
        if result.returncode == 0:
            print(f"[{_ts()}] [현황판] GitHub Pages 업로드 완료 (generate.py 전체 {_dt:.1f}s)")
        else:
            print(f"[{_ts()}] [현황판] 업로드 실패 (generate.py 오류, {_dt:.1f}s)")
    except subprocess.TimeoutExpired:
        _dt = time.time() - _t0
        print(f"[{_ts()}] [현황판] 업로드 시간 초과 (30분, 실제 경과 {_dt:.1f}s) — 건너뜀")
