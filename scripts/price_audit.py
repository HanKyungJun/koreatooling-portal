#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""재연마 단가표 전수 검수 (price_audit)

판독 규칙 전문: wiki/standards/재연마-정가표-읽는법.md 를 먼저 읽으세요.
이 스크립트는 그 문서 §7 「검수 5종 세트」를 자동화한 것입니다.

  A 위생       정가 결측·0이하·10원단위·코드중복·품목명 6필드·코드↔품목명 대조
  B 정가표대조  블록을 '제목'으로 찾아 구간·날수·코팅을 맞춰 기대값 산출 후 전수 비교
  C 파일 간     같은 품목명은 파일이 달라도 정가가 같아야 함
  D 가격 역전   직경↑ 인데 값↓ 이면 무조건 오류
  E 코팅 정합   비코팅 <= 일반 <= 고경도

D·E 는 블록 매핑 없이도 통하는 판별법이라 미매핑 영역의 오류까지 잡습니다.

사용법:
  python scripts/price_audit.py
  python scripts/price_audit.py --xlsx --json out.json

주의: 블록 좌표를 하드코딩하지 않고 '제목 텍스트'로 찾습니다.
      정가표가 개정돼 행·열이 밀려도 따라가며, 제목이 바뀌면 경고로 알려줍니다.
⚠️ 금액은 콘솔·xlsx 에만 나옵니다. 위키에는 기재하지 않습니다 (CLAUDE.md §4).
"""
import argparse, collections, datetime, json, os, re, sys

KST = datetime.timezone(datetime.timedelta(hours=9))
def now_kst():
    """파일명·표기 시각은 항상 KST. (컨테이너·VM 시계가 UTC인 환경 대비)"""
    return datetime.datetime.now(KST)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl 이 필요합니다:  pip install openpyxl")

# ── 확인된 예외 (오류 아님) ────────────────────────────────────
EXCEPTIONS_E2 = {('코너', '초경', '밑옆날', 6.0)}      # 고경도<일반이 맞음 (2026-09-02 확인)
NON_STANDARD_NAME_PREFIX = ('원통연삭/', '절단/')       # 형상·재질 개념 없는 정상 품목
CODE_DIA_TOLERATED = {'5.45', '5.86'}                  # 품목코드 직경 3자리 한계 (값 정상)

SHEET_ITEMS = '품목일괄등록'
COL_CODE, COL_NAME, COL_PRICE, ROW_START = 2, 3, 28, 3

BAND = re.compile(r'^(\d+(?:\.\d+)?)\s*(이하|미만|이상)?$')
BLADE = re.compile(r'^(2날|3날|4날|6날|3,4,6날)$')

# 🔴 「月 수량 …」 수량 할인표 사본 구역. 정가 블록과 제목이 글자 그대로 같아서
#    잘못 잡히면 기대값 전체가 조용히 틀린다 (2026-09-02, 거짓 불일치 1,726건 발생).
#    정가 블록은 이 구역 밖(왼쪽 c1~c14 또는 r26 이하)에 있다.
DISCOUNT_ZONE_ROWS = (1, 25)
DISCOUNT_ZONE_COLS = (15, 32)

def in_discount_zone(r, c):
    return (DISCOUNT_ZONE_ROWS[0] <= r <= DISCOUNT_ZONE_ROWS[1]
            and DISCOUNT_ZONE_COLS[0] <= c <= DISCOUNT_ZONE_COLS[1])

# ── 코팅 그룹 라벨 정규화 ──────────────────────────────────
# 정가표 개정마다 같은 그룹의 라벨이 `코팅` ↔ `일반 코팅` ↔ `고경도 코팅` 으로 바뀐다.
# (2026-09-02 저장에서 `코팅` → `고경도 코팅` 으로 바뀌어 조회가 전부 실패했다.)
# 그래서 읽을 때 한 그룹을 가능한 모든 표기로 등록해 둔다.
COAT_ALIASES = {
    '비코팅': ('비코팅',),
    '일반':   ('일반', '일반코팅', '일반 코팅'),
    '고경도': ('고경도', '고경도코팅', '고경도 코팅'),
    '코팅':   ('코팅',),
}

def canon_coat(g):
    """그룹 라벨 → 표준 키(비코팅/일반/고경도/코팅). 코팅 라벨이 아니면 None."""
    t = re.sub(r'\s+', '', str(g or ''))
    if t == '비코팅': return '비코팅'
    if t.startswith('고경도'): return '고경도'
    if t.startswith('일반'): return '일반'
    if t == '코팅': return '코팅'
    return None

def alias_groups(cols):
    """vtable 의 그룹 dict 에 라벨 별칭을 추가한다 (기존 키는 건드리지 않는다)."""
    out = dict(cols)
    canon = {}
    for k in cols:
        c = canon_coat(k)
        if c: canon.setdefault(c, k)
    for c, k in canon.items():
        for a in COAT_ALIASES.get(c, (c,)):
            out.setdefault(a, cols[k])
    # 코팅 그룹이 하나뿐인 블록이면 어느 표기로 물어도 그 그룹을 준다
    coats = [c for c in canon if c in ('일반', '고경도', '코팅')]
    if len(coats) == 1:
        only = cols[canon[coats[0]]]
        for a in ('코팅', '일반', '일반 코팅', '일반코팅', '고경도', '고경도 코팅', '고경도코팅'):
            out.setdefault(a, only)
    return out

def alias_rows(rows):
    """htable 의 `그룹+날수` 합성 키에 라벨 별칭을 추가한다."""
    out = dict(rows)
    groups = set()
    for k in rows:
        m = re.match(r'^(.+?)([24])$', k)
        if m: groups.add(m.group(1))
    coats = [g for g in groups if g in ('일반', '고경도', '코팅')]
    for k, v in rows.items():
        m = re.match(r'^(.+?)([24])$', k)
        if not m: continue
        g, n = m.group(1), m.group(2)
        if g not in ('일반', '고경도', '코팅'): continue
        for a in ('일반', '고경도', '코팅') if len(coats) == 1 else COAT_ALIASES.get(g, (g,)):
            out.setdefault(a + n, v)
    return out


def band(v):
    if v is None: return None
    m = BAND.match(str(v).strip())
    return (float(m.group(1)), m.group(2) or '이하') if m else None

def in_band(val, kind, d):
    if kind == '이하': return d <= val + 1e-9
    if kind == '미만': return d < val - 1e-9
    return d >= val - 1e-9


class PriceList:
    """정가표 시트 판독기. 블록은 제목 텍스트로 찾는다."""

    def __init__(self, path, sheet=None):
        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet is None:
            # '정가표(26.09)' 처럼 개정 연월이 붙은 시트 중 최신을 고른다.
            # ('할인 0.75', '최종금액' 같은 파생 시트를 잘못 잡지 않도록)
            cand = []
            for s in wb.sheetnames:
                m = re.match(r'^정가표\s*\((\d{2})\.(\d{2})\)\s*$', s.strip())
                if m: cand.append(((int(m.group(1)), int(m.group(2))), s))
            if not cand:
                sys.exit('개정 연월 시트를 찾지 못했습니다. --sheet 로 지정하세요.\n'
                         '  시트 목록: ' + ', '.join(wb.sheetnames))
            sheet = max(cand)[1]
        self.name, self.ws = sheet, wb[sheet]
        self.titles = collections.defaultdict(list)
        self.errors = []
        self.dropped = []      # 할인표 구역에서 걸러낸 제목 (진단용)
        for r in range(1, self.ws.max_row + 1):
            for c in range(1, self.ws.max_column + 1):
                v = self.ws.cell(r, c).value
                if isinstance(v, str):
                    if v.startswith('#'): self.errors.append((r, c, v))
                    if len(v.strip()) > 8: self.titles[re.sub(r'\s+', ' ', v.strip())].append((r, c))

    def find(self, pattern, nth=0):
        p = re.compile(pattern)
        hits = sorted((q for t, qs in self.titles.items() if p.search(t) for q in qs))
        # 수량 할인표 사본은 제목이 같으므로 반드시 걸러낸다 (DISCOUNT_ZONE 주석 참조)
        kept = [q for q in hits if not in_discount_zone(*q)]
        if len(kept) != len(hits):
            self.dropped.append((pattern, len(hits) - len(kept)))
        return kept[nth] if len(kept) > nth else None

    def _is_title_row(self, r, c_from, c_to):
        """해당 행의 지정 열 범위에 '블록 제목'으로 보이는 긴 문자열이 있는가."""
        for c in range(max(1, c_from), min(self.ws.max_column, c_to) + 1):
            v = self.ws.cell(r, c).value
            if isinstance(v, str) and len(v.strip()) > 8: return True
        return False

    def _axis(self, tr, col, r_max, need_suffix=True):
        """지정 열이 직경 축인지 보고 축을 돌려준다.
        · 값 열을 축으로 오인하지 않도록 「이하/미만/이상」 접미가 붙은 칸만 인정한다
          (금액도 숫자라서 band() 를 그냥 통과한다).
        · r_max 로 블록 아래 경계를 반드시 막는다. 안 막으면 아래 블록의 축이 섞여 들어와
          큰 직경이 엉뚱한 행을 읽는다 (2026-09-02 확인)."""
        s = []
        for r in range(tr + 3, min(r_max, self.ws.max_row) + 1):
            if self._is_title_row(r, col, col): break
            b = band(self.ws.cell(r, col).value)
            if b and (b[1] or not need_suffix) and b[0] <= 300: s.append((b[0], b[1], r))
        return s

    def vtable(self, tr, tc, depth=12):
        """세로형: 직경이 제목 열, (비코팅/코팅) → (2날/3,4,6날) 2단 헤더."""
        seg = []
        for r in range(tr + 3, min(tr + 3 + depth, self.ws.max_row + 1)):
            if self._is_title_row(r, tc, tc + 1): break     # 다음 블록 시작
            b = band(self.ws.cell(r, tc).value)
            if b: seg.append((b[0], b[1], r))
        c_end = tc + 8
        for c in range(tc + 1, min(tc + 12, self.ws.max_column) + 1):
            v = self.ws.cell(tr, c).value
            if isinstance(v, str) and len(v.strip()) > 8:   # 오른쪽 옆 블록 시작
                c_end = c - 1; break
        # 🔴 같은 블록 안에서 그룹마다 직경 구간이 다른 경우가 있다.
        #    예) 초경 밑옆날 — 비코팅·일반은 6/10/12/16/20/25/32, 고경도는 6/8/10/12/16/20/25.
        #    블록 대표 축으로 고경도를 읽으면 한 칸씩 밀린다. 그룹 왼쪽의 축 열을 각자 쓴다.
        r_max = seg[-1][2] if seg else tr + 3 + depth
        axes = {tc: seg}
        for c in range(tc + 1, c_end + 1):
            a = self._axis(tr, c, r_max)
            if len(a) >= 3: axes[c] = a
        def axis_for(col):
            cand = [c for c in axes if c <= col]
            return axes[max(cand)] if cand else seg
        cols, cur = {}, None
        for c in range(tc, c_end + 1):
            g, l = self.ws.cell(tr + 1, c).value, self.ws.cell(tr + 2, c).value
            if isinstance(g, str) and g.strip(): cur = g.strip()
            if isinstance(l, str) and BLADE.match(l.strip()) and cur:
                d = cols.setdefault(cur, {})
                d['2' if l.strip() == '2날' else '4'] = c
                d.setdefault('S', axis_for(c))
        return seg, alias_groups(cols)

    def bigtable(self, tr, tc, depth=12):
        """블록 안에 딸린 **대구경 소구간 표**.
        예) 초경 FLAT E/M 밑날 블록의 `30미만`·`31이상` 라벨 + 오른쪽 2열(2날 / 3,4,6날).
        · 라벨 열은 날수 열들보다 오른쪽에 있다 → 그 조건으로 본표와 구분한다
        · **코팅 구분이 없다** (비코팅·일반코팅 품목이 같은 값을 쓴다 — 2026-09-02 실측 확인)
        반환: (구간, {'2': 열, '4': 열}) / 없으면 (None, {})"""
        # ⚠️ 오른쪽 옆 블록까지 넘어가면 옆 블록(고경도)의 소구간 표를 읽는다 → 경계를 먼저 막는다
        c_end = tc + 8
        for c in range(tc + 1, min(tc + 12, self.ws.max_column) + 1):
            v = self.ws.cell(tr, c).value
            if isinstance(v, str) and len(v.strip()) > 8:
                c_end = c - 1; break
        blade_cols = [c for c in range(tc, c_end + 1)
                      if isinstance(self.ws.cell(tr + 2, c).value, str)
                      and BLADE.match(str(self.ws.cell(tr + 2, c).value).strip())]
        if not blade_cols: return None, {}
        c_from = max(blade_cols) + 1
        for c in range(c_from, min(c_end, c_from + 6) + 1):
            seg = []
            for r in range(tr + 1, min(tr + 1 + depth, self.ws.max_row + 1)):
                b = band(self.ws.cell(r, c).value)
                if b and b[0] <= 300 and isinstance(self.ws.cell(r, c).value, str) \
                   and isinstance(self.ws.cell(r, c + 1).value, (int, float)):
                    seg.append((b[0], b[1], r))
            if len(seg) >= 2: return seg, {'2': c + 1, '4': c + 2}
        return None, {}

    def gvn(self, blk, group, n, dia):
        """세로형 블록에서 그룹·날수·직경으로 값 조회 (그룹 전용 직경 축 사용)."""
        if not blk: return None
        seg, cols = blk
        grp = cols.get(group)
        if not grp: return None
        col = grp.get(n)
        if not col: return None
        return self.gv(grp.get('S') or seg, col, dia)

    def htable(self, tr, tc, span=14, depth=10):
        """가로형: 직경 헤더 행을 아래→위로 탐색(공유 헤더 대응), 라벨은 데이터 열 왼쪽."""
        def hdr(r):
            run, best = [], []
            for c in range(tc, min(tc + span + 4, self.ws.max_column + 2)):
                b = band(self.ws.cell(r, c).value) if c <= self.ws.max_column else None
                if b and b[0] <= 300: run.append((b[0], b[1], c))
                else:
                    if len(run) > len(best): best = run
                    run = []
            return best if len(best) >= 3 else None
        seg = None
        for r in list(range(tr + 1, tr + 4)) + list(range(tr - 1, max(0, tr - 9), -1)):
            if 1 <= r <= self.ws.max_row:
                s = hdr(r)
                if s: seg = s; break
        if not seg: return None, {}
        c0 = seg[0][2]
        hdr_r = None
        for v, k, c in seg:
            if c == c0: hdr_r = None
        # 헤더 행 번호 복원 (가로형이므로 seg 는 열 정보만 가짐)
        for r in list(range(tr + 1, tr + 4)) + list(range(tr - 1, max(0, tr - 9), -1)):
            if 1 <= r <= self.ws.max_row and band(self.ws.cell(r, c0).value) \
               and band(self.ws.cell(r, c0).value)[0] == seg[0][0]:
                hdr_r = r; break
        rows, cur, npos = {}, None, [0]
        for r in range(tr + 1, min(tr + 1 + depth, self.ws.max_row + 1)):
            if self._is_title_row(r, tc, c0): break         # 다음 블록 시작
            g = l = None
            for c in (c0 - 2, c0 - 1, tc):
                if c < 1: continue
                v = self.ws.cell(r, c).value
                if isinstance(v, str) and v.strip():
                    if BLADE.match(v.strip()): l = v.strip()
                    elif canon_coat(v.strip()): g = canon_coat(v.strip())
            if g: cur = g
            if r == hdr_r: continue
            if sum(1 for _, _, c in seg if isinstance(self.ws.cell(r, c).value, (int, float))) >= 3:
                key = (cur or '') + ('2' if l == '2날' else '4' if l else '')
                rows.setdefault(key or 'X', r)
                # 라벨 없는 블록 대비 — 데이터 행을 위에서부터 순서대로 기록.
                # 관례상 첫 행 = 2날, 둘째 행 = 4날 (초경 E/M 외경연삭 비코팅 블록의
                # 명시 라벨 및 값 크기 관계로 확인, 2026-09-02)
                rows['P%d' % npos[0]] = r
                npos[0] += 1
        return seg, alias_rows(rows)

    def gh(self, seg, row, dia):
        for v, k, c in seg:
            if in_band(v, k, dia):
                x = self.ws.cell(row, c).value
                return None if isinstance(x, str) else x
        return None

    def gv(self, seg, col, dia):
        for v, k, r in seg:
            if in_band(v, k, dia):
                x = self.ws.cell(r, col).value
                return None if isinstance(x, str) else x
        return None


class Blocks:
    """품목 속성 → 정가표 블록. 못 찾은 블록은 missing 에 남는다."""

    def __init__(self, pl):
        self.pl, self.missing, self.v, self.h = pl, [], {}, {}
        self.at = {}           # key → 정가표에서 실제로 잡힌 (행, 열). --blocks 로 확인
        def V(key, pat, nth=0):
            p = pl.find(pat, nth)
            if p is None: self.missing.append(pat)
            else: self.v[key] = pl.vtable(*p); self.at[key] = p
        def H(key, pat, nth=0):
            p = pl.find(pat, nth)
            if p is None: self.missing.append(pat)
            else: self.h[key] = pl.htable(*p); self.at[key] = p
        # 초경 라핑 — 세로형
        V(('초경','L코너볼','밑날','일반'),     r'초경 라핑 코너/볼.*밑날.*일반')
        V(('초경','L코너볼','밑날','고경도'),   r'초경 라핑 코너/볼.*밑날.*고경도')
        V(('초경','L평','밑날','일반'),         r'초경 라핑 평.*밑날.*일반')
        V(('초경','L평','밑날','고경도'),       r'초경 라핑 평.*밑날.*고경도')
        V(('초경','L코너볼','밑골수리','일반'),   r'초경 라핑 코너/볼.*밑골수리.*일반')
        V(('초경','L코너볼','밑골수리','고경도'), r'초경 라핑 코너/볼.*밑골수리.*고경도')
        # 2026-09-02: 평 밑골수리 우측 제목이 (고경도 코팅) 으로 정정됨 → nth 추정 불필요
        V(('초경','L평','밑골수리','일반'),     r'초경 라핑 평.*밑골수리', 0)
        V(('초경','L평','밑골수리','고경도'),   r'초경 라핑 평.*밑골수리.*고경도')
        # HSS — 가로형
        H(('HSS','L평','밑날','비코팅'),        r'^HSS 라핑 E/M 외경연삭$')
        H(('HSS','L평','밑날','코팅'),          r'^HSS 라핑 E/M 외경연삭 코팅$')
        H(('HSS','L코너볼','밑날','비코팅'),     r'HSS 코너R.*BALL.*밑날$')
        H(('HSS','L코너볼','밑날','코팅'),       r'HSS 코너R.*BALL.*밑날 코팅$')
        H(('HSS','L평','밑골수리','비코팅'),     r'^(HSS )?라핑 (평 )?밑골수리 비코팅$')
        H(('HSS','L평','밑골수리','코팅'),       r'^(HSS )?라핑 (평 )?밑골수리 코팅$')
        H(('HSS','L코너볼','밑골수리','비코팅'), r'HSS 라핑 BALL.*밑골수리 비코팅')
        H(('HSS','L코너볼','밑골수리','코팅'),   r'HSS 코너R.*밑골수리 코팅')
        # 골수리 = 외경연삭 (형상 무관 공통). 2026-09-02 골수리 887건 전수 일치로 확인
        # ── 비라핑 (일반 품목) ─────────────────────────────
        # ⚠️ 같은 제목이 c1/c15/c21/c27 에 4번 반복된다. c1 만 「정가」이고
        #    나머지는 月수량 구간별 할인 적용가다(3행 라벨). nth=0 = 가장 왼쪽 = c1.
        # ⚠️ 제목 뒤 `(일반코팅)` 표기는 개정마다 붙었다 떨어진다 → 선택 항목으로 둔다.
        #    할인표 사본과 제목이 같으므로 find() 의 DISCOUNT_ZONE 배제가 함께 있어야 안전하다.
        V(('초경','코너볼','밑날','일반'),   r'^초경 BALL E/M, 코너R E/M 밑날( \(일반코팅\))?$', 0)
        V(('초경','코너볼','밑날','고경도'), r'^BALL, 코너R 밑날 \(고경도', 0)
        V(('초경','평','밑날','일반'),       r'^초경 FLAT E/M 밑날( \(일반코팅\))?$', 0)
        V(('초경','평','밑날','고경도'),     r'^초경 FLAT E/M 밑날 \(고경도', 0)
        V(('초경','평','밑옆날','코팅'),     r'초경 FLAT E/M 밑옆날')
        V(('초경','코너볼','밑옆날','코팅'), r'초경 BALL E/M, 코너R E/M 밑옆날')
        # 외경연삭 = 단일 가공(밑날만/옆날만/골수리만) 단가 [사내 확인 2026-09-02]
        H(('초경','*','외경','비코팅'), r'^초경 E/M 외경연삭$')
        H(('초경','*','외경','일반'),   r'^초경 E/M 외경연삭_일반코팅$')
        H(('초경','*','외경','고경도'), r'^초경 E/M 외경연삭_고경도코팅$')
        H(('HSS','*','외경','비코팅'),  r'^HSS E/M 외경연삭$')
        H(('HSS','*','외경','코팅'),    r'^HSS E/M 외경연삭 코팅$')
        H(('HSS','평','밑옆날','비코팅'),   r'^HSS 밑옆날 비코팅$')
        H(('HSS','평','밑옆날','코팅'),     r'^HSS 밑옆날 코팅$')
        H(('HSS','코너볼','밑옆날','비코팅'), r'^HSS 코너R E/M, BALL E/M 밑옆날 비코팅$')
        H(('HSS','코너볼','밑옆날','코팅'),   r'HSS 코너R E/M, BALL E/M,\s+밑옆날 코팅')
        # 리머 — 형상 계열이 따로다 (평·코너볼 어디에도 안 들어간다)
        V(('초경','리머','밑날','일반'), r'^초경 리머 E/M 밑날( \(일반코팅\))?$')
        H(('HSS','리머','밑날','비코팅'), r'^HSS 리머 E/M 밑날 비코팅$')
        H(('HSS','리머','밑날','코팅'),   r'^HSS 리머 E/M 밑날 코팅$')
        # 초경 평 밑날 블록에 딸린 대구경 소구간 표 (30미만/31이상 × 2날/3,4,6날)
        pos = pl.find(r'^초경 FLAT E/M 밑날( \(일반코팅\))?$', 0)
        if pos: self.v[('초경','평','밑날','대구경')] = pl.bigtable(*pos); self.at[('초경','평','밑날','대구경')] = pos
        H(('초경','*','골수리','비코팅'), r'^초경 라핑 E/M 외경연삭$')
        H(('초경','*','골수리','코팅'),   r'^초경 라핑 E/M 외경연삭 코팅$')
        H(('HSS','*','골수리','비코팅'),  r'^HSS 라핑 E/M 외경연삭$')
        H(('HSS','*','골수리','코팅'),    r'^HSS 라핑 E/M 외경연삭 코팅$')

    def expect(self, mat, shape, part, coat, blade, dia, **_):
        """기대 정가. 매핑 대상이 아니면 None (= 검사 제외)."""
        n = '2' if blade == 2 else '4'          # ★ 3날 이상은 전부 4날 행
        # 🔴 「밑옆날 = 밑골수리 = 밑외경」 은 같은 가공이지만, 정가표 블록 제목은
        #    라핑만 「밑골수리」로 개명됐고 비라핑은 여전히 「밑옆날」이다 (2026-09-01 개정).
        #    라핑 기준으로 전부 밑골수리로 바꾸면 비라핑 밑옆날 블록을 못 찾는다.
        pt = part
        if part in ('밑골수리', '밑옆날', '밑골수리날'):
            pt = '밑골수리' if shape.startswith('라핑') else '밑옆날'
        lap = shape.startswith('라핑')
        fam = 'L코너볼' if shape in ('라핑볼', '라핑코너') else 'L평'
        # ── 비라핑(일반 품목) ─────────────────────────────
        if not lap:
            # 리머 — 전용 블록. 초경은 세로형(비코팅/코팅), HSS는 가로형(비코팅/코팅 블록 분리)
            if shape == '리머' and pt == '밑날':
                if mat == '초경':
                    g = '비코팅' if coat == '비코팅' else '코팅'
                    return self.pl.gvn(self.v.get((mat, '리머', '밑날', '일반')), g, n, dia)
                k = (mat, '리머', '밑날', '비코팅' if coat == '비코팅' else '코팅')
                if k in self.h and self.h[k][0]:
                    seg, rows = self.h[k]
                    r = rows.get(n) or rows.get('P0' if n == '2' else 'P1')
                    if r: return self.pl.gh(seg, r, dia)
                return None
            fam2 = '코너볼' if shape in ('볼', '코너') else '평' if shape == '평' else None
            if fam2:
                # 초경 밑날 — 전용 블록
                if mat == '초경' and pt == '밑날':
                    k = (mat, fam2, '밑날', '고경도' if coat == '고경도코팅' else '일반')
                    g = '비코팅' if coat == '비코팅' else '코팅'
                    e = self.pl.gvn(self.v.get(k), g, n, dia)
                    if e is not None: return e
                    # 본표 최대 구간(25이하)을 넘는 평 품목은 대구경 소구간 표를 쓴다
                    big = self.v.get((mat, '평', '밑날', '대구경'))
                    if fam2 == '평' and big and big[0]:
                        seg, cols = big
                        return self.pl.gv(seg, cols[n], dia)
                    return None
                # 초경 밑옆날 — 한 블록에 비코팅/일반/고경도가 모두 있다
                if mat == '초경' and pt == '밑옆날':
                    k = (mat, fam2, '밑옆날', '코팅')
                    g = {'비코팅': '비코팅', '일반코팅': '일반', '고경도코팅': '고경도'}.get(coat)
                    return self.pl.gvn(self.v.get(k), g, n, dia) if g else None
                # HSS 밑옆날 — 형상별 전용 블록
                if mat == 'HSS' and pt == '밑옆날':
                    k = (mat, fam2, '밑옆날', '비코팅' if coat == '비코팅' else '코팅')
                    if k in self.h and self.h[k][0]:
                        seg, rows = self.h[k]
                        r = rows.get(n) if coat == '비코팅' else \
                            rows.get(('일반' if coat == '일반코팅' else '고경도') + n)
                        if r: return self.pl.gh(seg, r, dia)
                    return None
                # HSS 코너·볼 밑날 — 라핑과 같은 블록을 쓴다.
                # 블록 제목이 「HSS 코너R E/M, BALL E/M, 라핑 BALL E/M 밑날」로
                # 비라핑·라핑을 함께 담고 있다 [한경준님 확인, 2026-09-02].
                if mat == 'HSS' and pt == '밑날' and fam2 == '코너볼':
                    k = (mat, 'L코너볼', '밑날', '비코팅' if coat == '비코팅' else '코팅')
                    if k in self.h and self.h[k][0]:
                        seg, rows = self.h[k]
                        if coat == '비코팅':
                            r = rows.get(n) or rows.get('비코팅' + n) or rows.get('P0' if n == '2' else 'P1')
                        else:
                            g = '일반' if coat == '일반코팅' else '고경도'
                            r = rows.get(g + n) or rows.get(g)
                        if r: return self.pl.gh(seg, r, dia)
                    return None
                # 단일 가공(밑날만/옆날만/골수리만) = 외경연삭 표 [사내 확인 2026-09-02]
                if pt in ('옆날', '골수리') or (mat == 'HSS' and pt == '밑날' and fam2 == '평'):
                    if mat == '초경':
                        g = {'비코팅': '비코팅', '일반코팅': '일반', '고경도코팅': '고경도'}.get(coat)
                        k = (mat, '*', '외경', g)
                    else:
                        k = (mat, '*', '외경', '비코팅' if coat == '비코팅' else '코팅')
                    if k in self.h and self.h[k][0]:
                        seg, rows = self.h[k]
                        if mat == '초경':
                            r = rows.get(n) or rows.get('P0' if n == '2' else 'P1') or rows.get('X')
                        else:
                            r = rows.get(n) if coat == '비코팅' else \
                                rows.get(('일반' if coat == '일반코팅' else '고경도') + n)
                        if r: return self.pl.gh(seg, r, dia)
                    return None
            return None

        # 골수리 = 외경연삭. 형상 무관 공통이고 날수 구분도 없다.
        # 옆날도 외경연삭 표를 쓴다 — 주석 "밑날가격 = 옆날가격" 과 같은 계열
        if lap and pt in ('골수리', '옆날'):
            k = (mat, '*', '골수리', '비코팅' if coat == '비코팅' else '코팅')
            if k not in self.h: return None
            seg, rows = self.h[k]
            if not seg: return None
            if coat == '비코팅':
                r = rows.get('X') or rows.get('비코팅') or rows.get('4') or rows.get('2')
            else:
                g = '일반' if coat == '일반코팅' else '고경도'
                r = rows.get(g) or rows.get(g + '4') or rows.get(g + '2')
            return self.pl.gh(seg, r, dia) if r else None
        if mat == '초경' and lap and pt in ('밑날', '밑골수리'):
            k = (mat, fam, pt, '고경도' if coat == '고경도코팅' else '일반')
            g = '비코팅' if coat == '비코팅' else '코팅'
            return self.pl.gvn(self.v.get(k), g, n, dia)
        # HSS 라핑 밑날은 전용 블록이 없다 — 주석 "밑날가격 = 옆날가격" 에 따라 외경연삭 블록 사용
        if mat == 'HSS' and (lap or (shape in ('코너', '볼') and pt == '밑날' and coat == '비코팅')):
            k = (mat, fam, pt, '비코팅' if coat == '비코팅' else '코팅')
            if k not in self.h: return None
            seg, rows = self.h[k]
            if not seg: return None
            if coat == '비코팅':
                r = rows.get('비코팅' + n) or rows.get(n) or rows.get('비코팅') or rows.get('X')
            else:
                g = '일반' if coat == '일반코팅' else '고경도'
                r = rows.get(g + n) or rows.get(g + '4') or rows.get(g) or rows.get('X')
            return self.pl.gh(seg, r, dia) if r else None
        return None


def parse_name(nm):
    p = [x.strip() for x in str(nm).split('/')]
    if len(p) < 6: return None
    m, b = re.match(r'^([\d.]+)', p[2]), re.match(r'^(\d+)', p[0])
    if not m or not b: return None
    return dict(blade=int(b.group(1)), shape=p[1], dia=float(m.group(1)), mat=p[3],
                part=p[4], coat=p[5].replace(' ', ''), dia_txt=m.group(1))


def load_items(paths):
    rows = []
    for path in paths:
        base = os.path.basename(path)
        m = re.search(r'수정_([^_]+)_', base)
        label = m.group(1) if m else os.path.splitext(base)[0][:8]
        ws = openpyxl.load_workbook(path, data_only=True)[SHEET_ITEMS]
        for r in range(ROW_START, ws.max_row + 1):
            code, nm, pr = ws.cell(r, COL_CODE).value, ws.cell(r, COL_NAME).value, ws.cell(r, COL_PRICE).value
            if not code and not nm: continue
            x = dict(f=label, r=r, code=code, nm=nm, price=pr, path=path)
            if nm: x['p'] = parse_name(nm)
            rows.append(x)
    return rows


def audit(rows, blocks):
    I, seen, checked = collections.OrderedDict(), collections.defaultdict(list), 0
    uncovered = collections.Counter()
    def add(t, x): I.setdefault(t, []).append(x)
    for x in rows:
        nm = str(x['nm'])
        if not x['code']: add('A1 품목코드 결측', x); continue
        seen[(x['f'], x['code'])].append(x['r'])
        if not x.get('p'):
            if not nm.startswith(NON_STANDARD_NAME_PREFIX): add('A2 품목명 형식오류', x)
            continue
        p = x['p']
        if x['price'] in (None, ''): add('A3 정가 결측', x); continue
        try: v = float(x['price'])
        except (TypeError, ValueError): add('A4 정가 비숫자', x); continue
        if v <= 0: add('A5 정가 0 이하', x)
        if round(v) % 10 != 0: add('A6 10원 단위 아님', x)
        cm = re.match(r'^[A-Z]?(\d)([A-Z]{2})(\d{3})', str(x['code']))
        if cm:
            if abs(int(cm.group(3)) / 10 - p['dia']) > 1e-6 and p['dia_txt'] not in CODE_DIA_TOLERATED:
                add('A7 코드↔품목명 직경 불일치', x)
            if int(cm.group(1)) != p['blade']: add('A8 코드↔품목명 날수 불일치', x)
        # 재질·코팅 표기 이상 — 표기가 틀리면 블록 조회가 조용히 실패해 검사에서 빠진다
        if p['mat'] not in ('초경', 'HSS'):
            add('A9 재질 표기 이상', x)
        if p['coat'] not in ('비코팅', '일반코팅', '고경도코팅'):
            add('A10 코팅 표기 이상', x)
        e = blocks.expect(**p)
        if e is None:
            uncovered[(p['mat'], p['shape'], p['part'])] += 1
        else:
            checked += 1
            if abs(v - float(e)) > 0.5:
                y = dict(x); y['exp'] = round(float(e)); add('B 정가표 대조 불일치', y)
    for k, v in seen.items():
        if len(v) > 1:
            add('A9 품목코드 중복', {'f': k[0], 'r': '', 'code': k[1], 'nm': f'행 {v}', 'price': ''})

    bn = collections.defaultdict(dict)
    for x in rows:
        if x.get('p') and x['price'] is not None:
            bn[str(x['nm']).replace(' ', '')][x['f']] = (x['r'], x['price'])
    for nm, d in bn.items():
        if len(d) > 1 and len({t[1] for t in d.values()}) > 1:
            add('C 파일 간 정가 불일치', {'f': '/'.join(d), 'r': '', 'code': '', 'nm': nm,
                'price': ' vs '.join(f'{a}:{b[1]}' for a, b in d.items())})

    grp = collections.defaultdict(list)
    for x in rows:
        if x.get('p') and isinstance(x['price'], (int, float)):
            q = x['p']
            grp[(x['f'], q['blade'], q['shape'], q['mat'], q['part'], q['coat'])].append((q['dia'], x))
    for lst in grp.values():
        lst.sort(key=lambda t: t[0]); prev = None
        for dia, x in lst:
            if prev and x['price'] < prev[1] - 1e-6 and dia > prev[0] + 1e-9:
                add('D 가격 역전', {'f': x['f'], 'r': x['r'], 'code': x['code'], 'nm': x['nm'],
                    'price': f"Ø{prev[0]:g} {prev[1]:.0f} -> Ø{dia:g} {x['price']:.0f}"})
            if prev is None or x['price'] >= prev[1]: prev = (dia, x['price'])

    ck = collections.defaultdict(dict)
    for x in rows:
        if x.get('p') and isinstance(x['price'], (int, float)):
            q = x['p']
            ck[(x['f'], q['blade'], q['shape'], q['mat'], q['part'], q['dia'])][q['coat']] = (x['r'], x['price'], x['nm'])
    for g, d in ck.items():
        if (g[2], g[3], g[4], g[5]) in EXCEPTIONS_E2: continue
        if '비코팅' in d and '일반코팅' in d and d['비코팅'][1] > d['일반코팅'][1]:
            add('E1 비코팅 > 일반', {'f': g[0], 'r': d['일반코팅'][0], 'code': '', 'nm': d['일반코팅'][2],
                'price': f"비{d['비코팅'][1]:.0f} > 일반{d['일반코팅'][1]:.0f}"})
        if '일반코팅' in d and '고경도코팅' in d and d['일반코팅'][1] > d['고경도코팅'][1]:
            add('E2 일반 > 고경도', {'f': g[0], 'r': d['고경도코팅'][0], 'code': '', 'nm': d['고경도코팅'][2],
                'price': f"일반{d['일반코팅'][1]:.0f} > 고경도{d['고경도코팅'][1]:.0f}"})
    return I, checked, uncovered


def save_xlsx(I, base):
    from openpyxl.styles import Font, PatternFill
    F = '맑은 고딕'
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '검수결과'
    ws['A1'] = f'재연마 단가표 검수 결과 ({now_kst():%Y-%m-%d %H:%M} KST)'
    ws['A1'].font = Font(name=F, size=22, bold=True); ws.merge_cells('A1:G1')
    hdr = ('유형', '파일', '행', '품목코드', '품목명', '값', '정가표')
    for i, h in enumerate(hdr, 1):
        c = ws.cell(3, i, h); c.font = Font(name=F, size=13, bold=True)
        c.fill = PatternFill('solid', fgColor='FFEDEDED')
    r = 4
    for t, v in I.items():
        for x in v:
            for i, val in enumerate([t, x['f'], x['r'], x.get('code', ''), x['nm'],
                                     x['price'], x.get('exp', '')], 1):
                ws.cell(r, i, val).font = Font(name=F, size=13)
            r += 1
    for col, wd in zip('ABCDEFG', (26, 7, 7, 14, 46, 24, 10)):
        ws.column_dimensions[col].width = wd
    ws.freeze_panes = 'A4'
    if r > 4: ws.auto_filter.ref = f'A3:G{r-1}'
    out = os.path.join(base, f'검수결과_{now_kst():%Y%m%d_%H%M}.xlsx')
    wb.save(out); return out


def main():
    ap = argparse.ArgumentParser(description='재연마 단가표 전수 검수')
    ap.add_argument('--dir', default='raw/단가표', help='작업 폴더 (기본 raw/단가표)')
    ap.add_argument('--pricelist', default=None, help='정가표 파일 경로 (기본 자동 탐색)')
    ap.add_argument('--sheet', default=None, help='정가표 시트명 (기본 최신 정가표* 시트)')
    ap.add_argument('--pattern', default=r'검수표시\.xlsx$', help='품목 파일 파일명 정규식')
    ap.add_argument('--json', default=None, help='결과 JSON 저장 경로')
    ap.add_argument('--xlsx', action='store_true', help='결과를 xlsx로도 저장')
    ap.add_argument('--blocks', action='store_true',
                    help='등록된 블록이 정가표에서 실제로 어디에 잡혔는지 출력 (오참조 점검)')
    a = ap.parse_args()

    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    base = a.dir if os.path.isabs(a.dir) else os.path.join(root, a.dir)
    if not os.path.isdir(base): sys.exit(f'폴더 없음: {base}')
    files = [f for f in sorted(os.listdir(base)) if not f.startswith('~$')]
    pl_path = a.pricelist or next((os.path.join(base, f) for f in files
                                   if '정가표' in f and f.endswith('.xlsx')), None)
    if not pl_path: sys.exit('정가표 파일을 찾지 못했습니다')
    pat = re.compile(a.pattern)
    items = [os.path.join(base, f) for f in files if pat.search(f)]
    if not items: sys.exit(f'품목 파일 없음 (패턴 {a.pattern})')

    print(f'정가표   : {os.path.basename(pl_path)}')
    pl = PriceList(pl_path, a.sheet)
    print(f'  시트   : {pl.name}')
    blocks = Blocks(pl)
    if blocks.missing:
        print(f'  ⚠️ 못 찾은 블록 {len(blocks.missing)}개 — 제목이 바뀌었을 수 있습니다')
        for m in blocks.missing: print(f'       {m}')
    if pl.dropped:
        print(f'  ℹ️ 할인표 구역에서 걸러낸 제목 {len(pl.dropped)}건 (정상 동작)')
    if a.blocks:
        print('  ── 등록된 블록 위치 ──')
        for k in sorted(blocks.at, key=lambda x: blocks.at[x]):
            r, c = blocks.at[k]
            print(f'       {get_column_letter(c)}{r:<5} {"·".join(k)}')
    if pl.errors:
        print('  🔴 정가표 오류값 %d곳: %s' % (len(pl.errors),
              ', '.join(f'{get_column_letter(c)}{r}' for r, c, _ in pl.errors)))
    print('품목파일 : ' + ', '.join(os.path.basename(p) for p in items))

    rows = load_items(items)
    print(f'\n총 품목 {len(rows)}건  {dict(collections.Counter(x["f"] for x in rows))}')
    I, checked, uncovered = audit(rows, blocks)
    unc = sum(uncovered.values())
    print(f'정가표 대조 {checked}건 / 미대조 {unc}건  (미대조는 B검사에서 빠지고 A·C·D·E만 적용)')
    if unc:
        print('  미대조 상위: ' + ', '.join(f'{k[0]}·{k[1]}·{k[2]} {v}' for k, v in uncovered.most_common(5)))
        print('  ↳ 「골수리(=외경연삭)」 등 아직 블록을 매핑하지 않은 계열입니다. 넓히려면 Blocks 에 등록하세요.')
    print()

    print('=== 검사 결과 ===')
    if not I:
        print('  ✅ 이상 0건')
    for t, v in I.items():
        print(f'  {t}: {len(v)}건')
        for x in v[:8]:
            ex = f"  -> 정가표 {x['exp']}" if 'exp' in x else ''
            print(f"      {x['f']} r{x['r']} {str(x['nm'])[:44]:44s} {x['price']}{ex}")
        if len(v) > 8: print(f'      … 외 {len(v)-8}건')

    if a.json:
        json.dump({t: [{k: str(w) for k, w in x.items() if k != 'p'} for x in v] for t, v in I.items()},
                  open(a.json, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
        print(f'\nJSON 저장: {a.json}')
    if a.xlsx:
        print(f'\nxlsx 저장: {save_xlsx(I, base)}')

    total = sum(len(v) for v in I.values())
    print(f'\n합계 {total}건' + ('   ✅ 통과' if total == 0 else ''))
    # 종료코드: 값 관련 이상(B·C·D·E)이 있으면 1
    return 1 if any(t[0] in 'BCDE' for t in I) else 0


if __name__ == '__main__':
    sys.exit(main())
