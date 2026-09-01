import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
import warnings
warnings.filterwarnings('ignore')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HDR_BG  = "1F497D"
HDR_FG  = "FFFFFF"
TOP5_BG = "FFFF00"
TOP10_BG= "FFF2CC"
TOT_BG  = "D9D9D9"
FONT    = "맑은 고딕"

thin   = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_style(cell, center=True):
    cell.font      = Font(name=FONT, bold=True, color=HDR_FG, size=10)
    cell.fill      = PatternFill("solid", fgColor=HDR_BG)
    cell.alignment = Alignment(horizontal='center' if center else 'left',
                               vertical='center', wrap_text=True)
    cell.border    = border

def tot_style(cell, center=True):
    cell.font      = Font(name=FONT, bold=True, size=10)
    cell.fill      = PatternFill("solid", fgColor=TOT_BG)
    cell.alignment = Alignment(horizontal='center' if center else 'left',
                               vertical='center')
    cell.border    = border

def data_style(cell, center=True, bg=None):
    cell.font      = Font(name=FONT, size=10)
    if bg:
        cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal='center' if center else 'left',
                               vertical='center')
    cell.border    = border

NUM_FMT  = '#,##0'
PCT_FMT  = '0.0%'
DASH_FMT = '#,##0;-#,##0;"-"'
MONTH_KR = ['1월','2월','3월','4월','5월','6월',
            '7월','8월','9월','10월','11월','12월']


def make_year(year):
    import glob as _glob
    yr2 = str(year)[2:]
    src = f'raw/출하현황/{yr2}년_출하현황.xlsx'

    frames = [pd.read_excel(src, sheet_name='출하현황', header=0)]

    # 추가 월별 파일 (예: 26년_출하현황_4월.xlsx) 자동 병합
    for extra in sorted(_glob.glob(f'raw/출하현황/{yr2}년_출하현황_*.xlsx')):
        ex = pd.read_excel(extra, sheet_name='출하현황', header=0)
        frames.append(ex)
        print(f"    + 추가 파일 병합: {extra}  ({len(ex)}행)")

    df = pd.concat(frames, ignore_index=True)
    df['출하일자'] = pd.to_datetime(df['출하일자'], errors='coerce')
    df['월'] = df['출하일자'].dt.month

    pivot = df.pivot_table(index='납품처명', columns='월', values='출고',
                           aggfunc='sum', fill_value=0)
    for m in range(1, 13):
        if m not in pivot.columns:
            pivot[m] = 0
    pivot = pivot[[m for m in range(1, 13)]]
    pivot['연간합계'] = pivot.sum(axis=1)
    pivot = pivot.sort_values('연간합계', ascending=False).reset_index()
    pivot.columns = ['납품처명'] + [f'{m}월' for m in range(1, 13)] + ['연간합계']

    # 데이터 있는 월 확인 (합계 > 0)
    active_months = [m for m in range(1, 13) if pivot[f'{m}월'].sum() > 0]

    wb = Workbook()

    # ── 요약 시트 ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "요약"
    ws.freeze_panes = "B3"
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 28

    month_cols = [f'{m}월' for m in range(1, 13)]
    headers    = ['순위', '납품처명'] + month_cols + ['연간합계', '비율(%)']
    n_col      = len(headers)

    # 제목
    ws.merge_cells(f'A1:{get_column_letter(n_col)}1')
    tc = ws['A1']
    tc.value     = f"{year}년 출하현황 — 납품처별 월별 출고량  (단위: 개)"
    tc.font      = Font(name=FONT, bold=True, size=13, color=HDR_FG)
    tc.fill      = PatternFill("solid", fgColor=HDR_BG)
    tc.alignment = Alignment(horizontal='center', vertical='center')

    # 헤더
    for ci, h in enumerate(headers, 1):
        hdr_style(ws.cell(row=2, column=ci, value=h), center=(ci != 2))

    # 데이터
    tot_er = len(pivot) + 3
    for ri, row in pivot.iterrows():
        er   = ri + 3
        rank = ri + 1
        bg   = TOP5_BG if rank <= 5 else (TOP10_BG if rank <= 10 else None)

        ws.cell(row=er, column=1, value=rank);            data_style(ws.cell(row=er, column=1), bg=bg)
        ws.cell(row=er, column=2, value=row['납품처명']); data_style(ws.cell(row=er, column=2), center=False, bg=bg)

        for ci, mc in enumerate(month_cols, 3):
            c = ws.cell(row=er, column=ci, value=row[mc])
            data_style(c, bg=bg); c.number_format = DASH_FMT

        c = ws.cell(row=er, column=15, value=f'=SUM(C{er}:N{er})')
        data_style(c, bg=bg); c.number_format = NUM_FMT

        c = ws.cell(row=er, column=16, value=f'=O{er}/O${tot_er}')
        data_style(c, bg=bg); c.number_format = PCT_FMT

    # 합계 행
    ws.merge_cells(f'A{tot_er}:B{tot_er}')
    tot_style(ws.cell(row=tot_er, column=1, value='합계'))
    tot_style(ws.cell(row=tot_er, column=2))
    for ci in range(3, 16):
        cl = get_column_letter(ci)
        c  = ws.cell(row=tot_er, column=ci, value=f'=SUM({cl}3:{cl}{tot_er-1})')
        tot_style(c); c.number_format = NUM_FMT
    c = ws.cell(row=tot_er, column=16, value=1.0)
    tot_style(c); c.number_format = PCT_FMT

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 28
    for ci in range(3, 15):
        ws.column_dimensions[get_column_letter(ci)].width = 7
    ws.column_dimensions['O'].width = 10
    ws.column_dimensions['P'].width = 9

    # ── 월별 시트 (데이터 있는 월만) ──────────────────────────
    for m in range(1, 13):
        mc    = f'{m}월'
        mdata = pivot[['납품처명', mc, '연간합계']].copy()
        mdata = mdata[mdata[mc] > 0].sort_values(mc, ascending=False).reset_index(drop=True)

        ws_m = wb.create_sheet(title=MONTH_KR[m-1])
        ws_m.freeze_panes = "B3"
        ws_m.row_dimensions[1].height = 18
        ws_m.row_dimensions[2].height = 28

        m_hdrs = ['순위', '납품처명', f'{m}월 출고(개)', '월간비율(%)', '누적비율(%)', '연간합계(개)']

        ws_m.merge_cells('A1:F1')
        tc = ws_m['A1']
        if m not in active_months:
            tc.value = f"{year}년 {m}월 — 출하 데이터 없음"
        else:
            tc.value = f"{year}년 {m}월 출하현황 — 납품처별 출고량  (단위: 개)"
        tc.font      = Font(name=FONT, bold=True, size=13, color=HDR_FG)
        tc.fill      = PatternFill("solid", fgColor=HDR_BG)
        tc.alignment = Alignment(horizontal='center', vertical='center')

        for ci, h in enumerate(m_hdrs, 1):
            hdr_style(ws_m.cell(row=2, column=ci, value=h), center=(ci != 2))

        if len(mdata) == 0:
            ws_m.merge_cells('A3:F3')
            c = ws_m.cell(row=3, column=1, value='해당 월 출하 데이터 없음')
            c.font      = Font(name=FONT, size=10, color='888888')
            c.alignment = Alignment(horizontal='center', vertical='center')
        else:
            tot_row_m = len(mdata) + 3
            for ri, row in mdata.iterrows():
                er   = ri + 3
                rank = ri + 1
                bg   = TOP5_BG if rank <= 5 else None

                ws_m.cell(row=er, column=1, value=rank);            data_style(ws_m.cell(row=er, column=1), bg=bg)
                ws_m.cell(row=er, column=2, value=row['납품처명']); data_style(ws_m.cell(row=er, column=2), center=False, bg=bg)

                c = ws_m.cell(row=er, column=3, value=row[mc])
                data_style(c, bg=bg); c.number_format = NUM_FMT

                c = ws_m.cell(row=er, column=4, value=f'=C{er}/C${tot_row_m}')
                data_style(c, bg=bg); c.number_format = PCT_FMT

                prev = f'=E{er-1}+D{er}' if ri > 0 else f'=D{er}'
                c = ws_m.cell(row=er, column=5, value=prev)
                data_style(c, bg=bg); c.number_format = PCT_FMT

                c = ws_m.cell(row=er, column=6, value=row['연간합계'])
                data_style(c, bg=bg); c.number_format = NUM_FMT

            # 합계 행
            ws_m.merge_cells(f'A{tot_row_m}:B{tot_row_m}')
            tot_style(ws_m.cell(row=tot_row_m, column=1, value='합계'))
            tot_style(ws_m.cell(row=tot_row_m, column=2))
            c = ws_m.cell(row=tot_row_m, column=3, value=f'=SUM(C3:C{tot_row_m-1})')
            tot_style(c); c.number_format = NUM_FMT
            c = ws_m.cell(row=tot_row_m, column=4, value=1.0)
            tot_style(c); c.number_format = PCT_FMT
            tot_style(ws_m.cell(row=tot_row_m, column=5))
            c = ws_m.cell(row=tot_row_m, column=6, value=f'=SUM(F3:F{tot_row_m-1})')
            tot_style(c); c.number_format = NUM_FMT

        ws_m.column_dimensions['A'].width = 6
        ws_m.column_dimensions['B'].width = 28
        ws_m.column_dimensions['C'].width = 13
        ws_m.column_dimensions['D'].width = 11
        ws_m.column_dimensions['E'].width = 11
        ws_m.column_dimensions['F'].width = 13

    out = f'wiki/comparisons/출하현황_납품처별_월별분석_{year}.xlsx'
    wb.save(out)
    print(f"  저장 완료: {out}  (활성 월: {active_months})")


if __name__ == '__main__':
    print("=== 연도별 출하현황 분석 파일 생성 ===")
    for yr in [2022, 2023, 2024, 2026]:
        print(f"\n[{yr}년] 처리 중...")
        make_year(yr)
    print("\n모든 파일 생성 완료!")
