"""
납기 알림 자동화  v2.0
──────────────────────
매일 오전 9시 실행 (Windows 작업 스케줄러)
- ERP에서 수주·생산실적 데이터 조회
- 납기 D-3 이내 항목 필터
- Excel 생성 (긴급 항목 빨간 강조)
- Gmail API (OAuth)로 자동 발송  ← v2.0: SMTP 제거, Gmail API 사용

사전 준비 (최초 1회):
  python setup_gmail_token.py  ← 브라우저에서 권한 허용 후 자동 저장
"""

import os, sys, base64, io

# Windows 콘솔 CP949 환경에서도 한글·이모지 정상 출력
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
from datetime import date, datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from trico_client import TricoClient

# ── 설정 ──────────────────────────────────────────────────────────────────
GMAIL_USER   = "hzn2001@toolkorea.co.kr"
TO_EMAIL     = "hzn2001@toolkorea.co.kr"
DLV_DAYS     = 3   # D-N 이내 납기 강조

TOKEN_FILE   = Path(__file__).parent / "token_gmail.json"
SECRET_FILE  = Path(__file__).parent.parent / "client_secret_gmail.json"
SCOPES       = ["https://www.googleapis.com/auth/gmail.send"]

SAVE_DIR = Path(__file__).parent.parent / "outputs" / "dlv_alerts"
SAVE_DIR.mkdir(parents=True, exist_ok=True)

# 완료 상태 코드 (stat_bc) — 이 코드인 수주는 알림에서 제외
# ⚠️ ERP stat_bc 완료 코드를 확인 후 수정하세요 (현재 추정값)
DONE_STAT = {"90", "99", "완료"}


# ══════════════════════════════════════════════════════════════════════════════
def get_gmail_service():
    """Gmail API 서비스 객체 반환. 토큰 없으면 안내 후 종료."""
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    if not TOKEN_FILE.exists():
        print("⚠️  Gmail 토큰 없음. 먼저 실행하세요:")
        print(f"   python {Path(__file__).parent / 'setup_gmail_token.py'}")
        sys.exit(1)

    creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)

    # 토큰 만료 시 자동 갱신
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
        TOKEN_FILE.write_text(creds.to_json(), encoding="utf-8")

    return build("gmail", "v1", credentials=creds)


# ══════════════════════════════════════════════════════════════════════════════
def fetch_data():
    client = TricoClient()
    today  = date.today()
    first  = today.replace(day=1).strftime("%Y-%m-%d")
    # 수주는 2개월치
    fr2m   = (today.replace(day=1) - timedelta(days=1)).replace(day=1).strftime("%Y-%m-%d")

    print("  수주 조회...")
    df_so = filter_active(client.수주(fr_dt=fr2m))

    print("  생산실적 조회...")
    df_pp = client.생산실적(fr_dt=first)

    print("  출하 조회...")
    df_sh = client.출하(fr_dt=first)

    return df_so, df_pp, df_sh


def parse_date(val):
    if not val or str(val).strip() in ("", "None", "NaN"):
        return None
    s = str(val).split("T")[0]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def filter_active(df: pd.DataFrame) -> pd.DataFrame:
    """완료(stat_bc ∈ DONE_STAT) 수주 제외"""
    if df.empty or "stat_bc" not in df.columns:
        return df
    before = len(df)
    df = df[~df["stat_bc"].astype(str).isin(DONE_STAT)].copy()
    removed = before - len(df)
    if removed:
        print(f"  [필터] 완료 상태 {removed}건 제외")
    return df


def flag_urgent(df: pd.DataFrame, date_col: str, days: int) -> pd.DataFrame:
    """납기 D-N 이내 → urgent 컬럼 추가"""
    today = date.today()
    limit = today + timedelta(days=days)
    df = df.copy()

    def _check(v):
        d = parse_date(v)
        if d is None:
            return False
        return today <= d <= limit

    df["_urgent"] = df[date_col].apply(_check) if date_col in df.columns else False
    return df


# ══════════════════════════════════════════════════════════════════════════════
def build_excel(df_so, df_pp, df_sh, path: Path):
    wb = openpyxl.Workbook()

    RED_FILL   = PatternFill("solid", fgColor="FFCCCC")
    RED_FONT   = Font(color="CC0000", bold=True)
    HDR_FILL   = PatternFill("solid", fgColor="1976D2")
    HDR_FONT   = Font(color="FFFFFF", bold=True)
    WARN_FILL  = PatternFill("solid", fgColor="FFF3CD")
    WARN_FONT  = Font(color="856404", bold=True)
    ALIGN_C    = Alignment(horizontal="center", vertical="center")
    THIN       = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    today = date.today()

    def write_sheet(ws, df, title, date_col):
        if df.empty:
            ws.append(["데이터 없음"])
            return 0

        # 상단 타이틀
        ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
        tc = ws["A1"]
        tc.value       = f"{title}  [{today.strftime('%Y-%m-%d')} 기준]"
        tc.font        = Font(size=13, bold=True)
        tc.alignment   = ALIGN_C
        ws.row_dimensions[1].height = 24

        # 헤더
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(2, ci, col)
            c.fill      = HDR_FILL
            c.font      = HDR_FONT
            c.alignment = ALIGN_C
            c.border    = THIN

        urgent_count = 0
        for ri, (_, row) in enumerate(df.iterrows(), 3):
            is_urg = row.get("_urgent", False)
            if is_urg:
                urgent_count += 1
            for ci, col in enumerate(df.columns, 1):
                if col == "_urgent":
                    continue
                val = row[col]
                if val and str(val) not in ("None", "NaN", "nan"):
                    val = str(val).split("T")[0]
                else:
                    val = ""
                c = ws.cell(ri, ci, val)
                c.border = THIN
                if is_urg:
                    c.fill = RED_FILL
                    c.font = RED_FONT
                elif col == date_col and val:
                    d = parse_date(val)
                    if d and d < today:
                        c.fill = WARN_FILL
                        c.font = WARN_FONT

        # 열 너비
        for ci, col in enumerate(df.columns, 1):
            if col == "_urgent":
                continue
            max_len = max(len(str(col)), 10)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 30)

        return urgent_count

    # ── 시트 1: 수주 (이미 flag_urgent 적용된 df_so 수신)
    ws1    = wb.active
    ws1.title = "수주"
    u1 = write_sheet(ws1, df_so, "재연마수주 등록", "dlv_dt")

    # ── 시트 2: 생산실적 ───────────────────────────────────────────────────
    df_pp2 = flag_urgent(df_pp, "dlv_dt", DLV_DAYS)
    ws2    = wb.create_sheet("생산실적")
    u2 = write_sheet(ws2, df_pp2, "생산실적 등록", "dlv_dt")

    # ── 시트 3: 출하 ──────────────────────────────────────────────────────
    df_sh2 = flag_urgent(df_sh, "out_dt", DLV_DAYS)
    ws3    = wb.create_sheet("출하")
    u3 = write_sheet(ws3, df_sh2, "재연마 출하등록", "out_dt")

    wb.save(path)
    return u1 + u2 + u3  # 긴급 항목 총 건수


# ══════════════════════════════════════════════════════════════════════════════
def _build_urgent_table(df_so: pd.DataFrame) -> str:
    """D-3 이내 수주 항목을 텍스트 표로 반환"""
    if df_so.empty or "_urgent" not in df_so.columns:
        return ""
    urgent_rows = df_so[df_so["_urgent"] == True]
    if urgent_rows.empty:
        return ""

    cols = ["cust_nm", "so_dt", "dlv_dt", "so_qty"]
    labels = {"cust_nm": "거래처", "so_dt": "수주일", "dlv_dt": "납기일", "so_qty": "수량"}
    available = [c for c in cols if c in urgent_rows.columns]

    lines = ["  " + "  ".join(f"{labels.get(c, c):<12}" for c in available)]
    lines.append("  " + "-" * (16 * len(available)))
    for _, row in urgent_rows.iterrows():
        vals = []
        for c in available:
            raw = row[c]
            if c == "so_qty":
                try:
                    v = str(int(float(raw))) if raw else "-"
                except Exception:
                    v = str(raw)
            else:
                v = str(raw).split("T")[0] if raw else "-"
            vals.append(f"{v:<12}")
        lines.append("  " + "  ".join(vals))
    return "\n".join(lines)


def send_email(urgent_count: int, df_so: pd.DataFrame = None):
    """Gmail API (OAuth)로 메일 발송 — SMTP 불필요"""
    service = get_gmail_service()

    today_str = date.today().strftime("%Y년 %m월 %d일")
    subject = f"[납기 알림] {date.today().strftime('%Y-%m-%d')} — D-{DLV_DAYS} 이내 {urgent_count}건"

    urgent_table = _build_urgent_table(df_so) if df_so is not None else ""
    urgent_section = f"\n■ D-{DLV_DAYS}일 이내 납기 항목 ({urgent_count}건)\n{urgent_table}\n" if urgent_table else f"\n■ D-{DLV_DAYS}일 이내 납기 항목: {urgent_count}건\n"

    body = f"""[{today_str} 납기 알림]{urgent_section}
본 메일은 자동 발송됩니다.
"""

    msg = MIMEMultipart()
    msg["From"]    = GMAIL_USER
    msg["To"]      = TO_EMAIL
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    # Gmail API send
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    service.users().messages().send(userId="me", body={"raw": raw}).execute()

    print(f"  ✅ 메일 발송 완료 → {TO_EMAIL}")


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    today    = date.today()
    filename = f"납기알림_{today.strftime('%Y%m%d')}.xlsx"
    out_path = SAVE_DIR / filename

    print(f"[{today}] 납기 알림 리포트 생성 중...")

    print("① ERP 데이터 조회")
    df_so, df_pp, df_sh = fetch_data()

    # 수주 D-3 플래그 — Excel과 메일 본문에 공통 사용
    df_so_flagged = flag_urgent(df_so, "dlv_dt", DLV_DAYS)

    print("② Excel 생성")
    urgent = build_excel(df_so_flagged, df_pp, df_sh, out_path)
    print(f"   저장: {out_path}  (긴급 {urgent}건)")

    print("③ 이메일 발송")
    send_email(urgent, df_so_flagged)

    print("완료.")
