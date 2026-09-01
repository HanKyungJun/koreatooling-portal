"""
Trico ERP 뷰어 GUI  v2.0
────────────────────────
실행: python erp_gui.py

변경이력:
  v2.0 (2026-06-16): 설정 탭 추가 (Gmail 인증, 수신자 설정, 납기 알림 수동 실행, 스케줄러 관리)
  v1.0 (초기): 생산실적/수주/출하 조회 탭

패치:
  2026-07-02: 영업팀 "출하등록(거래명세서)" 탭에 거래처(코드/명) 검색 필터 추가.
              메모(memo) 컬럼 표시 폭 확대 (내용이 잘리지 않도록).
  2026-07-16: 생산팀 "재연마 출하 거래명세서" 탭에도 단가0(확인 필요) 행 맨 위 정렬 적용
              (기존엔 영업팀 탭에만 적용돼 있었음 — 강조 표시는 있었으나 정렬은 안 됐던 불일치 수정).
              모든 탭 앱 시작 시 백그라운드 자동 조회(prefetch) 추가 — 서버 응답 자체는
              단축할 수 없으나(ERP 서버 처리시간), 탭 전환 전에 미리 로딩해둬 체감 대기시간 감소.
              기본 조회 날짜범위(이번달 1일~오늘)는 변경하지 않음(사용자 확인).
"""

import json
import re
import subprocess
import sys
import threading
import io
from datetime import date, timedelta
from pathlib import Path

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from trico_client import TricoClient

# ── 경로 ──────────────────────────────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    # PyInstaller exe: 모든 경로를 exe 위치 기준으로
    BASE_DIR = Path(sys.executable).resolve().parent
    ERP_DIR  = BASE_DIR
else:
    ERP_DIR  = Path(__file__).resolve().parent
    BASE_DIR = ERP_DIR.parent
TOKEN_FILE  = ERP_DIR / "token_gmail.json"
SECRET_FILE = BASE_DIR / "client_secret_gmail.json"
CONFIG_FILE = ERP_DIR / "config.json"
SCHED_PS1   = BASE_DIR / "scripts" / "setup_scheduler.ps1"
SAVE_DIR    = BASE_DIR / "outputs" / "dlv_alerts"
SCOPES      = ["https://www.googleapis.com/auth/gmail.send"]


# ── 설정 저장/로드 ─────────────────────────────────────────────────────────────
def load_config() -> dict:
    defaults = {
        "to_email":   "hzn2001@toolkorea.co.kr",
        "dlv_days":   3,
        "alert_time": "09:00",
    }
    if CONFIG_FILE.exists():
        try:
            data = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            defaults.update(data)
        except Exception:
            pass
    return defaults


def save_config(cfg: dict):
    CONFIG_FILE.write_text(
        json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8"
    )


# ── 컬럼명 한글 매핑 ───────────────────────────────────────────────────────────
COL_LABELS = {
    # 공통
    "cust_cd":     "거래처코드",  "cust_nm":     "거래처명",
    "cust_cd2":    "납품처코드",  "cust_nm2":    "납품처명",
    "dlv_dt":      "납기일",      "dlv_yard":    "납품지",
    "so_no":       "수주번호",    "so_sq":       "수주순번",
    "itm_cd":      "품목코드",    "itm_nm":      "품목명",
    "spec":        "규격",        "spec2":       "규격2",
    "model_cd":    "모델코드",    "rmks_sub":    "비고",
    "단가0":       "단가 0",
    "itm_id":      "품목ID",      "itm_nm_tmp":  "품목명(요약)",
    "stat_bc":     "상태",        "cdt":         "등록일시",
    "cnm":         "등록자",      "mdt":         "수정일시",
    "mnm":         "수정자",      "mid":         "수정자ID",
    "cid":         "등록자ID",    "memo":        "메모",
    "de_bc":       "납품구분",    "cury_bc":     "통화",
    "tran_bc":     "운송구분",    "tran_nm":     "운송사",
    "tran_tel":    "운송사연락처","tran_cust":   "운송거래처",
    "ent_bc":      "입력구분",
    # 생산실적
    "wo_no":       "작업지시번호","wo_dt":       "작업지시일",
    "fr_dt":       "시작일",      "to_dt":       "종료일",
    "plan_fac":    "계획공장",    "wo_qty":      "작업수량",
    "rout_cd":     "공정코드",    "wo_bc":       "작업구분",
    "wc_cd":       "작업장코드",  "sch_dt":      "예정일",
    "jae_coating": "코팅",        "jae_angle":   "각도",
    "wo_fr_dt":    "작업시작일",  "last_proc":   "최종공정",
    # 수주
    "so_dt":       "수주일",      "so_dept":     "수주부서코드",
    "so_dept_nm":  "수주부서",    "so_nm":       "담당자",
    "so_cnt":      "수주품목수",  "so_qty":      "수주수량",
    "out_dt":      "출하일",      "out_qty":     "출하수량",
    "order_nm":    "발주자",      "so_bc":       "수주구분",
    "sal_bc":      "판매구분",
    # 출하
    "mov_no":      "출하번호",    "out_bs":      "출하구분",
    "out_wh":      "출하창고",    "out_rid":     "담당자ID",
    "out_nm":      "담당자",      "out_dept":    "출하부서코드",
    "out_dept_nm": "출하부서",    "mov_bc":      "이동구분",
    "mov_own_qty": "자사수량",    "mov_qty":     "출하수량",
    "sal_dt":      "판매일",      "bad_qty":     "불량수량",
    "cnt":         "품목수",      "ex_rt":       "환율",
}

# ── 색상/폰트 ──────────────────────────────────────────────────────────────────
BG       = "#f5f5f5"
BG2      = "#ececec"
ACCENT   = "#1976d2"
GREEN    = "#388e3c"
RED      = "#c62828"
WHITE    = "#ffffff"
FONT     = ("맑은 고딕", 10)
FONT_SM  = ("맑은 고딕", 9)
FONT_HD  = ("맑은 고딕", 10, "bold")
FONT_SEC = ("맑은 고딕", 11, "bold")


# ══════════════════════════════════════════════════════════════════════════════
def _fmt(v) -> str:
    if v is None:
        return ""
    s = str(v)
    s = re.sub(r'T\d{2}:\d{2}:\d{2}.*', '', s)
    return s


# ══════════════════════════════════════════════════════════════════════════════
class DataTab(ttk.Frame):
    """단일 탭 — 날짜 입력 + 조회 + Treeview + Excel 저장"""

    def __init__(self, parent, name: str, fetch_fn, show_cust_filter: bool = False,
                 sort_zero_top: bool = False, **kwargs):
        super().__init__(parent, **kwargs)
        self.name             = name
        self.fetch_fn         = fetch_fn
        self.show_cust_filter = show_cust_filter
        self.sort_zero_top    = sort_zero_top
        self.df               = None
        self._build()

    def _build(self):
        bar = tk.Frame(self, bg=BG, pady=6, padx=8)
        bar.pack(fill="x")

        tk.Label(bar, text="시작일", bg=BG, font=FONT).pack(side="left")
        self.e_fr = tk.Entry(bar, width=12, font=FONT)
        self.e_fr.insert(0, date.today().replace(day=1).strftime("%Y-%m-%d"))
        self.e_fr.pack(side="left", padx=(4, 10))

        tk.Label(bar, text="종료일", bg=BG, font=FONT).pack(side="left")
        self.e_to = tk.Entry(bar, width=12, font=FONT)
        self.e_to.insert(0, date.today().strftime("%Y-%m-%d"))
        self.e_to.pack(side="left", padx=(4, 16))

        if self.show_cust_filter:
            tk.Label(bar, text="거래처(코드/명)", bg=BG, font=FONT).pack(side="left")
            self.e_cust = tk.Entry(bar, width=16, font=FONT)
            self.e_cust.pack(side="left", padx=(4, 16))
            self.e_cust.bind("<Return>", lambda e: self._query_thread())

        self.btn_query = tk.Button(
            bar, text="조회", font=FONT_HD, bg=ACCENT, fg=WHITE,
            padx=14, relief="flat", cursor="hand2", command=self._query_thread,
        )
        self.btn_query.pack(side="left")

        self.btn_excel = tk.Button(
            bar, text="Excel 저장", font=FONT, bg=GREEN, fg=WHITE,
            padx=10, relief="flat", cursor="hand2", command=self._save_excel,
        )
        self.btn_excel.pack(side="left", padx=(8, 0))

        self.lbl_status = tk.Label(bar, text="", bg=BG, font=FONT_SM, fg="#666")
        self.lbl_status.pack(side="left", padx=12)

        frame = tk.Frame(self)
        frame.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(frame, show="headings", selectmode="browse")
        vsb = ttk.Scrollbar(frame, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)

        self.tree.tag_configure("odd",  background="#fafafa")
        self.tree.tag_configure("even", background=WHITE)

    def _query_thread(self):
        self.btn_query.config(state="disabled", text="조회 중...")
        self.lbl_status.config(text="")
        threading.Thread(target=self._do_query, daemon=True).start()

    def _do_query(self):
        fr = self.e_fr.get().strip()
        to = self.e_to.get().strip()
        try:
            df = self.fetch_fn(fr_dt=fr, to_dt=to)
            if self.show_cust_filter:
                kw = self.e_cust.get().strip()
                if kw:
                    df = self._filter_cust(df, kw)
            df = self._postprocess_df(df)
            self.df = df
            self.after(0, lambda: self._load_table(df))
        except Exception as e:
            self.after(0, lambda: self.lbl_status.config(text=f"❌ {e}", fg="red"))
        finally:
            self.after(0, lambda: self.btn_query.config(state="normal", text="조회"))

    def _postprocess_df(self, df):
        """서브클래스가 조회 결과를 추가로 가공(정렬 등)할 때 오버라이드하는 훅. 기본은 그대로 반환."""
        return df

    @staticmethod
    def _filter_cust(df, keyword: str):
        """거래처코드(cust_cd) 또는 거래처명(cust_nm)에 keyword가 포함된 행만 남김 (클라이언트 측 필터, 대소문자 무시)."""
        if df is None or df.empty:
            return df
        cols = [c for c in ("cust_cd", "cust_nm") if c in df.columns]
        if not cols:
            return df
        mask = None
        for c in cols:
            m = df[c].astype(str).str.contains(keyword, case=False, na=False)
            mask = m if mask is None else (mask | m)
        return df[mask] if mask is not None else df

    def _load_table(self, df):
        self.tree.delete(*self.tree.get_children())
        if df.empty:
            self.lbl_status.config(text="데이터 없음", fg="#888")
            return

        cols   = list(df.columns)
        labels = [COL_LABELS.get(c, c) for c in cols]
        self.tree["columns"] = cols
        for c, lbl in zip(cols, labels):
            w = 240 if c == "memo" else max(80, len(lbl) * 13)
            self.tree.heading(c, text=lbl)
            self.tree.column(c, width=w, minwidth=60,
                             anchor="w", stretch=False)

        for i, row in df.iterrows():
            tag = "odd" if i % 2 else "even"
            self.tree.insert("", "end", values=[_fmt(v) for v in row], tags=(tag,))

        self.lbl_status.config(text=f"✅ {len(df)}행 × {len(cols)}열", fg=ACCENT)

    def _save_excel(self):
        if self.df is None or self.df.empty:
            messagebox.showwarning("알림", "먼저 조회를 실행하세요.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=f"{self.name}_{date.today().strftime('%Y%m%d')}.xlsx",
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.name[:31]

            hdr_fill = PatternFill("solid", fgColor="1976D2")
            hdr_font = Font(color="FFFFFF", bold=True)

            for ci, col in enumerate(self.df.columns, 1):
                cell = ws.cell(1, ci, col)
                cell.fill      = hdr_fill
                cell.font      = hdr_font
                cell.alignment = Alignment(horizontal="center")

            for ri, row in self.df.iterrows():
                for ci, val in enumerate(row, 1):
                    ws.cell(ri + 2, ci, _fmt(val))

            wb.save(path)
            messagebox.showinfo("저장 완료", f"저장됨:\n{path}")
        except Exception as e:
            messagebox.showerror("오류", str(e))


# ══════════════════════════════════════════════════════════════════════════════
class ShipmentTab(DataTab):
    """출하등록(거래명세서) 탭 — 단가 0 행 강조 + 건수 배지"""

    def _build(self):
        super()._build()
        self.tree.tag_configure("zero_price", background="#FFEBEE", foreground="#C62828")
        self.lbl_zero = tk.Label(self, text="", bg=BG, font=FONT_HD, fg=RED)
        self.lbl_zero.pack(side="bottom", anchor="w", padx=10, pady=2)

    def _postprocess_df(self, df):
        """sort_zero_top=True인 탭은 단가0 행을 맨 위로 정렬 (그 외 순서는 원래 순서 유지)."""
        if self.sort_zero_top and df is not None and not df.empty and "단가0" in df.columns:
            df = df.sort_values(by="단가0", ascending=False, kind="mergesort").reset_index(drop=True)
        return df

    def _load_table(self, df):
        self.tree.delete(*self.tree.get_children())
        if df.empty:
            self.lbl_status.config(text="데이터 없음", fg="#888")
            self.lbl_zero.config(text="")
            return

        cols   = list(df.columns)
        labels = [COL_LABELS.get(c, c) for c in cols]
        self.tree["columns"] = cols
        for c, lbl in zip(cols, labels):
            w = 50 if c == "단가0" else (240 if c == "memo" else max(80, len(lbl) * 13))
            self.tree.heading(c, text=lbl)
            self.tree.column(c, width=w, minwidth=40, anchor="center" if c == "단가0" else "w", stretch=False)

        zero_cnt = 0
        for i, row in df.iterrows():
            is_zero = bool(row.get("단가0", False))
            if is_zero:
                zero_cnt += 1
            tag = "zero_price" if is_zero else ("odd" if i % 2 else "even")
            vals = ["⚠️" if (c == "단가0" and is_zero) else ("" if c == "단가0" else _fmt(v))
                    for c, v in zip(cols, row)]
            self.tree.insert("", "end", values=vals, tags=(tag,))

        self.lbl_status.config(text=f"✅ {len(df)}행 × {len(cols)}열", fg=ACCENT)
        if zero_cnt:
            self.lbl_zero.config(text=f"⚠️  단가 0 전표: {zero_cnt}건 — 확인 필요")
        else:
            self.lbl_zero.config(text="✅  단가 0 전표 없음", fg=GREEN)


# ══════════════════════════════════════════════════════════════════════════════
class Section(tk.LabelFrame):
    """설정 탭용 구역 프레임"""
    def __init__(self, parent, title, **kwargs):
        super().__init__(
            parent, text=f"  {title}  ",
            bg=BG, fg="#1976d2",
            font=FONT_SEC,
            bd=1, relief="groove",
            padx=12, pady=8,
            **kwargs,
        )


class SettingsTab(ttk.Frame):
    """⚙️ 설정 탭 — Gmail 인증 / 수신자 / 스케줄러 / 납기 알림 수동 실행"""

    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self.cfg = load_config()
        self._build()
        self.after(500, self._refresh_all)   # 시작 시 상태 자동 조회

    # ── 전체 레이아웃 ──────────────────────────────────────────────────────────
    def _build(self):
        canvas = tk.Canvas(self, bg=BG, highlightthickness=0)
        vsb    = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = tk.Frame(canvas, bg=BG)
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_canvas_resize(e):
            canvas.itemconfig(win_id, width=e.width)

        inner.bind("<Configure>", _on_configure)
        canvas.bind("<Configure>", _on_canvas_resize)
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(
            -1 if e.delta > 0 else 1, "units"))

        PAD = {"pady": 6, "padx": 6, "fill": "x"}

        # ── § 1. ERP 연결 ──────────────────────────────────────────────────
        sec1 = Section(inner, "1.  ERP 연결")
        sec1.pack(**PAD)

        row1 = tk.Frame(sec1, bg=BG)
        row1.pack(fill="x")

        tk.Button(
            row1, text="연결 테스트", font=FONT_HD, bg=ACCENT, fg=WHITE,
            padx=14, relief="flat", cursor="hand2",
            command=self._test_erp,
        ).pack(side="left")

        self.lbl_erp = tk.Label(row1, text="", bg=BG, font=FONT)
        self.lbl_erp.pack(side="left", padx=12)

        # ── § 2. Gmail 설정 ────────────────────────────────────────────────
        sec2 = Section(inner, "2.  Gmail 납기 알림 설정")
        sec2.pack(**PAD)

        # 2-a 토큰 상태
        tok_row = tk.Frame(sec2, bg=BG)
        tok_row.pack(fill="x", pady=(0, 6))
        tk.Label(tok_row, text="Gmail 토큰:", bg=BG, font=FONT_HD, width=14, anchor="w").pack(side="left")
        self.lbl_token = tk.Label(tok_row, text="확인 중…", bg=BG, font=FONT)
        self.lbl_token.pack(side="left")

        btn_row = tk.Frame(sec2, bg=BG)
        btn_row.pack(fill="x", pady=(0, 8))
        tk.Button(
            btn_row, text="Gmail 인증 시작", font=FONT_HD, bg=ACCENT, fg=WHITE,
            padx=12, relief="flat", cursor="hand2",
            command=self._setup_gmail,
        ).pack(side="left")
        tk.Button(
            btn_row, text="토큰 삭제 후 재인증", font=FONT, bg="#757575", fg=WHITE,
            padx=10, relief="flat", cursor="hand2",
            command=self._reset_token,
        ).pack(side="left", padx=(8, 0))

        # 2-b 수신자 이메일
        email_row = tk.Frame(sec2, bg=BG)
        email_row.pack(fill="x", pady=(0, 6))
        tk.Label(email_row, text="수신자 이메일:", bg=BG, font=FONT_HD, width=14, anchor="w").pack(side="left")
        self.e_email = tk.Entry(email_row, width=36, font=FONT)
        self.e_email.insert(0, self.cfg.get("to_email", ""))
        self.e_email.pack(side="left", padx=(0, 10))

        # 2-c D-N 임계값
        days_row = tk.Frame(sec2, bg=BG)
        days_row.pack(fill="x", pady=(0, 6))
        tk.Label(days_row, text="납기 D-", bg=BG, font=FONT_HD, width=14, anchor="w").pack(side="left")
        self.e_days = tk.Entry(days_row, width=4, font=FONT, justify="center")
        self.e_days.insert(0, str(self.cfg.get("dlv_days", 3)))
        self.e_days.pack(side="left")
        tk.Label(days_row, text=" 일 이내 항목 빨간 강조", bg=BG, font=FONT).pack(side="left")

        # 저장 버튼
        save_row = tk.Frame(sec2, bg=BG)
        save_row.pack(fill="x", pady=(6, 0))
        tk.Button(
            save_row, text="설정 저장", font=FONT_HD, bg=GREEN, fg=WHITE,
            padx=14, relief="flat", cursor="hand2",
            command=self._save_settings,
        ).pack(side="right")
        self.lbl_save = tk.Label(save_row, text="", bg=BG, font=FONT_SM, fg=GREEN)
        self.lbl_save.pack(side="right", padx=8)

        # ── § 3. 납기 알림 수동 실행 ───────────────────────────────────────
        sec3 = Section(inner, "3.  납기 알림 수동 실행")
        sec3.pack(**PAD)

        run_row = tk.Frame(sec3, bg=BG)
        run_row.pack(fill="x", pady=(0, 8))

        self.btn_alert = tk.Button(
            run_row, text="▶  지금 납기 알림 실행", font=FONT_HD, bg=RED, fg=WHITE,
            padx=16, relief="flat", cursor="hand2",
            command=self._run_alert_thread,
        )
        self.btn_alert.pack(side="left")

        self.lbl_last = tk.Label(run_row, text="", bg=BG, font=FONT_SM, fg="#666")
        self.lbl_last.pack(side="left", padx=14)

        # 로그 박스
        log_frame = tk.Frame(sec3, bg=BG)
        log_frame.pack(fill="x")
        tk.Label(log_frame, text="실행 로그:", bg=BG, font=FONT_SM, fg="#666").pack(anchor="w")
        self.txt_log = tk.Text(
            log_frame, height=8, font=("Consolas", 9),
            bg="#1e1e1e", fg="#d4d4d4", insertbackground=WHITE,
            relief="flat", state="disabled",
        )
        self.txt_log.pack(fill="x", pady=(2, 0))

        # ── § 4. 단가 0 알림 수동 실행 ────────────────────────────────────────
        sec4 = Section(inner, "4.  단가 0 출하 전표 알림 수동 실행")
        sec4.pack(**PAD)

        zp_run_row = tk.Frame(sec4, bg=BG)
        zp_run_row.pack(fill="x", pady=(0, 8))

        self.btn_zero = tk.Button(
            zp_run_row, text="▶  지금 단가 0 알림 실행", font=FONT_HD, bg="#6a1b9a", fg=WHITE,
            padx=16, relief="flat", cursor="hand2",
            command=self._run_zero_alert_thread,
        )
        self.btn_zero.pack(side="left")

        self.lbl_zero_last = tk.Label(zp_run_row, text="", bg=BG, font=FONT_SM, fg="#666")
        self.lbl_zero_last.pack(side="left", padx=14)

        zp_log_frame = tk.Frame(sec4, bg=BG)
        zp_log_frame.pack(fill="x")
        tk.Label(zp_log_frame, text="실행 로그:", bg=BG, font=FONT_SM, fg="#666").pack(anchor="w")
        self.txt_zero_log = tk.Text(
            zp_log_frame, height=5, font=("Consolas", 9),
            bg="#1e1e1e", fg="#d4d4d4", insertbackground=WHITE,
            relief="flat", state="disabled",
        )
        self.txt_zero_log.pack(fill="x", pady=(2, 0))

        # ── § 5. Windows 스케줄러 ──────────────────────────────────────────
        sec4 = Section(inner, "5.  Windows 작업 스케줄러")
        sec4.pack(**PAD)

        sched_btn_row = tk.Frame(sec4, bg=BG)
        sched_btn_row.pack(fill="x", pady=(0, 8))

        tk.Button(
            sched_btn_row, text="상태 새로고침", font=FONT, bg="#546e7a", fg=WHITE,
            padx=10, relief="flat", cursor="hand2",
            command=self._check_scheduler,
        ).pack(side="left")

        tk.Button(
            sched_btn_row, text="스케줄러 등록 (관리자 권한)", font=FONT_HD, bg=ACCENT, fg=WHITE,
            padx=14, relief="flat", cursor="hand2",
            command=self._register_scheduler,
        ).pack(side="left", padx=(10, 0))

        self.lbl_sched1 = tk.Label(sec4, text="CNC_Daily_Report:        확인 중…", bg=BG, font=FONT, anchor="w")
        self.lbl_sched1.pack(fill="x")
        self.lbl_sched2 = tk.Label(sec4, text="CNC_Daily_Report_OnBoot: 확인 중…", bg=BG, font=FONT, anchor="w")
        self.lbl_sched2.pack(fill="x", pady=(2, 0))

        tk.Label(sec4, text="⚠️  스케줄러 등록은 관리자 권한이 필요합니다. UAC 창이 뜨면 '예'를 클릭하세요.",
                 bg=BG, font=FONT_SM, fg="#b71c1c", anchor="w").pack(fill="x", pady=(8, 0))

    # ── 상태 일괄 새로고침 ─────────────────────────────────────────────────────
    def _refresh_all(self):
        self._check_token_status()
        self._check_scheduler()

    # ── § 1 ERP 연결 테스트 ────────────────────────────────────────────────────
    def _test_erp(self):
        self.lbl_erp.config(text="연결 중…", fg="#666")
        threading.Thread(target=self._do_test_erp, daemon=True).start()

    def _do_test_erp(self):
        try:
            client = TricoClient()
            today  = date.today()
            fr     = today.replace(day=1).strftime("%Y-%m-%d")
            df     = client.수주(fr_dt=fr)
            msg    = f"✅ 연결됨  ({len(df)}건 조회)"
            self.after(0, lambda: self.lbl_erp.config(text=msg, fg=GREEN))
        except Exception as e:
            err = str(e)[:80]
            self.after(0, lambda: self.lbl_erp.config(text=f"❌ {err}", fg=RED))

    # ── § 2 Gmail 토큰 상태 확인 ──────────────────────────────────────────────
    def _check_token_status(self):
        if not TOKEN_FILE.exists():
            self.lbl_token.config(text="❌ 토큰 없음  →  'Gmail 인증 시작' 클릭", fg=RED)
            return
        try:
            from google.oauth2.credentials import Credentials
            creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
            if creds.expired and not creds.refresh_token:
                self.lbl_token.config(text="⚠️ 토큰 만료 (갱신 불가)  →  재인증 필요", fg="#e65100")
            elif creds.expired:
                self.lbl_token.config(text="⚠️ 토큰 만료  →  다음 실행 시 자동 갱신", fg="#f9a825")
            else:
                self.lbl_token.config(
                    text=f"✅ 정상  ({creds.client_id[:30]}…)", fg=GREEN)
        except Exception as e:
            self.lbl_token.config(text=f"⚠️ 토큰 읽기 오류: {e}", fg="#e65100")

    # ── § 2 Gmail 인증 시작 ───────────────────────────────────────────────────
    def _setup_gmail(self):
        if not SECRET_FILE.exists():
            messagebox.showerror(
                "파일 없음",
                f"client_secret_gmail.json 파일이 없습니다.\n\n"
                f"위치: {SECRET_FILE}\n\n"
                "Google Cloud Console → OAuth 2.0 클라이언트 자격증명 → JSON 다운로드 후\n"
                "위 경로에 저장하세요."
            )
            return

        if TOKEN_FILE.exists():
            if not messagebox.askyesno("확인", "이미 토큰이 있습니다. 재인증하시겠습니까?"):
                return
            TOKEN_FILE.unlink(missing_ok=True)

        self.lbl_token.config(text="브라우저에서 Google 로그인 중…", fg="#1565c0")
        threading.Thread(target=self._do_setup_gmail, daemon=True).start()

    def _do_setup_gmail(self):
        try:
            from google_auth_oauthlib.flow import InstalledAppFlow
            flow  = InstalledAppFlow.from_client_secrets_file(str(SECRET_FILE), SCOPES)
            creds = flow.run_local_server(port=0)
            TOKEN_FILE.write_text(creds.to_json(), encoding="utf-8")
            self.after(0, lambda: (
                self.lbl_token.config(text="✅ 인증 완료  토큰이 저장되었습니다.", fg=GREEN),
                messagebox.showinfo("완료", "Gmail 인증이 완료되었습니다.\n납기 알림을 발송할 수 있습니다.")
            ))
        except Exception as e:
            err = str(e)
            self.after(0, lambda: (
                self.lbl_token.config(text=f"❌ 인증 실패: {err}", fg=RED),
                messagebox.showerror("인증 실패", err)
            ))

    def _reset_token(self):
        if TOKEN_FILE.exists():
            TOKEN_FILE.unlink()
        self._setup_gmail()

    # ── § 2 설정 저장 ─────────────────────────────────────────────────────────
    def _save_settings(self):
        email = self.e_email.get().strip()
        try:
            days = int(self.e_days.get().strip())
        except ValueError:
            messagebox.showerror("오류", "D-N 일수는 숫자로 입력하세요.")
            return

        self.cfg["to_email"] = email
        self.cfg["dlv_days"] = days
        save_config(self.cfg)

        self.lbl_save.config(text="✅ 저장됨")
        self.after(2500, lambda: self.lbl_save.config(text=""))

    # ── § 3 납기 알림 수동 실행 ───────────────────────────────────────────────
    def _append_log(self, text: str):
        self.txt_log.config(state="normal")
        self.txt_log.insert("end", text + "\n")
        self.txt_log.see("end")
        self.txt_log.config(state="disabled")

    def _clear_log(self):
        self.txt_log.config(state="normal")
        self.txt_log.delete("1.0", "end")
        self.txt_log.config(state="disabled")

    def _run_alert_thread(self):
        if not TOKEN_FILE.exists():
            messagebox.showwarning("Gmail 토큰 없음", "먼저 'Gmail 인증 시작'을 실행하세요.")
            return
        self.btn_alert.config(state="disabled", text="실행 중…")
        self._clear_log()
        threading.Thread(target=self._do_run_alert, daemon=True).start()

    def _do_run_alert(self):
        import sys as _sys
        # stdout 리디렉션 → 로그 박스
        buf = io.StringIO()
        old_stdout = _sys.stdout
        old_stderr = _sys.stderr
        _sys.stdout = buf
        _sys.stderr = buf

        try:
            # daily_dlv_alert 함수들을 동적 임포트 (sys.exit 방지)
            import importlib.util, types

            spec = importlib.util.spec_from_file_location(
                "daily_dlv_alert", ERP_DIR / "daily_dlv_alert.py"
            )
            mod = types.ModuleType("daily_dlv_alert")
            mod.__spec__ = spec
            spec.loader.exec_module(mod)

            # 설정 동기화
            mod.TO_EMAIL  = self.cfg.get("to_email", mod.TO_EMAIL)
            mod.DLV_DAYS  = int(self.cfg.get("dlv_days", mod.DLV_DAYS))
            mod.TOKEN_FILE = TOKEN_FILE
            mod.SAVE_DIR.mkdir(parents=True, exist_ok=True)

            today    = date.today()
            filename = f"납기알림_{today.strftime('%Y%m%d')}.xlsx"
            out_path = mod.SAVE_DIR / filename

            print(f"[{today}] 납기 알림 리포트 생성 중…")
            print("① ERP 데이터 조회")
            df_so, df_pp, df_sh = mod.fetch_data()

            print("② Excel 생성")
            urgent = mod.build_excel(df_so, df_pp, df_sh, out_path)
            print(f"   저장: {out_path}  (긴급 {urgent}건)")

            print("③ 이메일 발송")
            mod.send_email(out_path, urgent)

            result = f"✅ 완료  |  긴급 {urgent}건  →  {mod.TO_EMAIL}"
            last   = f"마지막 실행: {today.strftime('%Y-%m-%d')} {date.today().strftime('%H:%M')}"
            self.after(0, lambda: (
                self.btn_alert.config(state="normal", text="▶  지금 납기 알림 실행"),
                self.lbl_last.config(text=last, fg=GREEN),
            ))

        except Exception as e:
            import traceback
            print(f"\n❌ 오류 발생:\n{traceback.format_exc()}")
            self.after(0, lambda: self.btn_alert.config(
                state="normal", text="▶  지금 납기 알림 실행"))
        finally:
            _sys.stdout = old_stdout
            _sys.stderr = old_stderr
            log_text = buf.getvalue()
            self.after(0, lambda: self._append_log(log_text))

    def _run_zero_alert_thread(self):
        if not TOKEN_FILE.exists():
            messagebox.showwarning("Gmail 토큰 없음", "먼저 'Gmail 인증 시작'을 실행하세요.")
            return
        self.btn_zero.config(state="disabled", text="실행 중…")
        self._clear_zero_log()
        threading.Thread(target=self._do_run_zero_alert, daemon=True).start()

    def _do_run_zero_alert(self):
        import sys as _sys
        buf = io.StringIO()
        old_stdout, old_stderr = _sys.stdout, _sys.stderr
        _sys.stdout = buf
        _sys.stderr = buf

        try:
            import importlib.util, types

            spec = importlib.util.spec_from_file_location(
                "zero_price_alert", ERP_DIR / "zero_price_alert.py"
            )
            mod = types.ModuleType("zero_price_alert")
            mod.__spec__ = spec
            spec.loader.exec_module(mod)

            mod.TOKEN_FILE = TOKEN_FILE
            mod.SAVE_DIR.mkdir(parents=True, exist_ok=True)

            today    = date.today()
            filename = f"단가0알림_{today.strftime('%Y%m%d')}.xlsx"
            out_path = mod.SAVE_DIR / filename

            print(f"[{today}] 단가 0 출하 전표 알림 실행 중…")
            print("① ERP 조회")
            df = mod.fetch_zero_price()

            if df.empty:
                print("   단가 0 / 1일 경과 전표 없음 — 발송 생략")
                result = "✅ 완료  |  단가 0 전표 없음"
            else:
                print(f"   단가 0 전표: {len(df)}건")
                print("② Excel 생성")
                mod.build_excel(df, out_path)
                print(f"   저장: {out_path}")
                print("③ 이메일 발송")
                mod.send_email(df, out_path)
                result = f"✅ 완료  |  단가 0 {len(df)}건 발송"

            self.after(0, lambda: (
                self.btn_zero.config(state="normal", text="▶  지금 단가 0 알림 실행"),
                self.lbl_zero_last.config(
                    text=f"마지막 실행: {today.strftime('%Y-%m-%d %H:%M')}", fg=GREEN),
            ))

        except Exception as e:
            import traceback
            print(f"\n❌ 오류 발생:\n{traceback.format_exc()}")
            self.after(0, lambda: self.btn_zero.config(
                state="normal", text="▶  지금 단가 0 알림 실행"))
        finally:
            _sys.stdout = old_stdout
            _sys.stderr = old_stderr
            self.after(0, lambda: self._append_zero_log(buf.getvalue()))

    def _clear_zero_log(self):
        self.txt_zero_log.config(state="normal")
        self.txt_zero_log.delete("1.0", "end")
        self.txt_zero_log.config(state="disabled")

    def _append_zero_log(self, text: str):
        self.txt_zero_log.config(state="normal")
        self.txt_zero_log.insert("end", text)
        self.txt_zero_log.see("end")
        self.txt_zero_log.config(state="disabled")

    # ── § 5 스케줄러 상태 ─────────────────────────────────────────────────────
    def _check_scheduler(self):
        threading.Thread(target=self._do_check_scheduler, daemon=True).start()

    def _do_check_scheduler(self):
        tasks = {
            "CNC_Daily_Report":       self.lbl_sched1,
            "CNC_Daily_Report_OnBoot": self.lbl_sched2,
        }
        for task_name, lbl in tasks.items():
            try:
                r = subprocess.run(
                    ["powershell", "-NoProfile", "-Command",
                     f"Get-ScheduledTask -TaskName '{task_name}' "
                     "| Select-Object TaskName,State | ConvertTo-Json"],
                    capture_output=True, text=True, timeout=8,
                    creationflags=subprocess.CREATE_NO_WINDOW,
                )
                if r.returncode == 0 and r.stdout.strip():
                    data  = json.loads(r.stdout.strip())
                    state = data.get("State", "?")
                    icon  = "✅" if state in ("Ready", "Running") else "⚠️"
                    text  = f"{icon}  {task_name}: {state}"
                    fg    = GREEN if state in ("Ready", "Running") else "#e65100"
                else:
                    text = f"❌  {task_name}: 등록되지 않음"
                    fg   = RED
            except Exception as e:
                text = f"⚠️  {task_name}: 조회 오류 ({e})"
                fg   = "#e65100"

            _lbl = lbl  # closure capture
            _text, _fg = text, fg
            self.after(0, lambda l=_lbl, t=_text, f=_fg: l.config(text=t, fg=f))

    def _register_scheduler(self):
        if not SCHED_PS1.exists():
            messagebox.showerror("파일 없음", f"스케줄러 설정 파일을 찾을 수 없습니다:\n{SCHED_PS1}")
            return
        if not messagebox.askyesno(
            "스케줄러 등록",
            "Windows 작업 스케줄러에 납기 알림을 등록합니다.\n\n"
            "관리자 권한(UAC) 요청 창이 뜨면 '예'를 클릭하세요.\n\n"
            "계속하시겠습니까?"
        ):
            return
        try:
            subprocess.Popen(
                ["powershell", "-Command",
                 f"Start-Process powershell "
                 f"-ArgumentList '-NoProfile -ExecutionPolicy Bypass -File \"{SCHED_PS1}\"' "
                 f"-Verb RunAs"],
                creationflags=subprocess.CREATE_NO_WINDOW,
            )
            messagebox.showinfo("실행됨",
                "스케줄러 등록 스크립트를 관리자 권한으로 실행했습니다.\n"
                "약 5초 후 '상태 새로고침' 버튼을 눌러 확인하세요.")
            self.after(6000, self._check_scheduler)
        except Exception as e:
            messagebox.showerror("오류", str(e))


# ══════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Trico ERP 뷰어  v2.0")
        self.geometry("1280x760")
        self.configure(bg=BG)

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TNotebook",        background=BG, borderwidth=0)
        style.configure("TNotebook.Tab",    font=FONT_HD, padding=(14, 6))
        style.configure("Treeview",         font=FONT_SM, rowheight=24)
        style.configure("Treeview.Heading", font=FONT_HD)

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=8, pady=8)

        style.configure("Inner.TNotebook",     background=BG2, borderwidth=0)
        style.configure("Inner.TNotebook.Tab", font=FONT, padding=(10, 4))

        # 탭 최초 표시 전 백그라운드 자동 조회(prefetch) 시작 지연(ms) — 화면 렌더링이
        # 끝난 뒤 순차적으로 조회를 걸어 서버에 요청이 한꺼번에 몰리지 않도록 함.
        prefetch_delay_ms = 300
        PREFETCH_STEP_MS  = 250

        # 생산팀 상위 탭
        prod_frame = ttk.Frame(nb, style="TFrame")
        nb.add(prod_frame, text="  🏭 생산팀  ")

        try:
            client = TricoClient()
            inner_nb = ttk.Notebook(prod_frame, style="Inner.TNotebook")
            inner_nb.pack(fill="both", expand=True, padx=4, pady=4)

            screens = [
                ("생산실적 등록",        client.생산실적,   DataTab,     False),
                ("재연마수주 등록",       client.수주,      DataTab,     False),
                ("재연마 출하등록",       client.출하,      DataTab,     False),
                ("재연마 출하 거래명세서", client.출하_명세,  ShipmentTab, True),
            ]
            for name, fn, TabClass, sort_zero in screens:
                tab = TabClass(inner_nb, name=name, fetch_fn=fn,
                                sort_zero_top=sort_zero, style="TFrame")
                inner_nb.add(tab, text=f"  {name}  ")
                tab.after(prefetch_delay_ms, tab._query_thread)
                prefetch_delay_ms += PREFETCH_STEP_MS
        except Exception as e:
            ttk.Label(prod_frame,
                      text=f"⚠️  ERP 연결 실패: {e}\n\n사무실 네트워크(LAN/VPN) 연결 후 재시작하세요.",
                      font=FONT, foreground=RED, justify="center"
                      ).pack(expand=True)

        # 영업팀 상위 탭
        sales_frame = ttk.Frame(nb, style="TFrame")
        nb.add(sales_frame, text="  💼 영업팀  ")

        sales_nb = ttk.Notebook(sales_frame, style="Inner.TNotebook")
        sales_nb.pack(fill="both", expand=True, padx=4, pady=4)

        try:
            sales_screens = [
                ("출하등록(거래명세서)", client.출하_영업_명세, ShipmentTab),
            ]
            for name, fn, TabClass in sales_screens:
                tab = TabClass(sales_nb, name=name, fetch_fn=fn,
                                show_cust_filter=True, sort_zero_top=True, style="TFrame")
                sales_nb.add(tab, text=f"  {name}  ")
                tab.after(prefetch_delay_ms, tab._query_thread)
                prefetch_delay_ms += PREFETCH_STEP_MS
        except Exception as e:
            ttk.Label(sales_frame,
                      text=f"⚠️  ERP 연결 실패: {e}\n\n사무실 네트워크(LAN/VPN) 연결 후 재시작하세요.",
                      font=FONT, foreground=RED, justify="center"
                      ).pack(expand=True)

        # 설정 탭
        settings_tab = SettingsTab(nb, style="TFrame")
        nb.add(settings_tab, text="  ⚙️ 설정  ")

        # 상태바
        tk.Label(self,
                 text="Trico ERP 뷰어 v2.0  |  단가·금액 컬럼 자동 차단  |  사무실 네트워크 전용",
                 bg="#e0e0e0", font=FONT_SM, anchor="w", padx=8, pady=3
                 ).pack(fill="x", side="bottom")


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    App().mainloop()
