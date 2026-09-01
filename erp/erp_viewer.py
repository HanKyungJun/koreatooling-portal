"""
Trico ERP Viewer  v1.0
──────────────────────
MS SQL Server 읽기 전용 뷰어
탭: 수주/작업지시 · 생산실적 · 재고/자재 · 공구 수불
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import json
import os
import threading
import datetime

# ── 가격 관련 차단 키워드 (컬럼명에 포함되면 내보내기 제외) ─────────────────
PRICE_BLOCKED_KEYWORDS = [
    # 한국어
    "단가", "정가", "금액", "가격", "원가", "공급가", "부가세", "세액",
    "할인", "매출액", "매입액", "공급액", "판매가", "구매가", "견적가",
    "낙찰가", "계약금", "잔금", "선급금", "청구금", "지급금", "수수료",
    # 영문 (Trico 영문 컬럼명 대비)
    "price", "cost", "amount", "rate", "value", "fee", "tax", "vat",
    "discount", "charge", "revenue", "payment", "invoice",
]

def is_price_column(col_name: str) -> bool:
    """컬럼명에 가격 관련 키워드가 포함되면 True 반환"""
    lower = col_name.lower().replace(" ", "").replace("_", "")
    return any(kw.replace(" ", "") in lower for kw in PRICE_BLOCKED_KEYWORDS)

# ── 설정 파일 경로 ─────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE  = os.path.join(BASE_DIR, "erp_config.json")
QUERIES_FILE = os.path.join(BASE_DIR, "erp_queries.json")

# ── 기본 쿼리 (Trico DB 스키마에 맞게 수정 필요) ─────────────────────────────
DEFAULT_QUERIES = {
    "수주/작업지시": {
        "label": "수주/작업지시",
        "sql": (
            "SELECT TOP 200 "
            "  wo_no       AS [작업지시번호], "
            "  order_no    AS [수주번호], "
            "  item_code   AS [품번], "
            "  item_name   AS [품명], "
            "  qty_order   AS [수주수량], "
            "  qty_comp    AS [완료수량], "
            "  due_date    AS [납기일], "
            "  status      AS [상태] "
            "FROM dbo.WorkOrder "
            "ORDER BY due_date DESC"
        )
    },
    "생산실적": {
        "label": "생산실적",
        "sql": (
            "SELECT TOP 200 "
            "  prod_date   AS [생산일자], "
            "  wo_no       AS [작업지시번호], "
            "  item_code   AS [품번], "
            "  item_name   AS [품명], "
            "  qty_good    AS [양품수량], "
            "  qty_defect  AS [불량수량], "
            "  worker      AS [작업자] "
            "FROM dbo.ProdResult "
            "ORDER BY prod_date DESC"
        )
    },
    "재고/자재": {
        "label": "재고/자재",
        "sql": (
            "SELECT TOP 200 "
            "  item_code   AS [품번], "
            "  item_name   AS [품명], "
            "  location    AS [위치], "
            "  qty_stock   AS [재고수량], "
            "  unit        AS [단위], "
            "  last_in     AS [최종입고일] "
            "FROM dbo.Inventory "
            "ORDER BY item_code"
        )
    },
    "공구 수불": {
        "label": "공구 수불",
        "sql": (
            "SELECT TOP 200 "
            "  tran_date   AS [일자], "
            "  tool_code   AS [공구코드], "
            "  tool_name   AS [공구명], "
            "  tran_type   AS [구분], "
            "  qty         AS [수량], "
            "  balance     AS [잔량], "
            "  worker      AS [담당자] "
            "FROM dbo.ToolTran "
            "ORDER BY tran_date DESC"
        )
    }
}

TAB_ORDER = ["수주/작업지시", "생산실적", "재고/자재", "공구 수불"]

# ── 설정 로드/저장 ─────────────────────────────────────────────────────────────
def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"server": "", "database": "", "username": "", "password": "", "driver": "ODBC Driver 17 for SQL Server"}

def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def load_queries():
    if os.path.exists(QUERIES_FILE):
        with open(QUERIES_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    save_queries(DEFAULT_QUERIES)
    return DEFAULT_QUERIES

def save_queries(q):
    with open(QUERIES_FILE, "w", encoding="utf-8") as f:
        json.dump(q, f, ensure_ascii=False, indent=2)

# ── DB 연결 ────────────────────────────────────────────────────────────────────
def get_connection(cfg):
    try:
        import pyodbc
    except ImportError:
        raise RuntimeError("pyodbc가 설치되지 않았습니다.\npip install pyodbc 실행 후 재시작하세요.")
    conn_str = (
        f"DRIVER={{{cfg['driver']}}};"
        f"SERVER={cfg['server']};"
        f"DATABASE={cfg['database']};"
        f"UID={cfg['username']};"
        f"PWD={cfg['password']};"
        f"TrustServerCertificate=yes;"
        f"Connection Timeout=10;"
    )
    return pyodbc.connect(conn_str)

def run_query(cfg, sql):
    conn = get_connection(cfg)
    cur  = conn.cursor()
    cur.execute(sql)
    cols = [desc[0] for desc in cur.description]
    rows = cur.fetchall()
    conn.close()
    return cols, [list(r) for r in rows]

# ══════════════════════════════════════════════════════════════════════════════
# 연결 설정 다이얼로그
# ══════════════════════════════════════════════════════════════════════════════
class ConnDialog(tk.Toplevel):
    def __init__(self, parent, cfg):
        super().__init__(parent)
        self.title("DB 연결 설정")
        self.resizable(False, False)
        self.grab_set()
        self.result = None

        fields = [
            ("서버 (IP\\인스턴스)", "server"),
            ("데이터베이스",        "database"),
            ("사용자명",            "username"),
            ("비밀번호",            "password"),
            ("ODBC 드라이버",       "driver"),
        ]
        self._vars = {}
        for i, (label, key) in enumerate(fields):
            ttk.Label(self, text=label, anchor="w").grid(row=i, column=0, padx=12, pady=5, sticky="w")
            show = "*" if key == "password" else ""
            var  = tk.StringVar(value=cfg.get(key, ""))
            entry = ttk.Entry(self, textvariable=var, width=36, show=show)
            entry.grid(row=i, column=1, padx=12, pady=5)
            self._vars[key] = var

        btn_frame = ttk.Frame(self)
        btn_frame.grid(row=len(fields), column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text="연결 저장", command=self._ok).pack(side="left", padx=6)
        ttk.Button(btn_frame, text="취소",     command=self.destroy).pack(side="left", padx=6)

        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.wait_window()

    def _ok(self):
        self.result = {k: v.get() for k, v in self._vars.items()}
        self.destroy()

# ══════════════════════════════════════════════════════════════════════════════
# 쿼리 편집 다이얼로그
# ══════════════════════════════════════════════════════════════════════════════
class QueryDialog(tk.Toplevel):
    def __init__(self, parent, tab_key, sql):
        super().__init__(parent)
        self.title(f"쿼리 편집 — {tab_key}")
        self.grab_set()
        self.result = None

        ttk.Label(self, text="SELECT 쿼리 (읽기 전용 권장)").pack(anchor="w", padx=10, pady=(10, 2))
        self._text = tk.Text(self, width=80, height=18, font=("Consolas", 10))
        self._text.insert("1.0", sql)
        self._text.pack(padx=10, pady=4, fill="both", expand=True)

        btn = ttk.Frame(self)
        btn.pack(pady=8)
        ttk.Button(btn, text="저장",  command=self._save).pack(side="left", padx=6)
        ttk.Button(btn, text="취소",  command=self.destroy).pack(side="left", padx=6)
        ttk.Button(btn, text="기본값 복원", command=self._reset).pack(side="left", padx=6)
        self._tab_key = tab_key
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.wait_window()

    def _save(self):
        self.result = self._text.get("1.0", "end").strip()
        self.destroy()

    def _reset(self):
        self._text.delete("1.0", "end")
        self._text.insert("1.0", DEFAULT_QUERIES[self._tab_key]["sql"])

# ══════════════════════════════════════════════════════════════════════════════
# 데이터 탭 패널
# ══════════════════════════════════════════════════════════════════════════════
class DataTab(ttk.Frame):
    def __init__(self, parent, tab_key, app):
        super().__init__(parent)
        self._tab_key = tab_key
        self._app     = app
        self._all_rows = []
        self._cols     = []
        self._build_ui()

    def _build_ui(self):
        # 툴바
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=6, pady=(6, 2))

        ttk.Button(bar, text="⟳ 새로고침", command=self.refresh).pack(side="left")
        ttk.Button(bar, text="쿼리 편집",  command=self._edit_query).pack(side="left", padx=4)
        ttk.Button(bar, text="📥 스크랩 (Excel)", command=self._export_excel,
                   style="Accent.TButton").pack(side="left", padx=4)

        ttk.Label(bar, text="검색:").pack(side="left", padx=(12, 2))
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._filter())
        ttk.Entry(bar, textvariable=self._search_var, width=24).pack(side="left")
        ttk.Button(bar, text="✕", width=2, command=lambda: self._search_var.set("")).pack(side="left")

        self._count_lbl = ttk.Label(bar, text="— 건")
        self._count_lbl.pack(side="right", padx=8)

        # Treeview + 스크롤
        frame = ttk.Frame(self)
        frame.pack(fill="both", expand=True, padx=6, pady=4)

        self._tree = ttk.Treeview(frame, show="headings", selectmode="browse")
        vsb = ttk.Scrollbar(frame, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        hsb.pack(side="bottom", fill="x")
        vsb.pack(side="right",  fill="y")
        self._tree.pack(fill="both", expand=True)

        # 상태바
        self._status = ttk.Label(self, text="준비", anchor="w", foreground="gray")
        self._status.pack(fill="x", padx=8, pady=(0, 4))

    # ── 새로고침 ───────────────────────────────────────────────────────────────
    def refresh(self):
        self._status.config(text="조회 중…", foreground="blue")
        self._count_lbl.config(text="— 건")
        threading.Thread(target=self._fetch, daemon=True).start()

    def _fetch(self):
        queries = load_queries()
        sql = queries.get(self._tab_key, {}).get("sql", "")
        try:
            cols, rows = run_query(self._app.config, sql)
            self.after(0, lambda: self._populate(cols, rows))
        except Exception as e:
            self.after(0, lambda: self._on_error(str(e)))

    def _populate(self, cols, rows):
        self._cols     = cols
        self._all_rows = rows
        self._search_var.set("")
        self._render(cols, rows)
        self._status.config(text=f"마지막 조회: 성공", foreground="green")

    def _render(self, cols, rows):
        self._tree.delete(*self._tree.get_children())
        self._tree["columns"] = cols
        for c in cols:
            self._tree.heading(c, text=c, command=lambda _c=c: self._sort(_c))
            self._tree.column(c, width=120, minwidth=60, anchor="w")
        for row in rows:
            self._tree.insert("", "end", values=row)
        self._count_lbl.config(text=f"{len(rows):,} 건")

    def _on_error(self, msg):
        self._status.config(text=f"오류: {msg[:80]}", foreground="red")
        messagebox.showerror("쿼리 오류", msg, parent=self)

    # ── 검색 필터 ──────────────────────────────────────────────────────────────
    def _filter(self):
        kw = self._search_var.get().lower()
        if not kw:
            filtered = self._all_rows
        else:
            filtered = [r for r in self._all_rows if any(kw in str(v).lower() for v in r)]
        self._render(self._cols, filtered)

    # ── 정렬 ──────────────────────────────────────────────────────────────────
    def _sort(self, col):
        items = [(self._tree.set(k, col), k) for k in self._tree.get_children("")]
        try:
            items.sort(key=lambda x: float(x[0]) if x[0] else 0)
        except ValueError:
            items.sort()
        for idx, (_, k) in enumerate(items):
            self._tree.move(k, "", idx)

    # ── Excel 내보내기 (가격 컬럼 자동 차단) ──────────────────────────────────
    def _export_excel(self):
        if not self._all_rows:
            messagebox.showinfo("데이터 없음", "먼저 새로고침으로 데이터를 조회하세요.", parent=self)
            return

        # 차단 컬럼 분류
        safe_idx   = [i for i, c in enumerate(self._cols) if not is_price_column(c)]
        blocked    = [c for c in self._cols if is_price_column(c)]
        safe_cols  = [self._cols[i] for i in safe_idx]

        # 차단 컬럼 사전 안내
        if blocked:
            msg = (
                f"⚠️ 아래 컬럼은 가격·금액 정보로 판단되어\n내보내기에서 제외됩니다:\n\n"
                + "\n".join(f"  • {c}" for c in blocked)
                + "\n\n계속 진행하시겠습니까?"
            )
            if not messagebox.askyesno("가격 컬럼 차단", msg, parent=self):
                return

        # 저장 경로 선택
        ts        = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default   = f"ERP_{self._tab_key.replace('/', '_')}_{ts}.xlsx"
        filepath  = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=default,
            title="저장 위치 선택",
        )
        if not filepath:
            return

        # openpyxl 로 저장
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror(
                "패키지 없음",
                "openpyxl이 설치되지 않았습니다.\npip install openpyxl 실행 후 재시작하세요.",
                parent=self
            )
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self._tab_key[:31]

        # 헤더 스타일
        hdr_fill = PatternFill("solid", fgColor="2F5496")
        hdr_font = Font(color="FFFFFF", bold=True)
        for ci, col in enumerate(safe_cols, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        # 현재 필터된 rows 기준으로 내보내기
        visible_rows = []
        for item in self._tree.get_children():
            vals = list(self._tree.item(item, "values"))
            visible_rows.append(vals)

        for ri, row in enumerate(visible_rows, start=2):
            for ci, idx in enumerate(safe_idx, start=1):
                val = row[idx] if idx < len(row) else ""
                ws.cell(row=ri, column=ci, value=val)

        # 열 너비 자동 조정
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

        # 메타 시트 (차단 이력)
        meta = wb.create_sheet("_차단컬럼기록")
        meta["A1"] = "내보내기 일시"
        meta["B1"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        meta["A2"] = "탭"
        meta["B2"] = self._tab_key
        meta["A3"] = "차단된 컬럼 (가격/금액)"
        meta["B3"] = ", ".join(blocked) if blocked else "(없음)"
        meta["A4"] = "포함된 컬럼 수"
        meta["B4"] = len(safe_cols)
        meta["A5"] = "내보낸 행 수"
        meta["B5"] = len(visible_rows)

        wb.save(filepath)

        summary = f"✅ 저장 완료\n\n경로: {filepath}\n행: {len(visible_rows):,}건 / 컬럼: {len(safe_cols)}개"
        if blocked:
            summary += f"\n제외된 컬럼: {', '.join(blocked)}"
        messagebox.showinfo("스크랩 완료", summary, parent=self)

    # ── 쿼리 편집 ─────────────────────────────────────────────────────────────
    def _edit_query(self):
        queries = load_queries()
        sql     = queries.get(self._tab_key, {}).get("sql", "")
        dlg     = QueryDialog(self, self._tab_key, sql)
        if dlg.result is not None:
            queries[self._tab_key]["sql"] = dlg.result
            save_queries(queries)
            messagebox.showinfo("저장됨", "쿼리가 저장되었습니다.\n새로고침 버튼으로 다시 조회하세요.", parent=self)

# ══════════════════════════════════════════════════════════════════════════════
# 메인 앱
# ══════════════════════════════════════════════════════════════════════════════
class ERPViewer(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Trico ERP Viewer")
        self.geometry("1100x680")
        self.minsize(800, 500)
        self.config = load_config()
        self._tabs: dict[str, DataTab] = {}
        self._build_ui()
        self._check_config_on_start()

    def _build_ui(self):
        # 메뉴바
        menubar = tk.Menu(self)
        self.config_menu = tk.Menu(menubar, tearoff=0)
        self.config_menu.add_command(label="DB 연결 설정…", command=self._open_conn_dialog)
        self.config_menu.add_separator()
        self.config_menu.add_command(label="종료", command=self.destroy)
        menubar.add_cascade(label="설정", menu=self.config_menu)

        refresh_menu = tk.Menu(menubar, tearoff=0)
        refresh_menu.add_command(label="현재 탭 새로고침", command=self._refresh_current)
        refresh_menu.add_command(label="전체 탭 새로고침", command=self._refresh_all)
        menubar.add_cascade(label="새로고침", menu=refresh_menu)
        self.configure(menu=menubar)

        # 헤더
        hdr = ttk.Frame(self)
        hdr.pack(fill="x", padx=10, pady=(8, 0))
        ttk.Label(hdr, text="Trico ERP Viewer", font=("", 14, "bold")).pack(side="left")
        self._conn_lbl = ttk.Label(hdr, text="⚫ 미연결", foreground="gray")
        self._conn_lbl.pack(side="right")

        ttk.Separator(self, orient="horizontal").pack(fill="x", padx=10, pady=4)

        # 탭 노트북
        self._nb = ttk.Notebook(self)
        self._nb.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        for key in TAB_ORDER:
            tab = DataTab(self._nb, key, self)
            self._nb.add(tab, text=f"  {key}  ")
            self._tabs[key] = tab

    def _check_config_on_start(self):
        if not self.config.get("server"):
            messagebox.showinfo(
                "연결 설정 필요",
                "처음 실행입니다.\nDB 연결 정보를 입력해 주세요.",
                parent=self
            )
            self._open_conn_dialog()
        else:
            self._update_conn_label()

    def _open_conn_dialog(self):
        dlg = ConnDialog(self, self.config)
        if dlg.result:
            self.config = dlg.result
            save_config(self.config)
            self._test_and_update()

    def _test_and_update(self):
        self._conn_lbl.config(text="⟳ 연결 확인 중…", foreground="orange")
        threading.Thread(target=self._test_conn, daemon=True).start()

    def _test_conn(self):
        try:
            conn = get_connection(self.config)
            conn.close()
            self.after(0, lambda: self._conn_lbl.config(
                text=f"🟢 {self.config['server']} / {self.config['database']}",
                foreground="green"
            ))
        except Exception as e:
            self.after(0, lambda: self._conn_lbl.config(text="🔴 연결 실패", foreground="red"))
            self.after(0, lambda: messagebox.showerror("연결 실패", str(e), parent=self))

    def _update_conn_label(self):
        srv = self.config.get("server", "")
        db  = self.config.get("database", "")
        if srv:
            self._conn_lbl.config(text=f"⚫ {srv} / {db}", foreground="gray")

    def _refresh_current(self):
        idx = self._nb.index("current")
        key = TAB_ORDER[idx]
        self._tabs[key].refresh()

    def _refresh_all(self):
        for tab in self._tabs.values():
            tab.refresh()


# ── 진입점 ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = ERPViewer()
    app.mainloop()
