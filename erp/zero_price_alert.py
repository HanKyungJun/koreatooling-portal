"""
단가 0 출하 전표 알림  v1.0
──────────────────────────
매일 오전 9시 실행 (Windows 작업 스케줄러)
- 영업팀 출하등록(lem120_g00)에서 단가 0 전표 조회
- 출하일자 기준 1일 이상 경과한 건만 알림
- Gmail API (OAuth)로 자동 발송 + Excel 첨부
"""

import os, sys, base64, io

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from datetime import date, datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from trico_client import TricoClient

# ── 설정 ──────────────────────────────────────────────────────────────────────
GMAIL_USER  = "hzn2001@toolkorea.co.kr"
LOOK_BACK   = 30   # 최근 N일 조회

_cfg_file = Path(__file__).parent / "config.json"
def _load_cfg():
    try:
        import json
        return json.loads(_cfg_file.read_text(encoding="utf-8"))
    except Exception:
        return {}

TOKEN_FILE  = Path(__file__).parent / "token_gmail.json"
SECRET_FILE = Path(__file__).parent.parent / "client_secret_gmail.json"
SCOPES      = ["https://www.googleapis.com/auth/gmail.send"]
SAVE_DIR    = Path(__file__).parent.parent / "outputs" / "zero_price_alerts"
SAVE_DIR.mkdir(parents=True, exist_ok=True)

DISPLAY_COLS = [
    "mov_no", "out_dt", "sal_dt", "cust_nm", "cust_nm2",
    "itm_nm", "spec", "mov_qty", "bad_qty", "so_no", "out_nm", "memo",
]
COL_LABELS = {
    "mov_no": "출하번호", "out_dt": "출하일", "sal_dt": "판매일",
    "cust_nm": "거래처명", "cust_nm2": "납품처명",
    "itm_nm": "품목명", "spec": "규격",
    "mov_qty": "출하수량", "bad_qty": "불량수량",
    "so_no": "수주번호", "out_nm": "담당자", "memo": "메모",
}


# ══════════════════════════════════════════════════════════════════════════════
def get_gmail_service():
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    if not TOKEN_FILE.exists():
        print("⚠️  Gmail 토큰 없음. 먼저 설정 탭에서 Gmail 인증을 완료하세요.")
        sys.exit(1)

    creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
        TOKEN_FILE.write_text(creds.to_json(), encoding="utf-8")

    return build("gmail", "v1", credentials=creds)


def parse_date(val):
    if not val or str(val).strip() in ("", "None", "NaN"):
        return None
    s = str(val).split("T")[0]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
def fetch_zero_price():
    """단가 0이고 출하일 기준 1일 이상 경과한 전표 반환"""
    today    = date.today()
    fr_dt    = (today - timedelta(days=LOOK_BACK)).strftime("%Y-%m-%d")
    to_dt    = (today - timedelta(days=1)).strftime("%Y-%m-%d")

    print(f"  영업팀 출하 조회 ({fr_dt} ~ {to_dt})...")
    client = TricoClient()
    df = client.출하_영업_명세(fr_dt=fr_dt, to_dt=to_dt)

    if df.empty:
        return pd.DataFrame()

    # 단가 0 필터
    df_zero = df[df["단가0"] == True].copy()
    if df_zero.empty:
        return pd.DataFrame()

    # 출하일 1일 이상 경과 필터
    df_zero["_out_date"] = df_zero["out_dt"].apply(parse_date)
    df_zero = df_zero[
        df_zero["_out_date"].apply(lambda d: d is not None and d < today)
    ].drop(columns=["_out_date", "단가0"])

    # 표시할 컬럼만 추출 (없는 컬럼 제외)
    cols = [c for c in DISPLAY_COLS if c in df_zero.columns]
    return df_zero[cols].reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
def build_excel(df: pd.DataFrame, path: Path):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "단가0 전표"

    HDR_FILL  = PatternFill("solid", fgColor="C62828")
    HDR_FONT  = Font(color="FFFFFF", bold=True)
    ROW_FILL  = PatternFill("solid", fgColor="FFEBEE")
    ROW_FONT  = Font(color="C62828")
    ALIGN_C   = Alignment(horizontal="center", vertical="center")
    THIN      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    today = date.today()

    labels = [COL_LABELS.get(c, c) for c in df.columns]

    # 타이틀
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    tc = ws["A1"]
    tc.value     = f"단가 0 출하 전표  [{today.strftime('%Y-%m-%d')} 기준 / 출하일 1일 이상 경과]"
    tc.font      = Font(size=13, bold=True)
    tc.alignment = ALIGN_C
    ws.row_dimensions[1].height = 24

    # 헤더
    for ci, lbl in enumerate(labels, 1):
        c = ws.cell(2, ci, lbl)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = ALIGN_C; c.border = THIN

    # 데이터
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        for ci, col in enumerate(df.columns, 1):
            val = row[col]
            val = str(val).split("T")[0] if val and str(val) not in ("None","NaN","nan") else ""
            c = ws.cell(ri, ci, val)
            c.fill = ROW_FILL; c.font = ROW_FONT; c.border = THIN

    # 열 너비
    for ci, col in enumerate(df.columns, 1):
        w = max(len(COL_LABELS.get(col, col)), 10) + 2
        ws.column_dimensions[get_column_letter(ci)].width = min(w, 30)

    wb.save(path)


# ══════════════════════════════════════════════════════════════════════════════
def send_email(df: pd.DataFrame, path: Path):
    cfg     = _load_cfg()
    to_mail = cfg.get("to_email", GMAIL_USER)
    service = get_gmail_service()
    today   = date.today()
    cnt     = len(df)

    subject = f"[단가 0 알림] {today.strftime('%Y-%m-%d')} — 미처리 출하 전표 {cnt}건"
    body = f"""안녕하세요,

오늘({today.strftime('%Y년 %m월 %d일')}) 단가 0 출하 전표 알림입니다.

■ 출하일 기준 1일 이상 경과한 단가 0 전표: {cnt}건
■ 상세 내용은 첨부 Excel 파일을 확인해주세요.

단가가 0으로 등록된 전표는 ERP에서 직접 수정이 필요합니다.

본 메일은 자동 발송됩니다.
"""

    msg = MIMEMultipart()
    msg["From"]    = GMAIL_USER
    msg["To"]      = to_mail
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    with open(path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{path.name}"')
    msg.attach(part)

    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    service.users().messages().send(userId="me", body={"raw": raw}).execute()
    print(f"  ✅ 메일 발송 완료 → {to_mail}")


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    today    = date.today()
    filename = f"단가0알림_{today.strftime('%Y%m%d')}.xlsx"
    out_path = SAVE_DIR / filename

    print(f"[{today}] 단가 0 출하 전표 알림 실행 중...")

    print("① ERP 조회")
    df = fetch_zero_price()

    if df.empty:
        print("   단가 0 / 1일 경과 전표 없음 — 발송 생략")
        sys.exit(0)

    print(f"   단가 0 전표: {len(df)}건")

    print("② Excel 생성")
    build_excel(df, out_path)
    print(f"   저장: {out_path}")

    print("③ 이메일 발송")
    send_email(df, out_path)

    print("완료.")
